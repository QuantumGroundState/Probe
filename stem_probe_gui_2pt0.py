#!/usr/bin/env python3
"""
STEM Probe Simulator GUI  —  version 1.8
Physics core adapted from notebook by J. Barthel (ju.barthel@fz-juelich.de)
Ernst Ruska-Centre (ER-C 2), Forschungszentrum Jülich GmbH, 2025

Version history (newest first)
-------------------------------
1.9  W&C geometric theory overlay on d₅₀ vs α plot; plus three physics
     corrections to the temporal coherence model, source-size FWHM convention
     fix, and beam-current recalibration.
     - "Theory overlay" radio (Off / W&C total / W&C components) added to the
       Tab 2 control row.  Selecting W&C total overlays d_geo(α) as a dashed
       orange curve on the d₅₀ axis.  W&C components additionally draws the
       four individual contributions as lighter dotted lines:
           d_diff = 1.22·λ/α          (diffraction disc)
           d_s    = ½·Cs·α³           (spherical aberration disc)
           d_c    = Cc·(σ_ΔE/E)·f_t·α (chromatic disc — same chain as _get_physics)
           d_g    = FWHM_physical / M  (geometric image of the source)
       The overlay rerenders from cached sweep data (no re-sweep needed) and a
       caveat label on the plot notes that W&C is a ray-optics approximation
       that overestimates probe size at high Strehl.
     - New helper method _calc_wc_geometric(alphas_mrad) performs the
       closed-form calculation; disabled components default to zero.

1.9  Three physics corrections to the temporal coherence model, source-size
     FWHM convention fix, and beam-current recalibration.
     - Beam-current formula reverted to the Langmuir / reduced-brightness
       convention:  I = π² · B_r · V₀ · α² · r_s²  with  r_s = FWHM/(2·M).
       Earlier in v1.9 the current formula was (incorrectly) changed to use
       σ_s = FWHM/(2.355·M) to "match" the probe convolution path, but the
       two paths describe different physical roles and the textbook value
       for r_s in the brightness equation is genuinely FWHM/2.  Reverting
       recovers ~28 % of beam current that was erroneously lost.
       The coherence convolution in get_probe still uses σ_s = FWHM/2.355;
       this is intentional and documented in _compute_current_pA.
     - Gun reduced-brightness presets raised ~10–20× to match modern vendor
       data sheets (Hitachi HF5000, JEOL NeoARM, Thermo Fisher Spectra 300):
           Cold FEG     : 1×10⁸ → 2×10⁹
           Schottky FEG : 5×10⁷ → 5×10⁸
           LaB₆         : 1×10⁶ → 1×10⁷
           W hairpin    : 1×10⁴ → 5×10⁴
       Previous values were roughly an order of magnitude low, which caused
       beam currents to be ~10× below typical measured values at matching
       α and voltage.  Combined with the r_s reversion, default-case currents
       are now ~14× higher than before (1.4× from r_s, 10× from B_r).

1.9  Three physics corrections to the temporal coherence model, plus source-size
     FWHM convention fix.
     - Physical source size entry relabelled "Physical Source FWHM (nm)".
       Previously the field was passed directly as σ_s (std-dev) to get_probe
       but divided by 2 (as a diameter) in the beam-current formula — the two
       paths were inconsistent.  Now all paths convert:
           σ_s = FWHM / (2·√(2·ln2) · M)  ≈ FWHM / (2.355 · M)
       Affected: get_probe (via _get_physics), _worker_demag, _compute_current_pA.
       The apparent-size readout now shows both FWHM_apparent and σ_s.
       For the default 3 nm FWHM / M=100, apparent σ_s changes from
       0.0300 nm (old) to 0.0127 nm (new) — a 2.4× reduction in
       source-size broadening.

1.9  Three physics corrections to the temporal coherence model.
     - Issue 1 — Relativistic f_t factor added to focal spread.
       Reimer eq. 6.40 gives H = Cc·(ΔE/E)·f_t where
       f_t = (1 + E/E₀) / (1 + E/2E₀)  (E₀ = m₀c² = 511 keV).
       Previously f_t was omitted; the error grows from ~5 % at 60 keV
       to ~30 % at 300 keV, causing the temporal envelope to decay too
       slowly (probe looked artificially sharper at high voltage).
     - Issue 2 — Energy spread entry now takes FWHM (as measured by EELS
       zero-loss peak width) and converts internally to the Gaussian
       std-dev  σ_ΔE = δE_FWHM / (2√(2 ln 2)) ≈ δE_FWHM / 2.355.
       Previously the field expected a std-dev with no documentation of
       the convention; this caused a 2.355× error if the user pasted a
       FWHM directly from EELS software.  The GUI field is relabelled
       "δE FWHM (eV)", a help note is shown below it, and the live
       readout now shows σ_ΔE alongside σ_f so the chain is auditable.
     - Issue 3 — Defocus quadrature range widened from ±2σ_f (21 points)
       to ±3σ_f (31 points).  The extra 10 FFTs capture the outermost
       4.6 % of the Gaussian tails that were previously truncated, which
       could slightly underestimate damping in the probe tails at large
       focal spread values.

1.8  Export-to-spreadsheet, phase-plate threshold table, and UI polish.
     - "Export to Excel" button writes a multi-sheet .xlsx workbook via openpyxl:
       Parameters, Res_vs_Alpha, Res_vs_Defocus, Demag_Sweep, Current_Sweep,
       Map_d50_pm, Map_Strehl, Map_deff_pm.  Each data sheet has a pale-teal
       summary block (min d₅₀, max Strehl, min d_eff, etc.) above the table.
     - Phase plate radial-average panel: orange dash-dot reference line at
       χ = −π/2 and an inset table of q/α crossings for |χ| = π/8…π.
     - Demagnification tab right y-axis replaced with secondary_yaxis (exact
       M↔I correspondence); left y-axis inverted (largest M at bottom).
     - Δf = 0 reference line removed from Resolution vs Defocus plot.
     - d_eff minimum annotated on the Current sweep plot (star + Δ label).
     - Map Steps N×N combobox extended to 200; d₅₀ vs Demag and d₅₀ vs
       Current scale radio buttons default to Log₁₀ on startup.
     - "Export to Excel" button (left panel) writes a multi-sheet .xlsx workbook
       via openpyxl.  One sheet per program tab: Parameters, Res_vs_Alpha,
       Res_vs_Defocus, Demag_Sweep, Current_Sweep, Map_d50_pm, Map_Strehl.
       The Parameters sheet lists every GUI input (energy, aberrations,
       brightness, coherence, etc.).  Sweep sheets include probe current (pA)
       when Gun Brightness is enabled.  Map sheets are written as 2-D tables
       with α rows and Δf columns.
     - Phase plate radial-average panel gains a dashed reference line at
       χ = −π/2 (orange dash-dot) and a compact inset table listing the
       q (nm⁻¹) and α (mrad) at which |χ| first crosses π/8, π/4, π/2,
       3π/4, and π; values update automatically with aberration changes.
     - Demagnification-tab right y-axis replaced with secondary_yaxis driven
       by I = K/M²: every M tick on the left axis has its exact current value
       at the same height on the right.  Left y-axis inverted (largest M at
       bottom) so M and I both decrease upward, matching physical intuition.

1.7  Phase-plate secondary degree axis and probe-tab layout.
     - Phase radial-average plot gains a secondary right y-axis showing χ in
       degrees, linked to the radian axis via secondary_yaxis.
     - Probe-tab GridSpec rebalanced (height_ratios 2.5 : 1, tighter margins)
       to give more space to the 2-D images without cutting the radial plots.

1.6  d₅₀ vs Current sweep tab and geometric step control for demag/current sweeps.
     - New Tab 4 "d₅₀ vs Current": sweeps demagnification M using geometric steps,
       plots probe size d₅₀ on the y-axis vs probe current I (pA) on the x-axis,
       with a secondary top x-axis showing M.  Reuses _worker_demag for workers.
     - Step multiplier entry replaces fixed "powers of 2": user sets any multiplier
       ≥ 1.001 (default 2.0); generates M_start, M_start·mult, M_start·mult², …
     - Scale radio (Linear / Log₁₀) added to both the Current and Demagnification
       tabs; re-renders live from cached data without re-running the sweep.
     - Top x-axis M labels on the Current tab are placed at every actual data-point
       M value and rotated 90° to prevent overlap.
     - Demagnification tab axes transposed: probe size d₅₀ (pm) now on the x-axis;
       demagnification M on the left y-axis (teal); probe current I on the right
       y-axis (orange, dashed) when Gun Brightness is enabled.  Scale radio now
       applies to both y-axes simultaneously.

1.5  d_eff map display mode and expanded marker options.
     - Third display radio "d_eff" added to the map tab: computes the effective
       resolution d_eff = √(d₅₀² + d_dose²) per cell using beam parameters; I(α)
       is alpha-dependent so d_dose varies across the map.  Falls back to d₅₀ if
       beam parameters are not set.
     - Markers radio expanded from Respective/Both to four options:
         Respective  — star for the active display only (default)
         All         — white (min d₅₀), yellow (max Strehl), cyan (min d_eff)
         None        — no markers
         Difference  — respective star + open diamond at the Scherzer point +
                       dashed connecting line + Δ(Δf) / Δα annotation
     - Map tab label "Colourmap" shortened to "Color".
     - N_probe formula label (N = N_pixel·(d₅₀/dx)²) added to the Source
       Characteristics panel so the dose-resolution scaling is visible on screen.

1.4  Map tab enhancements and dose-resolution physics fix.
     - Strehl ratio computed in parallel inside _worker_map alongside d₅₀ (no
       extra FFT: the aperture and phase plate are already built in the worker).
       A "Display:" radio toggles between d₅₀ map and Strehl map without
       recalculating; Strehl map uses a fixed [0, 1] colour scale with an
       S = 0.8 contour line overlaid.
     - Elapsed time and ETA label added below the map progress bar; updates
       every completed cell using a running-average rate estimate.
     - Scherzer on/off radio: independently toggles the vertical defocus line
       (green) and horizontal aperture line (orange) on the map.
     - Min d₅₀ star annotation now includes the Strehl value at that cell:
       "min d₅₀ = X pm  (S = Y, Δf = Z nm, α = W mrad)".
     - Rose-criterion dose calculation fixed: N_pixel = I·T/e gives electrons
       per pixel; scaled to probe footprint as N_probe = N_pixel·(d₅₀/dx)²
       before applying d_dose = k·d₅₀ / (C·√N_probe).  Previously N_pixel was
       used directly, overestimating d_dose by a factor of (dx/d₅₀).
       Fix applied to both _compute_dose_limited_pm and _update_res_plot.

1.3  d₅₀ contour map gains parallel Strehl ratio computation and display toggle.
     - _worker_map now returns (d50, strehl) instead of d50 alone; Strehl is
       computed from get_phaseplate + calc_strehl with no extra FFT cost since
       the aperture is already built inside the worker.
     - "Display:" radio buttons (d₅₀ / Strehl) added to the map tab control row;
       switching between them re-renders instantly without recalculating.
     - Strehl map uses a fixed [0, 1] colour scale; d₅₀ = 0.8 contour line is
       overlaid on the Strehl map to show the diffraction-limited boundary.
     - _calc_map_thread, _update_map_plot, and _map_last_data updated to carry
       the strehl_grid alongside the d50_grid.
     - Moving average (1-D box filter via mode='valid') extended to Resolution,
       Defocus, and Demagnification sweep tabs; updates live without recalculating.
     - Map edge-cropping fixed: separable 2-D box filter now uses mode='valid'
       and trims defoci / alphas arrays to match, eliminating zero-pad artifacts.

1.2  Rose criterion SNR threshold made user-configurable.
     - "Rose criterion k" entry added to Source Characteristics; default k = 5.
     - d_dose = k · d₅₀ / (C · √N) and d_eff now use the entered k instead of
       the previously hardcoded value of 5.
     - Changing k live updates d_dose and d_eff in the info bar without
       requiring a full probe recalculation.
     - d50 soft-disk blurring multiplier (edge_mult) exposed as a radio button
       (0.1 × dx or 1 × dx) next to the sampling dx field; all four sweep
       workers and the single-probe path respect the selected value.
     - Phase radial-average plot gains a secondary right y-axis showing χ in
       degrees, linked to the radian axis via secondary_yaxis.
     - Probe-tab GridSpec rebalanced (height_ratios 2.5 : 1, tighter margins)
       to reduce white space without overlap.
     - Default grid spacing changed to dx = 0.02 nm/pixel, grid size to N = 1024.
       (v1.9: defaults updated to dx = 0.05 nm/pixel, N = 2048.)

1.1  Aberration file import, polar input mode, and plot improvements.
     - Aberration table toolbar: "Import Aberrations from File…" button and
       "Load Defaults" button (resets all rows to ABERRATION_DEFS values).
     - Two file formats are auto-detected and parsed:
         • ST format  — first line 'STEM'; Haider-style abbreviations (O2, A2 …)
         • Processed  — no header line; Krivanek-style abbreviations (C1, A1 …)
       Units pm / nm / um / mm are all converted to nm automatically.
     - Polar input mode (ρ / θ) added alongside Cartesian (Ax / Ay).  A radio
       button in the table toolbar switches between the two representations.
       Bidirectional traces keep both representations in sync at all times.
       Rotationally symmetric terms (n = 0) have no angle; their θ entry is
       disabled and ρ stores the signed magnitude.  Default mode is ρ / θ.
     - On file import, rows absent from the file are zeroed and unchecked;
       defocus is applied last so the Scherzer auto-update cannot overwrite it.
     - Minimum-probe vertical line added to the α-sweep (Tab 2) and defocus-
       sweep (Tab 3) plots, matching the existing red star annotation.
     - Phase plate plots (2-D image and radial average) now label the spatial-
       frequency axis as q rather than k, and carry a secondary top x-axis
       showing the equivalent aperture angle α in mrad.
     - All Entry widgets in Beam & Optics Parameters (including Source
       Characteristics) stretch to fill the panel width via columnconfigure.
     - Left panel widened from 460 px to 640 px so the aberration table columns
       are fully visible without horizontal scrolling.

1.0  Gun brightness panel and probe-current display.
     - New "Gun Brightness" section in Source Characteristics with radio buttons
       for four common gun types (Cold FEG 1×10⁸, Schottky FEG 5×10⁷,
       LaB₆ 1×10⁶, W hairpin 1×10⁴ A/m²/sr/V) plus a free-entry Custom field.
     - Live probe-current readout using I = B_r·V₀·π²·α²·σ²_apparent; updates
       whenever α, physical size, demagnification, voltage, or B_r changes.
     - Secondary (top) x-axis on the α-sweep plot (Tab 2) showing probe current
       in pA at each α value (current ∝ α²; source size and M are fixed).
     - Secondary (top) x-axis on the demagnification plot (Tab 4) showing probe
       current in pA at each M value (current ∝ 1/M²; α is fixed).

0.9  Probe size vs. demagnification tab (Tab 4).
     - Sweeps the spatial-coherence demagnification factor M from a user-defined
       range (default 10–100) while holding all other parameters fixed.
     - Left y-axis: probe diameter d₅₀ (pm).  Right y-axis: Strehl ratio.
     - Because Strehl is computed from the coherent phase plate only (independent
       of source size), the Strehl curve is a horizontal line whose constant value
       reflects the wavefront quality at the chosen α and aberration set.
     - The plot annotates the demagnification that achieves minimum d₅₀ and
       notes the implied beam-current penalty (I ∝ 1/M²) relative to M = 1.

0.8  Strehl ratio added to the Probe View info bar.
     - New physics function calc_strehl(phi, aperture) returns both the exact
       Strehl ratio S = |⟨exp(iχ)⟩|² and the Maréchal approximation
       S ≈ exp(−σ²_χ), where the averages are aperture-weighted.
     - Both values are displayed in the Tab 1 info bar after each probe
       calculation.  The Strehl value label is colour-coded: green (S ≥ 0.8,
       diffraction-limited), orange (0.5 ≤ S < 0.8), red (S < 0.5).
     - Strehl is computed from the coherent phase plate only; focus-spread
       and source-size envelopes are intentionally excluded, consistent with
       the conventional definition of wavefront Strehl ratio.

0.7  Source Characteristics panel replaces the old flat focus-spread /
     source-size rows.
     - Temporal coherence: user now enters the chromatic aberration
       coefficient Cc (nm) and the energy spread δE (eV).  The focal spread
       is computed as δf = Cc · δE / E₀ (E₀ in eV) and shown live.
     - Spatial coherence: user enters Physical (Virtual) Source Size (nm)
       and a Demagnification factor.  The apparent source size fed into the
       calculation is Physical / Demagnification, shown live.
     - Both sections share a "Source Characteristics" LabelFrame with
       individual enable checkboxes labelled Temporal and Spatial.
     - Default Cs (C₃) updated to 2 700 000 nm (2.7 mm).
     - Default physical source size 3 nm, demagnification 100
       → apparent size 0.03 nm.

0.6  Multi-core parallelisation of the two parameter sweeps.
     - Python's threading module cannot achieve true CPU parallelism due to
       the GIL; sweep loops were moved into a ProcessPoolExecutor so each
       alpha / defocus step runs in a separate OS process with its own GIL.
     - Two module-level worker functions (_worker_alpha, _worker_defocus) are
       required because multiprocessing uses pickle to send work to child
       processes, and only top-level functions are picklable.
     - Progress bars are preserved: futures are submitted individually and
       as_completed() is used so the bar ticks as each process finishes
       rather than waiting for the entire pool to complete.
     - The if __name__ == '__main__' guard (already present) prevents Windows
       spawn-mode workers from re-launching the GUI on import.

0.5  Generalised Scherzer optimum based on Weyland & Muller (2004).
     - New physics functions for arbitrary phase-error tolerance φ:
         Δf  = −(2|φ|/π)^½ · (Cs·λ)^½
         α₀  = (8|φ|·λ / (π·Cs))^¼
         d₀  = 0.61·(π/(8|φ|))^¼ · Cs^¼ · λ^¾
     - "Scherzer Optimal Probe" panel: live φ input (default −π/2) with
       formula labels and computed values for d₀, α₀, and Δf.  Changing φ,
       energy, or Cs auto-updates the convergence-angle and defocus fields.
     - Default Cs (C₃) changed to 2 100 000 nm (2.1 mm); all other
       aberrations (coma, astigmatisms, C5) default to zero.
     - Defocus sweep and probe-tab Scherzer display both use the generalised
       formula so they respect the user-chosen φ.

0.4  Focus-spread and source-size envelopes added.  Both TEM and STEM Scherzer
     defocus values computed and displayed in the probe tab info bar.  Live
     wavelength and dk readouts.  Dual Scherzer lines marked on the defocus
     sweep plot.

0.3  Scrollable left panel; extended aberration table to 6th-order terms
     (C5, A5, R5, S5).  Colormap selector and additional probe display modes
     (contour, 3-D surface, x/y line profiles).

0.2  Added Resolution vs α sweep (Tab 2) and Resolution vs Defocus sweep
     (Tab 3).  Progress bars and threaded calculation to keep GUI responsive.

0.1  Initial release.  Wave-optical probe calculation, phase plate display,
     d50 metric, basic aberration table.
"""

VERSION = "1.9"

import os
import time
import numpy as np
import threading
from concurrent.futures import ProcessPoolExecutor, as_completed
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401 — registers 3-D projection


# ============================================================
# Physics core
# ============================================================

def calc_wavelength(e_kev):
    """Relativistic electron wavelength in nm from beam energy in keV."""
    return 1.2398419843320025 / np.sqrt(e_kev * (e_kev + 1022.0))


def stopping_power_carbon(e_kev):
    """
    Mass electronic stopping power of amorphous carbon (MeV·cm²/g).

    Linearly interpolated from NIST ESTAR tabulated values.
    Accurate to ~2 % for electron energies 50–500 keV (typical STEM range).

    Dose conversion:
        D [J/kg] = F [e/Å²] × S_e [MeV·cm²/g] × 1.602 × 10⁶
    The specimen density ρ cancels (mass stopping power = (1/ρ)·dE/dx).
    """
    _E = np.array([50.,   80.,   100.,  150.,  200.,  300.,  400.,  500.])
    _S = np.array([2.601, 2.141, 1.956, 1.815, 1.781, 1.726, 1.720, 1.730])
    return float(np.interp(e_kev, _E, _S, left=_S[0], right=_S[-1]))


def calc_scherzer_tem(cs_nm, wl_nm):
    """
    Scherzer defocus for TEM phase-contrast imaging (nm).
    Maximises the CTF passband:  Δf_S(TEM) = −√(4/3 · Cs · λ)
    Returns None if Cs ≤ 0.
    """
    if cs_nm <= 0:
        return None
    return -np.sqrt((4.0 / 3.0) * cs_nm * wl_nm)


def calc_scherzer_stem(cs_nm, wl_nm):
    """
    Scherzer defocus for STEM probe optimisation (nm).
    Minimises the probe diameter:  Δf_S(STEM) = −√(Cs · λ)
    Returns None if Cs ≤ 0.
    """
    if cs_nm <= 0:
        return None
    return -np.sqrt(cs_nm * wl_nm)


def calc_scherzer_stem_general(cs_nm, wl_nm, phi_max):
    """
    Generalised STEM Scherzer defocus for an arbitrary maximum phase error φ_max (rad).
      Δf = −(2|φ_max|/π)^(1/2) · (Cs · λ)^(1/2)
    Recovers calc_scherzer_stem when φ_max = −π/2.
    Returns None if Cs ≤ 0 or φ_max = 0.
    """
    if cs_nm <= 0 or phi_max == 0.0:
        return None
    return -np.sqrt(2.0 * abs(phi_max) / np.pi) * np.sqrt(cs_nm * wl_nm)


def calc_optimal_alpha_mrad(cs_nm, wl_nm, phi_max):
    """
    Optimal convergence semi-angle in mrad for an arbitrary phase error φ_max (rad).
      α₀ = (8|φ_max| · λ / (π · Cs))^(1/4)
    """
    if cs_nm <= 0 or phi_max == 0.0:
        return None
    return 1000.0 * (8.0 * abs(phi_max) * wl_nm / (np.pi * cs_nm)) ** 0.25


def calc_optimal_probe_size(cs_nm, wl_nm, phi_max):
    """
    Optimal probe FWHM in nm for an arbitrary phase error φ_max (rad).
      d₀ = 0.61 · (π / (8|φ_max|))^(1/4) · Cs^(1/4) · λ^(3/4)
    """
    if cs_nm <= 0 or phi_max == 0.0:
        return None
    return 0.61 * (np.pi / (8.0 * abs(phi_max))) ** 0.25 * cs_nm ** 0.25 * wl_nm ** 0.75


def get_aperture(radius, edge_width, x, y):
    """
    Soft-edged circular aperture in reciprocal space.

    Uses a tanh roll-off rather than a hard step so that the real-space probe
    does not have Gibbs ringing from a sharp aperture edge.  The transition
    width equals edge_width (typically one reciprocal-pixel dk), giving a
    smooth but spatially tight cutoff.

    Parameters
    ----------
    radius     : float  — aperture radius in nm⁻¹
    edge_width : float  — roll-off half-width in nm⁻¹ (usually dk)
    x, y       : ndarray — reciprocal-space coordinate grids in nm⁻¹

    Returns
    -------
    ndarray, values in [0, 1]
    """
    return 0.5 * (1.0 - np.tanh((np.sqrt(x**2 + y**2) - radius) / edge_width))


def get_phaseplate(aberrations, aperture, kx, ky, wavelength):
    """
    Compute the aberration phase plate χ(k) in radians.

    The phase is evaluated only at pixels inside the aperture (aperture > 0.01)
    to avoid wasting time on zero-weight regions.  The eikonal expansion used
    is the standard polar form:

        χ(k) = Σ_{m,n}  λ^(m-1) · Re[ a_{m,n} · k^p · k̄^q ] / m

    where  a = ax + i·ay,  p = (m+n)/2,  q = (m−n)/2,
    and the (m,n) pairs come from the ABERRATION_DEFS table.

    Parameters
    ----------
    aberrations : dict  {(m,n): (ax, ay)} — enabled aberration coefficients in nm
    aperture    : ndarray — aperture mask from get_aperture
    kx, ky      : ndarray — reciprocal-space grids in nm⁻¹
    wavelength  : float  — electron wavelength in nm

    Returns
    -------
    phi : ndarray — phase in radians, same shape as kx; zero outside aperture
    """
    ap_mask = aperture > 0.01
    act_idx = np.flatnonzero(ap_mask)          # flat indices of active pixels
    kx_act  = kx.flat[act_idx]
    ky_act  = ky.flat[act_idx]
    # Complex reciprocal-space coordinate and its conjugate
    k       = kx_act + 1j * ky_act
    k_cg    = kx_act - 1j * ky_act
    phi_act = np.zeros(len(act_idx), dtype=float)
    for (m, n), (ax, ay) in aberrations.items():
        a = ax + 1j * ay
        p = (m + n) >> 1   # (m+n)/2
        q = (m - n) >> 1   # (m-n)/2
        phi_act += wavelength ** (m - 1) * (a * (k ** p) * (k_cg ** q)).real / m
    phi = np.zeros_like(kx, dtype=float)
    phi.flat[act_idx] = phi_act * 2.0 * np.pi   # convert from waves to radians
    return phi


def get_probe(aberrations, aperture, kx, ky, wavelength,
              focus_spread=None, source_size=None):
    """
    Compute the normalised probe intensity I(r) = |ψ(r)|².

    The probe wave-function is built in reciprocal space:
        ψ(k) = A(k) · exp(−i·χ(k))
    and transformed to real space with an IFFT.

    Temporal coherence (focus spread)
    ----------------------------------
    When focus_spread is given (σ_f in nm), a Gaussian envelope models the
    spread of defocus values caused by chromatic aberration and gun energy
    spread.  focus_spread is the RMS std-dev of the defocus distribution.

    The caller (_get_physics / _update_fs_display) computes σ_f from the
    user-entered FWHM energy spread and the relativistic correction factor f_t
    (Reimer & Kohl, "Transmission Electron Microscopy", 5th ed., eq. 6.40):

        σ_ΔE  = δE_FWHM / (2·√(2·ln 2))        [FWHM → std-dev]
        f_t   = (1 + E/E₀) / (1 + E/2·E₀)      [relativistic; E₀ = 511 keV]
        σ_f   = Cc · σ_ΔE / E₀ · f_t

    The integral over defocus is approximated by 31 quadrature points
    spanning ±3σ_f in steps of σ_f/5, capturing 99.7 % of the Gaussian.
    Each point adds an extra defocus phase  Δφ = 2π · Δf · λ · k² / 2
    and is weighted by the (unnormalised) Gaussian PDF
        w(Δf) = exp(−Δf² / (2·σ_f²))        [std-dev convention]
    The weighted sum is NOT renormalised here; the final division by `total`
    handles normalisation.

    Spatial coherence (source size)
    ---------------------------------
    When source_size is given (σ_s in nm, the RMS / standard deviation of
    the apparent source distribution), the intensity distribution is
    convolved with a Gaussian source profile
        g(r) = (1/(2π·σ_s²)) · exp(−r²/(2·σ_s²))
    whose 2-D continuous Fourier transform (cycles/length convention, as
    used by numpy.fft) is
        S(k) = exp(−2π² · σ_s² · |k|²)      [std-dev convention]
    Convolution is done in reciprocal space by multiplying the FT of the
    intensity by S(k).  References: Goodman, "Introduction to Fourier
    Optics", 4th ed., §2.1 and App. A; Bracewell, "The Fourier Transform
    and its Applications", 3rd ed., Ch. 6 (Gaussian FT pair).

    Convention note: both σ_f and σ_s are standard deviations.  The GUI
    converts the user's FWHM energy spread and applies the relativistic f_t
    factor before passing σ_f in.  For source size: σ = FWHM/2.355,
    or σ = (1/e half-width)/√2 if specified as a 1/e width.

    Parameters
    ----------
    aberrations  : dict  — enabled aberration coefficients {(m,n): (ax, ay)} in nm
    aperture     : ndarray — aperture mask from get_aperture
    kx, ky       : ndarray — reciprocal-space grids in nm⁻¹
    wavelength   : float  — electron wavelength in nm
    focus_spread : float or None — RMS std-dev σ_f of defocus distribution (nm)
    source_size  : float or None — RMS std-dev σ_s of apparent source (nm)

    Returns
    -------
    ndarray — normalised real-space probe intensity, same shape as kx
    """
    ap_mask = aperture > 0.01
    act_idx = np.flatnonzero(ap_mask)
    kx_act  = kx.flat[act_idx]
    ky_act  = ky.flat[act_idx]
    k       = kx_act + 1j * ky_act
    k_cg    = kx_act - 1j * ky_act
    # Accumulate aberration phase at active pixels
    phi_act = np.zeros(len(act_idx), dtype=float)
    for (m, n), (ax, ay) in aberrations.items():
        a = ax + 1j * ay
        p = (m + n) >> 1
        q = (m - n) >> 1
        phi_act += wavelength ** (m - 1) * (a * (k ** p) * (k_cg ** q)).real / m
    phi      = np.zeros_like(kx, dtype=float)
    intens_r = np.zeros_like(kx, dtype=float)
    if focus_spread is not None:
        # Gaussian quadrature over defocus: 31 points from -3·σ_f to +3·σ_f
        # (step = σ_f/5), capturing 99.7 % of the Gaussian weight.
        # Previously ±2σ_f / 21 points captured only 95.4 %; the truncated
        # tails slightly underestimated damping in the far probe wings.
        dfs = focus_spread / 5.0           # step size = σ_f / 5
        k2  = (k * k_cg).real             # |k|² at active pixels
        for ifs in range(-15, 16):         # -3σ to +3σ in steps of σ/5
            df     = dfs * ifs
            # Unnormalised Gaussian PDF weight: w(Δf) = exp(−Δf²/(2·σ_f²)).
            # Std-dev convention; renormalisation handled by `total` below.
            pf     = np.exp(-df ** 2 / (2.0 * focus_spread ** 2))
            # Additional phase due to this defocus offset
            phi_fs = df * wavelength * k2 / 2.0
            phi.flat[act_idx] = (phi_act + phi_fs) * 2.0 * np.pi
            wave_k   = aperture * np.exp(-1j * phi)
            intens_r += pf * np.abs(np.fft.ifft2(wave_k)) ** 2
    else:
        # No temporal envelope: single coherent wave-function
        phi.flat[act_idx] = phi_act * 2.0 * np.pi
        wave_k   = aperture * np.exp(-1j * phi)
        intens_r = np.abs(np.fft.ifft2(wave_k)) ** 2
    if source_size is not None:
        # Convolve intensity with Gaussian source profile in reciprocal space.
        # For a real-space Gaussian g(r) = exp(−r²/(2·σ_s²))/(2π·σ_s²) with
        # RMS std-dev σ_s, its 2-D continuous FT (numpy.fft's cycles/length
        # k convention) is  S(k) = exp(−2π²·σ_s²·|k|²).  See Goodman,
        # "Introduction to Fourier Optics" 4e App. A, and Bracewell,
        # "The Fourier Transform and its Applications" 3e Ch. 6.
        sk       = np.exp(-2.0 * np.pi ** 2 * source_size ** 2 * (kx ** 2 + ky ** 2))
        intens_r = np.fft.ifft2(np.fft.fft2(intens_r) * sk).real
    total = np.sum(intens_r)
    return intens_r / total if total > 0 else intens_r


def get_d50(pdf, dx, edge_mult=1.0):
    """
    Compute the d50 probe diameter in nm via bisection.

    d50 is defined as the diameter of the smallest circle centred on the
    probe that encloses 50 % of the total intensity.  The search uses a
    halving bisection starting from r = 5·dx and converging when the step
    is smaller than 1e-4·dx or the enclosed fraction is within 1e-4 of 0.5.

    The aperture function (tanh roll-off) is reused as a soft-edged disk
    mask so that the enclosed-intensity integral is smooth and the bisection
    converges without oscillation.

    Parameters
    ----------
    pdf       : ndarray — normalised probe intensity from get_probe
    dx        : float  — real-space pixel size in nm
    edge_mult : float  — tanh roll-off width = edge_mult × dx (default 1.0)

    Returns
    -------
    float — d50 diameter in nm  (= 2 × radius at 50 % encircled energy)
    """
    ndim      = pdf.shape
    edge_width = dx * edge_mult
    # Build real-space coordinate grids centred on the probe (FFT ordering)
    ix   = (np.fft.fftfreq(ndim[1]) * ndim[1]).astype(int)
    a_x  = np.tile(ix, (ndim[0], 1)) * dx
    a_y  = a_x.T
    r        = 5.0 * dx    # initial search radius
    r_step   = 0.5 * r     # initial bisection step
    itmax    = 100
    it       = 0
    int_r    = 0.0
    while it < itmax and abs(r_step) > 1e-4 * dx and abs(int_r - 0.5) > 1e-4:
        r_disk = get_aperture(r, edge_width, a_x, a_y)   # soft disk of radius r
        int_r  = np.sum(pdf * r_disk)             # enclosed intensity fraction
        # Reverse step direction when we overshoot 50 %
        if (int_r > 0.5 and r_step > 0.0) or (int_r < 0.5 and r_step < 0.0):
            r_step = -0.5 * r_step
        r  += r_step
        it += 1
    return 2.0 * r   # diameter = 2 × radius


def calc_strehl(phi, aperture):
    """
    Compute the Strehl ratio and phase-error statistics from the aberration
    phase plate.

    All quantities are computed from the coherent phase plate only —
    focus-spread and source-size envelopes are excluded, consistent with the
    conventional definition of wavefront Strehl ratio.

    Exact Strehl
    ------------
    The Strehl ratio is the squared magnitude of the aperture-weighted mean
    of the complex phasor exp(iχ):

        S = |⟨exp(iχ)⟩|²  =  |Σ A(k)·exp(iχ(k)) / Σ A(k)|²

    For a perfect lens χ = 0 everywhere, so every phasor points in the same
    direction and their mean has magnitude 1 → S = 1.  Aberrations rotate
    phasors in different directions; partial cancellation in the sum reduces
    the magnitude and therefore S.

    Maréchal approximation
    ----------------------
    Valid when S > ~0.8 (mild aberrations).  Comes from a second-order Taylor
    expansion of exp(iχ):

        S_M ≈ exp(−σ²_χ)

    where σ²_χ is the aperture-weighted phase variance.  The mean phase is
    subtracted before computing the variance to remove the piston term, which
    does not affect image quality.

    Phase-error statistics
    ----------------------
    Three scalar descriptors of the wavefront error, all piston-corrected:

        σ_χ        (radians)  — RMS wavefront error = sqrt(Var(χ))
        σ_χ / 2π   (waves)    — same quantity expressed as a fraction of λ
        PV = max(χ) − min(χ)  — peak-to-valley range within the aperture

    Parameters
    ----------
    phi      : ndarray — aberration phase χ(k) in radians (from get_phaseplate)
    aperture : ndarray — soft aperture mask A(k) in [0, 1] (from get_aperture)

    Returns
    -------
    s_exact     : float — exact Strehl ratio in [0, 1]
    s_marechal  : float — Maréchal approximation in [0, 1]
    sigma_rad   : float — RMS phase error in radians (piston removed)
    sigma_waves : float — RMS phase error in waves (= sigma_rad / 2π)
    pv_rad      : float — peak-to-valley phase range in radians (piston removed)
    mean_phi_rad: float — aperture-weighted mean phase (piston) in radians
    """
    # Restrict to pixels inside the aperture to avoid diluting the mean with
    # the zero-phase padded region outside (where exp(i·0) = 1 artificially
    # pulls S toward 1).
    ap_mask     = aperture > 0.01
    phi_active  = phi[ap_mask]
    w           = aperture[ap_mask]          # aperture weights for active pixels
    w_sum       = w.sum()

    # --- Exact Strehl: aperture-weighted mean of exp(iχ) ---
    # Split into cos and sin components to keep everything real-valued
    cos_mean = np.dot(w, np.cos(phi_active)) / w_sum
    sin_mean = np.dot(w, np.sin(phi_active)) / w_sum
    s_exact  = float(np.clip(cos_mean**2 + sin_mean**2, 0.0, 1.0))

    # --- Maréchal approximation: exp(−phase variance) ---
    # Remove piston (weighted mean phase) before computing variance so that a
    # uniform phase offset — which shifts the wave but not the intensity — does
    # not reduce the apparent Strehl.
    mean_phi   = np.dot(w, phi_active) / w_sum
    var_phi    = np.dot(w, (phi_active - mean_phi)**2) / w_sum
    s_marechal = float(np.clip(np.exp(-var_phi), 0.0, 1.0))

    # --- Phase-error statistics (piston-corrected) ---
    sigma_rad      = float(np.sqrt(var_phi))          # RMS in radians (piston removed)
    sigma_waves    = sigma_rad / (2.0 * np.pi)        # RMS in waves  (piston removed)
    phi_piston     = phi_active - mean_phi            # piston-corrected phase
    pv_rad         = float(phi_piston.max() - phi_piston.min())   # PV in radians
    mean_phi_rad   = float(mean_phi)                  # piston (aperture-weighted mean)
    # Total RMS including piston: sqrt(⟨χ²⟩) = sqrt(var + ⟨χ⟩²)
    # Decomposes as: sigma_total² = sigma_rad² + mean_phi_rad²
    sigma_total_rad = float(np.sqrt(var_phi + mean_phi**2))

    return s_exact, s_marechal, sigma_rad, sigma_waves, pv_rad, mean_phi_rad, sigma_total_rad


# ============================================================
# Radial-profile helper
# ============================================================

def radial_profile(data_2d, pixel_scale=1.0, max_r=None):
    """
    Compute the azimuthal (rotational) average of a 2D array from its
    geometrical centre pixel.

    Parameters
    ----------
    data_2d     : 2D ndarray — must already be fftshifted so the DC / probe
                  centre sits at pixel (N//2, N//2).
    pixel_scale : float — physical size of one pixel (nm⁻¹ for k-space,
                  nm for real-space).  Multiplied into the returned r-axis.
    max_r       : float or None — maximum radius in physical units to include.
                  Defaults to the largest radius that fits inside the array.

    Returns
    -------
    r_vals   : 1D ndarray — radii in physical units, starting from 0.
    profile  : 1D ndarray — mean data value in each radial bin.
    """
    N  = data_2d.shape[0]
    cy, cx = N // 2, N // 2
    y, x   = np.indices(data_2d.shape)
    r_px   = np.sqrt((x - cx)**2 + (y - cy)**2)

    max_px = int(max_r / pixel_scale) if max_r is not None else int(r_px.max())
    r_int  = np.clip(r_px.astype(int), 0, max_px)
    mask   = r_px <= max_px

    tbin = np.bincount(r_int[mask].ravel(),
                       weights=data_2d[mask].ravel(),
                       minlength=max_px + 1)
    cnt  = np.bincount(r_int[mask].ravel(), minlength=max_px + 1)
    profile = tbin / np.maximum(cnt, 1)
    r_vals  = np.arange(len(profile)) * pixel_scale
    return r_vals, profile


# ============================================================
# Multiprocessing worker functions (must be at module level)
# ============================================================
#
# These functions are called by child processes spawned by ProcessPoolExecutor.
# They MUST be defined at module level (not as class methods or closures)
# because multiprocessing uses pickle to serialise the work unit, and only
# top-level functions are picklable on Windows (spawn start method).
#
# Each worker receives all inputs as a single tuple (required by pool.submit
# when using a plain callable rather than starmap), unpacks it, runs the full
# probe calculation for one parameter point, and returns only the scalar d50.
# Returning the full probe array would waste inter-process pipe bandwidth.

def _worker_alpha(args):
    """
    Worker for the convergence-angle sweep (Tab 2).

    For a single value of the semi-convergence angle α, build the aperture,
    compute the probe, and return d50 plus the exact Strehl ratio.
    The k-grid arrays (akx, aky) are pre-computed on the main process and
    passed in to avoid redundant work inside each child process.

    Strehl is computed from the coherent phase plate only (no envelopes),
    consistent with the conventional wavefront-quality definition.

    Parameters (packed as a single tuple)
    --------------------------------------
    alpha : float  — semi-convergence angle in mrad
    abers : dict   — aberration coefficients {(m,n): (ax,ay)} in nm
    akx, aky : ndarray — reciprocal-space coordinate grids in nm⁻¹
    dk    : float  — reciprocal-space pixel size in nm⁻¹ (used as aperture edge width)
    wl    : float  — electron wavelength in nm
    dx    : float  — real-space pixel size in nm
    fs    : float or None — focal spread δf in nm (None → skip temporal envelope)
    ss    : float or None — apparent source size σ in nm (None → skip spatial envelope)

    Returns
    -------
    (d50, s_exact) : (float, float)
    """
    alpha, abers, akx, aky, dk, wl, dx, fs, ss, edge_mult = args
    k_alpha  = 1e-3 * alpha / wl          # convert mrad → nm⁻¹
    aperture = get_aperture(k_alpha, dk, akx, aky)
    probe    = get_probe(abers, aperture, akx, aky, wl,
                         focus_spread=fs, source_size=ss)
    # Strehl from the coherent phase plate only (envelopes excluded by design)
    phi        = get_phaseplate(abers, aperture, akx, aky, wl)
    s_exact, *_ = calc_strehl(phi, aperture)   # only s_exact needed in sweep
    return get_d50(probe, dx, edge_mult), s_exact


def _worker_defocus(args):
    """
    Worker for the defocus sweep (Tab 3).

    For a single defocus value Δf, override the (2,0) entry in the base
    aberration dict, compute the probe, and return d50 plus the exact Strehl
    ratio.  The aperture is pre-computed on the main process because it is
    the same for all steps (α is fixed during the defocus sweep).

    Strehl is computed from the coherent phase plate only (no envelopes),
    consistent with the conventional wavefront-quality definition.

    Parameters (packed as a single tuple)
    --------------------------------------
    df          : float — defocus Δf in nm
    abers_base  : dict  — all aberrations except (2,0), {(m,n): (ax,ay)} in nm
    aperture    : ndarray — pre-computed aperture mask (fixed α)
    akx, aky    : ndarray — reciprocal-space coordinate grids in nm⁻¹
    wl          : float — electron wavelength in nm
    dx          : float — real-space pixel size in nm
    fs          : float or None — focal spread δf in nm
    ss          : float or None — apparent source size σ in nm

    Returns
    -------
    (d50, s_exact) : (float, float)
    """
    df, abers_base, aperture, akx, aky, wl, dx, fs, ss, edge_mult = args
    # Inject this step's defocus into a fresh copy of the base aberration dict
    abers = dict(abers_base)
    abers[(2, 0)] = (df, 0.0)    # (2,0) = defocus; ay component is always 0
    probe = get_probe(abers, aperture, akx, aky, wl,
                      focus_spread=fs, source_size=ss)
    # Strehl from the coherent phase plate only (envelopes excluded by design)
    phi        = get_phaseplate(abers, aperture, akx, aky, wl)
    s_exact, *_ = calc_strehl(phi, aperture)   # only s_exact needed in sweep
    return get_d50(probe, dx, edge_mult), s_exact


def _worker_demag(args):
    """
    Worker for the demagnification sweep (Tab 4).

    For a single demagnification value M, compute the apparent source size
    σ = physical_size / M, then evaluate the probe and return d50 plus the
    exact Strehl ratio.  The aperture is pre-computed on the main process
    because α is fixed for the entire sweep.

    Note: Strehl is computed from the coherent phase plate only and is therefore
    independent of M — the returned s_exact will be the same for every step.
    The caller plots it as a horizontal reference line showing wavefront quality.

    Parameters (packed as a single tuple)
    --------------------------------------
    demag      : float  — demagnification factor M (dimensionless)
    phys_size  : float  — physical (virtual) source size in nm
    abers      : dict   — aberration coefficients {(m,n): (ax,ay)} in nm
    aperture   : ndarray — pre-computed aperture mask (fixed α)
    akx, aky   : ndarray — reciprocal-space coordinate grids in nm⁻¹
    wl         : float  — electron wavelength in nm
    dx         : float  — real-space pixel size in nm
    fs         : float or None — focal spread δf in nm

    Returns
    -------
    (d50, s_exact) : (float, float)
    """
    demag, phys_size, abers, aperture, akx, aky, wl, dx, fs, edge_mult = args
    # phys_size is the physical source FWHM; convert to apparent σ_s.
    # σ_s = FWHM / (2·√(2·ln2) · M)  ≈ FWHM / (2.355 · M)
    ss    = phys_size / (2.0 * np.sqrt(2.0 * np.log(2.0)) * demag)
    probe = get_probe(abers, aperture, akx, aky, wl, focus_spread=fs, source_size=ss)
    phi        = get_phaseplate(abers, aperture, akx, aky, wl)
    s_exact, *_ = calc_strehl(phi, aperture)   # constant across M; see docstring
    return get_d50(probe, dx, edge_mult), s_exact


def _worker_map(args):
    """
    Worker for one (defocus, alpha) cell of the d50 + Strehl contour map (Tab 5).

    The k-grid is reconstructed locally to avoid pickling large arrays across
    process boundaries (same pattern as _worker_focus_step).

    Strehl is derived from the coherent phase plate alone (independent of focus
    spread and source size), so it accurately reflects wavefront quality at the
    given (df, alpha) operating point regardless of the 'fast' flag.

    Parameters (packed as a single tuple)
    --------------------------------------
    df         : float — defocus in nm
    alpha_mrad : float — convergence semi-angle in mrad
    abers_base : dict  — aberration coefficients {(m,n): (ax, ay)} in nm
                         (defocus entry is overwritten by df for each cell)
    num_k      : int   — grid size N
    dx         : float — real-space pixel size in nm
    wl         : float — electron wavelength in nm
    fs         : float or None — focal spread δf in nm
    ss         : float or None — apparent source size in nm
    edge_mult  : float — soft-disk edge width multiplier for get_d50

    Returns
    -------
    (d50, strehl) : (float, float)
        d50    in nm
        strehl dimensionless [0, 1]
    """
    df, alpha_mrad, abers_base, num_k, dx, wl, fs, ss, edge_mult = args
    ikx      = (np.fft.fftfreq(num_k) * num_k).astype(int)
    _dk      = 1.0 / (dx * num_k)
    akx      = np.tile(ikx, (num_k, 1)) * _dk
    aky      = akx.T
    k_alpha  = 1e-3 * alpha_mrad / wl
    aperture = get_aperture(k_alpha, _dk, akx, aky)
    abers    = dict(abers_base)
    abers[(2, 0)] = (df, 0.0)
    probe    = get_probe(abers, aperture, akx, aky, wl, focus_spread=fs, source_size=ss)
    # Strehl from coherent phase plate — no extra FFT; phi is computed from abers only
    phi            = get_phaseplate(abers, aperture, akx, aky, wl)
    s_exact, *_    = calc_strehl(phi, aperture)
    return get_d50(probe, dx, edge_mult), s_exact


def _worker_focus_step(args):
    """
    Worker for one quadrature step of the focus-spread integral (single probe, Tab 1).

    The focus-spread envelope in get_probe is a sum of 21 coherent probes, each
    computed at a different defocus offset and weighted by a Gaussian.  Each step
    is independent, so they can be distributed across CPU cores.

    Large arrays (aperture, k-grids) are NOT passed as arguments to avoid the
    ~96 MB pickle cost per worker on Windows spawn mode.  Instead, only the
    scalar parameters needed to reconstruct them are sent; the worker rebuilds
    the k-grid and aperture locally (O(N²), negligible compared to the FFT).

    The defocus offset is injected directly into the (2,0) aberration coefficient
    rather than handled separately, which is mathematically equivalent to the
    phi_fs term in get_probe (both produce wl·Δf·|k|²/2 phase).

    Parameters (packed as a single tuple)
    --------------------------------------
    abers_step : dict  — aberrations with this step's defocus already folded into (2,0)
    k_alpha    : float — aperture radius in nm⁻¹
    dk         : float — reciprocal-space pixel size in nm⁻¹ (aperture edge width)
    num_k      : int   — grid size N
    dx         : float — real-space pixel size in nm
    wl         : float — electron wavelength in nm
    weight     : float — Gaussian quadrature weight exp(−Δf²/δf²)

    Returns
    -------
    ndarray — weight × coherent probe intensity for this defocus step
    """
    abers_step, k_alpha, dk, num_k, dx, wl, weight = args
    # Reconstruct k-grid and aperture locally — cheaper than pickling 96 MB
    ikx      = (np.fft.fftfreq(num_k) * num_k).astype(int)
    _dk      = 1.0 / (dx * num_k)
    akx      = np.tile(ikx, (num_k, 1)) * _dk
    aky      = akx.T
    aperture = get_aperture(k_alpha, _dk, akx, aky)
    # Single coherent probe at this defocus — no focus_spread, no source_size
    probe    = get_probe(abers_step, aperture, akx, aky, wl)
    return weight * probe


# ============================================================
# Aberration table — 4 naming conventions per entry
# ============================================================
ABERRATION_DEFS = [
    # (m,n)  Krivanek  Haider/TD    Traditional name                 def_ax   def_ay  init_on
    ((2, 0), 'C\u2081\u2080',  '\u0394f / C\u2081', 'Defocus',                       0.0,      0.0,   True),
    ((2, 2), 'C\u2081\u2082',  'A\u2081',            '2-fold astigmatism',            0.0,      0.1,   True),
    ((3, 1), 'C\u2082\u2081',  'B\u2082',            'Coma',                           0.0,      0.0,   True),
    ((3, 3), 'C\u2082\u2083',  'A\u2082',            '3-fold astigmatism',             0.0,      0.0,   True),
    ((4, 0), 'C\u2083\u2080',  'Cs / C\u2083',       'Spherical aberration (3rd)',  2700000.0,  0.0,   True),
    ((4, 2), 'C\u2083\u2082',  'S\u2083',            'Star aberration',               0.0,      0.0,   False),
    ((4, 4), 'C\u2083\u2084',  'A\u2083',            '4-fold astigmatism',            0.0,      0.0,   False),
    ((5, 1), 'C\u2084\u2081',  'B\u2084',            '4th-order coma',                0.0,      0.0,   False),
    ((5, 3), 'C\u2084\u2083',  'D\u2084',            '3-lobe aberration',             0.0,      0.0,   False),
    ((5, 5), 'C\u2084\u2085',  'A\u2084',            '5-fold astigmatism',            0.0,      0.0,   False),
    ((6, 0), 'C\u2085\u2080',  'C\u2085',            'Spherical aberration (5th)',    0.0,      0.0,   True),
    ((6, 2), 'C\u2085\u2082',  'S\u2085',            'Star aberration (6th order)',   0.0,      0.0,   False),
    ((6, 4), 'C\u2085\u2084',  'R\u2085',            'Rosette aberration',            0.0,      0.0,   False),
    ((6, 6), 'C\u2085\u2086',  'A\u2085',            '6-fold astigmatism',            0.0,      0.0,   True),
]

# Haider/ST-format abbreviation → (m, n)
ST_ABBR_TO_MN = {
    'O2': (2, 0), 'A2': (2, 2),
    'P3': (3, 1), 'A3': (3, 3),
    'O4': (4, 0), 'Q4': (4, 2), 'A4': (4, 4),
    'P5': (5, 1), 'R5': (5, 3), 'A5': (5, 5),
    'O6': (6, 0), 'A6': (6, 6),
}

# Krivanek-style abbreviation → (m, n)  (letter + wave-optical order n = m_eikonal − 1)
KRIVANEK_ABBR_TO_MN = {
    'C1': (2, 0), 'A1': (2, 2),
    'B2': (3, 1), 'A2': (3, 3),
    'C3': (4, 0), 'S3': (4, 2), 'A3': (4, 4),
    'B4': (5, 1), 'D4': (5, 3), 'A4': (5, 5),
    'C5': (6, 0), 'S5': (6, 2), 'R5': (6, 4), 'A5': (6, 6),
}

# Unit multipliers → nm
_ST_UNIT_TO_NM = {'pm': 1e-3, 'nm': 1.0, 'um': 1e3, 'mm': 1e6}


def parse_st_aberrations(filepath):
    """
    Parse an aberrations file and return {(m, n): (ax_nm, ay_nm)}.

    Supports two formats:
      • ST format  — first line is 'STEM'; uses Haider-style abbreviations (O2, A2 …)
      • Processed  — no 'STEM' header; uses Krivanek-style abbreviations (C1, A1 …)

    Units pm / nm / um / mm are all converted to nm.
    Rotationally symmetric terms (N/A angle) get ay = 0.
    Unknown abbreviations are collected in the returned skipped list.
    """
    result  = {}
    skipped = []
    with open(filepath, 'r') as fh:
        raw_lines = fh.readlines()

    lines = [l.strip() for l in raw_lines if l.strip()]

    # Detect format:
    #   • First line 'STEM'       → ST/Haider format with explicit header
    #   • Otherwise, peek at first data abbreviation:
    #       O2 / A2 / P3 … style  → ST/Haider format without STEM header
    #       C1 / A1 / B2 … style  → Krivanek format
    if lines and lines[0].upper() == 'STEM':
        data_lines = lines[3:]          # skip: STEM, header, separator
        abbr_map   = ST_ABBR_TO_MN
    else:
        data_lines = lines[2:]          # skip: header, separator
        first_abbr = data_lines[0].split()[0] if data_lines else ''
        abbr_map   = ST_ABBR_TO_MN if first_abbr in ST_ABBR_TO_MN else KRIVANEK_ABBR_TO_MN

    for line in data_lines:
        parts = line.split()
        if len(parts) < 4:
            continue
        abbr, mag_str, unit, angle_str = parts[0], parts[-3], parts[-2], parts[-1]
        mn = abbr_map.get(abbr)
        if mn is None:
            skipped.append(abbr)
            continue
        scale  = _ST_UNIT_TO_NM.get(unit, 1.0)
        mag_nm = float(mag_str) * scale
        if angle_str == 'N/A':
            ax, ay = mag_nm, 0.0
        else:
            angle_rad = float(angle_str) * (np.pi / 180.0)
            ax = float(mag_nm * np.cos(angle_rad))
            ay = float(mag_nm * np.sin(angle_rad))
        result[mn] = (ax, ay)
    return result, skipped


# Gun reduced-brightness presets (A / m² / sr / V).
# Reduced brightness B_r = B / V₀ is conserved through the column (Liouville).
# Probe current: I = B_r · V₀ · π² · α² · r_s²    (α in rad, r_s in m)
# where r_s = FWHM_apparent / 2 is the Langmuir-convention effective radius.
#
# v1.9 update: values raised ~10–20× to match modern vendor data sheets
# (Hitachi HF5000, JEOL NeoARM, Thermo Fisher Spectra 300).  Previous
# values were roughly an order of magnitude below modern published specs
# and caused probe-current estimates to come out ~10× too low.  These
# remain conservative upper bounds — real currents vary ~2–3× with tip
# age, emission conditions, and monochromator use.
BR_PRESETS = [
    ('Cold FEG',     'cfeg',     2.0e9),   # Hitachi CFEG; up to ~2×10⁹
    ('Schottky FEG', 'schottky', 5.0e8),   # ThermoFisher / JEOL X-FEG
    ('LaB\u2086',    'lab6',     1.0e7),   # typical LaB₆ emitter
    ('W hairpin',    'w',        5.0e4),   # thermionic tungsten
    ('Custom',       'custom',   None),
]

# Colormap options for the probe display
COLORMAP_OPTIONS = [
    ('Inferno',     'inferno'),
    ('Viridis',     'viridis'),
    ('Hot',         'hot'),
    ('Plasma',      'plasma'),
    ('Cube Helix',  'cubehelix'),
]


# ============================================================
# GUI
# ============================================================

class STEMProbeApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"STEM Probe Simulator  —  v{VERSION}")
        self.root.geometry("1480x860")
        self.root.minsize(1100, 680)

        self._last_d50_m    = None   # stored after each probe calc for live k update
        self._last_I_A      = None   # stored beam current (A) for live k update
        self._map_last_data = None   # stored map result for live colormap/MA update
        self._deff_last_grid = None  # d_eff (pm) grid cached separately for export
        self._res_last_data = None   # stored res result for live MA update
        self._defoc_last_data = None # stored defoc result for live MA update
        self._demag_last_data = None # stored demag result for live MA update
        self._calc_running  = False
        self._res_running   = False
        self._defoc_running = False
        self._demag_running   = False
        self._current_running = False
        self._map_running     = False
        self._current_last_data = None
        self._stop_event    = threading.Event()
        self._probe_render_data = None

        self._setup_style()
        self._build_layout()

        # Set default defocus to STEM Scherzer after all widgets exist
        self._set_default_scherzer_defocus()

    # ------------------------------------------------------------------
    # Style
    # ------------------------------------------------------------------

    def _setup_style(self):
        s = ttk.Style()
        try:
            s.theme_use('clam')
        except Exception:
            pass
        s.configure('Title.TLabel',      font=('Segoe UI', 11, 'bold'))
        s.configure('Header.TLabel',     font=('Segoe UI', 9, 'bold'))
        s.configure('Info.TLabel',       font=('Segoe UI', 8),  foreground='#666666')
        s.configure('Warn.TLabel',       font=('Segoe UI', 8),  foreground='#b85c00')
        s.configure('Dk.TLabel',         font=('Segoe UI', 8),  foreground='#444488')
        s.configure('ScherzerTEM.TLabel',font=('Segoe UI', 9, 'bold'), foreground='#8b1a1a')
        s.configure('ScherzerSTEM.TLabel',font=('Segoe UI', 9, 'bold'), foreground='#1a6e1a')
        s.configure('TLabelframe.Label', font=('Segoe UI', 9, 'bold'))
        s.configure('Accent.TButton',    font=('Segoe UI', 10, 'bold'), padding=6)
        s.configure('Stop.TButton',      font=('Segoe UI', 10, 'bold'), padding=6,
                    foreground='#ffffff', background='#8b1a1a')
        # Strehl ratio colour coding: green ≥ 0.8, orange 0.5–0.8, red < 0.5
        s.configure('StrehlGood.TLabel', font=('Segoe UI', 9, 'bold'), foreground='#1a6e1a')
        s.configure('StrehlWarn.TLabel', font=('Segoe UI', 9, 'bold'), foreground='#b85c00')
        s.configure('StrehlBad.TLabel',  font=('Segoe UI', 9, 'bold'), foreground='#8b1a1a')

    # ------------------------------------------------------------------
    # Top-level layout
    # ------------------------------------------------------------------

    def _build_layout(self):
        paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # ---- Left: scrollable controls ----
        left_outer = ttk.Frame(paned, width=640)
        left_outer.pack_propagate(False)
        paned.add(left_outer, weight=0)

        self._left_canvas = tk.Canvas(left_outer, highlightthickness=0, bg='#f0f0f0')
        scrollbar = ttk.Scrollbar(left_outer, orient=tk.VERTICAL,
                                   command=self._left_canvas.yview)
        self._left_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.left_frame = ttk.Frame(self._left_canvas)
        self._left_win  = self._left_canvas.create_window(
            (0, 0), window=self.left_frame, anchor='nw')

        self.left_frame.bind('<Configure>', self._on_left_configure)
        self._left_canvas.bind('<Configure>', self._on_canvas_configure)
        self._left_canvas.bind('<MouseWheel>',
            lambda e: self._left_canvas.yview_scroll(-1 * (e.delta // 120), 'units'))

        # ---- Right: notebook tabs ----
        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=1)

        self.notebook = ttk.Notebook(right_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        self._build_action_buttons()
        self._build_physics_section()
        self._build_aberrations_section()
        self._build_probe_tab()
        self._build_resolution_tab()
        self._build_defocus_tab()
        self._build_demag_tab()
        self._build_current_tab()
        self._build_map_tab()

    def _on_left_configure(self, event):
        self._left_canvas.configure(scrollregion=self._left_canvas.bbox('all'))

    def _on_canvas_configure(self, event):
        self._left_canvas.itemconfig(self._left_win, width=event.width)

    # ------------------------------------------------------------------
    # Left panel — Physics parameters
    # ------------------------------------------------------------------

    def _build_physics_section(self):
        frame = ttk.LabelFrame(self.left_frame, text='Beam & Optics Parameters', padding=8)
        frame.pack(fill=tk.X, padx=8, pady=(10, 4))
        frame.columnconfigure(1, weight=1)

        # Row helper
        def add_row(parent, label_text, var, row_idx, entry_width=None):
            ttk.Label(parent, text=label_text, width=26).grid(
                row=row_idx, column=0, sticky='w', padx=4, pady=2)
            ttk.Entry(parent, textvariable=var).grid(
                row=row_idx, column=1, sticky='ew', padx=4, pady=2)

        # --- Energy ---
        self.var_energy = tk.StringVar(value='300')
        add_row(frame, 'Beam energy (keV):', self.var_energy, 0)

        # --- Wavelength (derived) ---
        ttk.Label(frame, text='Wavelength (derived):').grid(
            row=1, column=0, sticky='w', padx=4, pady=2)
        self.var_wl_display = tk.StringVar(value='—')
        ttk.Label(frame, textvariable=self.var_wl_display,
                  style='Info.TLabel').grid(row=1, column=1, sticky='w', padx=4)
        self.var_energy.trace_add('write', self._update_wl_display)
        self._update_wl_display()

        ttk.Separator(frame, orient=tk.HORIZONTAL).grid(
            row=2, column=0, columnspan=3, sticky='ew', pady=4)

        # --- dx with live dk display and d50 soft-disk blur multiplier ---
        self.var_dx = tk.StringVar(value='0.05')
        ttk.Label(frame, text='Sampling dx (nm/pixel):', width=26).grid(
            row=3, column=0, sticky='w', padx=4, pady=2)
        ttk.Entry(frame, textvariable=self.var_dx).grid(
            row=3, column=1, sticky='ew', padx=4, pady=2)
        self.var_d50_blur = tk.StringVar(value='0.1')
        _blur_frame = ttk.Frame(frame)
        _blur_frame.grid(row=3, column=2, sticky='w', padx=(4, 2))
        ttk.Label(_blur_frame, text='d\u2085\u2080 blur\u00d7dx:',
                  style='Info.TLabel').pack(side=tk.LEFT, padx=(0, 2))
        for _val in ('0.1', '1'):
            ttk.Radiobutton(_blur_frame, text=_val,
                            variable=self.var_d50_blur, value=_val).pack(side=tk.LEFT)

        # dk formula label — shown on row below
        ttk.Label(frame, text='').grid(row=4, column=0)   # spacer
        self.var_dk_display = tk.StringVar(value='dk = 1 / (dx \u00d7 N) = —')
        ttk.Label(frame, textvariable=self.var_dk_display,
                  style='Dk.TLabel').grid(row=4, column=1, sticky='w', padx=4)

        # --- Grid size ---
        ttk.Label(frame, text='Grid size N (pixels):').grid(
            row=5, column=0, sticky='w', padx=4, pady=2)
        self.var_numk = tk.StringVar(value='2048')
        ttk.Combobox(frame, textvariable=self.var_numk, width=8,
                     values=['256', '512', '1024', '2048', '4096'],
                     state='readonly').grid(row=5, column=1, sticky='w', padx=4, pady=2)
        ttk.Label(frame, text='\u26a0 Smaller grids reduce accuracy of d\u2085\u2080 and Strehl.',
                  style='Warn.TLabel').grid(row=6, column=0, columnspan=3,
                                            sticky='w', padx=4, pady=(0, 2))

        # Attach traces for live dk update
        self.var_dx.trace_add('write',   self._update_dk_display)
        self.var_numk.trace_add('write', self._update_dk_display)
        self._update_dk_display()

        ttk.Separator(frame, orient=tk.HORIZONTAL).grid(
            row=7, column=0, columnspan=3, sticky='ew', pady=4)

        # --- Convergence angle ---
        self.var_alpha      = tk.StringVar(value='10.0')
        self.var_alpha_mode = tk.StringVar(value='manual')
        ttk.Label(frame, text='Conv. angle \u03b1 (mrad):', width=26).grid(
            row=8, column=0, sticky='w', padx=4, pady=2)
        self._alpha_entry = ttk.Entry(frame, textvariable=self.var_alpha)
        self._alpha_entry.grid(row=8, column=1, sticky='ew', padx=4, pady=2)
        _alpha_mode_frame = ttk.Frame(frame)
        _alpha_mode_frame.grid(row=8, column=2, sticky='w', padx=(0, 4))
        ttk.Radiobutton(_alpha_mode_frame, text='Manual',
                        variable=self.var_alpha_mode, value='manual',
                        command=self._on_alpha_mode_change).pack(side=tk.LEFT)
        ttk.Radiobutton(_alpha_mode_frame, text='Auto (Scherzer)',
                        variable=self.var_alpha_mode, value='auto',
                        command=self._on_alpha_mode_change).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(frame, text='(single probe calc & defocus sweep)',
                  style='Info.TLabel').grid(row=9, column=1, sticky='w', padx=4)

        ttk.Separator(frame, orient=tk.HORIZONTAL).grid(
            row=10, column=0, columnspan=3, sticky='ew', pady=4)

        # --- Source Characteristics box ---
        src = ttk.LabelFrame(frame, text='Source Characteristics', padding=6)
        src.grid(row=11, column=0, columnspan=3, sticky='ew', padx=4, pady=4)
        src.columnconfigure(1, weight=1)

        # -- Temporal coherence --
        self.var_fs_on = tk.BooleanVar(value=True)
        ttk.Checkbutton(src, variable=self.var_fs_on,
                        text='Temporal').grid(row=0, column=0, sticky='w', padx=4, pady=2)

        ttk.Label(src, text='C\u1d04 (nm):', width=22).grid(
            row=1, column=0, sticky='w', padx=(20, 4), pady=2)
        self.var_cc = tk.StringVar(value='2700000.0')
        ttk.Entry(src, textvariable=self.var_cc).grid(
            row=1, column=1, sticky='ew', padx=4, pady=2)

        ttk.Label(src, text='\u03b4E FWHM (eV):', width=22).grid(
            row=2, column=0, sticky='w', padx=(20, 4), pady=2)
        self.var_dE = tk.StringVar(value='0.3')
        ttk.Entry(src, textvariable=self.var_dE).grid(
            row=2, column=1, sticky='ew', padx=4, pady=2)

        # Help note: FWHM is what EELS software reports for the zero-loss peak
        # width.  The code converts to σ_ΔE = FWHM / 2.355 before computing
        # the focal spread, then applies the relativistic f_t correction.
        ttk.Label(src,
                  text=('\u24d8 Enter ZLP FWHM from EELS.  '
                        'Code converts: \u03c3\u2090E = FWHM \u00f7 2.355, '
                        'then \u03c3_f = C\u1d04\u00b7\u03c3\u2090E\u00b7f\u209c/E\u2080 '
                        '(Reimer eq.\u00a06.40)'),
                  style='Info.TLabel', justify='left', wraplength=320).grid(
            row=3, column=0, columnspan=2, sticky='w', padx=(20, 4), pady=(0, 3))

        # σ_f live readout — shows σ_ΔE and the resulting focal spread σ_f
        ttk.Label(src, text='\u03c3\u2090E / \u03c3_f (Reimer\u00a06.40):',
                  style='Info.TLabel', width=22).grid(
            row=4, column=0, sticky='w', padx=(20, 4), pady=1)
        self.var_fs_display = tk.StringVar(value='\u2014')
        ttk.Label(src, textvariable=self.var_fs_display,
                  style='Info.TLabel').grid(row=4, column=1, sticky='w', padx=4)

        ttk.Separator(src, orient=tk.HORIZONTAL).grid(
            row=5, column=0, columnspan=3, sticky='ew', pady=4)

        # -- Spatial coherence --
        self.var_ss_on = tk.BooleanVar(value=True)
        ttk.Checkbutton(src, variable=self.var_ss_on,
                        text='Spatial').grid(row=6, column=0, sticky='w', padx=4, pady=2)

        ttk.Label(src, text='Physical Source FWHM (nm):', width=22).grid(
            row=7, column=0, sticky='w', padx=(20, 4), pady=2)
        self.var_phys_size = tk.StringVar(value='3')
        ttk.Entry(src, textvariable=self.var_phys_size).grid(
            row=7, column=1, sticky='ew', padx=4, pady=2)

        # Help note: manufacturers and ronchigram measurements report source
        # size as FWHM.  Code converts to σ_s = FWHM / 2.355 before passing
        # to get_probe and to the beam-current formula.
        ttk.Label(src,
                  text=('\u24d8 Enter virtual source size as FWHM (gun spec / ronchigram). '
                        'Code converts: \u03c3_s\u00a0=\u00a0FWHM\u00a0\u00f7\u00a02.355'),
                  style='Info.TLabel', justify='left', wraplength=320).grid(
            row=8, column=0, columnspan=2, sticky='w', padx=(20, 4), pady=(0, 3))

        ttk.Label(src, text='Demagnification:', width=22).grid(
            row=9, column=0, sticky='w', padx=(20, 4), pady=2)
        self.var_demag = tk.StringVar(value='1000')
        ttk.Entry(src, textvariable=self.var_demag).grid(
            row=9, column=1, sticky='ew', padx=4, pady=2)

        ttk.Label(src, text='Apparent \u03c3_s (nm):', style='Info.TLabel', width=22).grid(
            row=10, column=0, sticky='w', padx=(20, 4), pady=1)
        self.var_apparent_size_display = tk.StringVar(value='\u2014')
        ttk.Label(src, textvariable=self.var_apparent_size_display,
                  style='Info.TLabel').grid(row=10, column=1, sticky='w', padx=4)

        ttk.Label(src,
                  text=('\u24d8 Beam current \u221d \u03c3\u00b2_apparent = '
                        '(FWHM\u00f7(2.355\u00b7M))\u00b2\n'
                        '  \u2191M \u2192 \u2191coherence but \u2193 current (fixed brightness)'),
                  style='Info.TLabel', justify='left').grid(
            row=11, column=0, columnspan=3, sticky='w', padx=(20, 4), pady=(1, 2))

        ttk.Separator(src, orient=tk.HORIZONTAL).grid(
            row=12, column=0, columnspan=3, sticky='ew', pady=4)

        # -- Gun brightness --
        self.var_br_on = tk.BooleanVar(value=True)
        ttk.Checkbutton(src, variable=self.var_br_on,
                        text='Gun Brightness').grid(
            row=13, column=0, sticky='w', padx=4, pady=2)

        self.var_br_preset = tk.StringVar(value='cfeg')
        self._br_radio_widgets = []   # kept so we can en/disable them together
        for row_i, (label, key, _) in enumerate(BR_PRESETS[:-1], start=14):
            rb = ttk.Radiobutton(src, variable=self.var_br_preset, value=key,
                                 text=f'{label}',
                                 command=self._update_br_from_preset)
            rb.grid(row=row_i, column=0, sticky='w', padx=(20, 4), pady=1)
            # Show the numeric B_r value next to each preset
            br_val = next(v for l, k, v in BR_PRESETS if k == key)
            ttk.Label(src, text=f'{br_val:.0e} A/m\u00b2/sr/V',
                      style='Info.TLabel').grid(
                row=row_i, column=1, columnspan=2, sticky='w', padx=4)
            self._br_radio_widgets.append(rb)

        # Custom row: radio + entry side by side
        custom_row = 14 + len(BR_PRESETS) - 1   # one past the last preset
        rb_custom = ttk.Radiobutton(src, variable=self.var_br_preset, value='custom',
                                    text='Custom:',
                                    command=self._update_br_from_preset)
        rb_custom.grid(row=custom_row, column=0, sticky='w', padx=(20, 4), pady=1)
        self._br_radio_widgets.append(rb_custom)
        self.var_br_custom = tk.StringVar(value='1e8')
        self.ent_br_custom = ttk.Entry(src, textvariable=self.var_br_custom)
        self.ent_br_custom.grid(row=custom_row, column=1, sticky='ew', padx=4)
        ttk.Label(src, text='A/m\u00b2/sr/V', style='Info.TLabel').grid(
            row=custom_row, column=2, sticky='w', padx=2)
        self.ent_br_custom.configure(state='disabled')   # enabled only for Custom

        # Formula + live current display
        # I = B_r · V₀ · π² · α² · σ²_apparent
        #   B_r        — reduced brightness (A/m²/sr/V); gun-type dependent
        #   V₀         — accelerating voltage (V) = e_kev × 1000
        #   α          — semi-convergence angle (rad)
        #   σ_apparent — apparent source radius (m) = phys_size / (2 × M)
        # This is the current accepted by the probe-forming aperture.
        formula_row = custom_row + 1
        ttk.Label(src,
                  text='I = B\u1d63\u00b7V\u2080\u00b7\u03c0\u00b2\u00b7\u03b1\u00b2\u00b7\u03c3\u00b2_app :',
                  style='Info.TLabel').grid(
            row=formula_row, column=0, sticky='w', padx=(20, 4), pady=(4, 1))
        self.var_current_display = tk.StringVar(value='\u2014')
        ttk.Label(src, textvariable=self.var_current_display,
                  style='Header.TLabel').grid(
            row=formula_row, column=1, columnspan=2, sticky='w', padx=4)

        # Pixel dwell time — used for shot-noise error bars on the α-sweep plot.
        # Shot noise in a Poisson electron beam: N electrons arrive per pixel in
        # dwell time T at current I  →  N = I·T/e  (e = 1.602e-19 C).
        # The fractional uncertainty in the measured source area (and hence in
        # apparent d₅₀) scales as 1/√N, so longer dwell times give tighter bars.
        dwell_row = formula_row + 1
        ttk.Label(src, text='Pixel dwell time:').grid(
            row=dwell_row, column=0, sticky='w', padx=(20, 4), pady=(2, 4))
        self.var_dwell_us = tk.StringVar(value='10')
        ttk.Entry(src, textvariable=self.var_dwell_us).grid(
            row=dwell_row, column=1, sticky='ew', padx=4)
        ttk.Label(src, text='\u03bcs', style='Info.TLabel').grid(
            row=dwell_row, column=2, sticky='w', padx=2)

        # Image contrast C — the fraction of electrons scattered by the feature
        # of interest (0–1).  Used with the user-entered Rose SNR threshold k to
        # compute the dose-limited resolution:
        #   d_dose = k · d₅₀ / (C · √N)
        # where N = I·T/e is the electrons per pixel.  When d_dose >> d₅₀ the
        # measurement is statistics-limited, not optics-limited.  The crossover
        # N_cross = (k/C)² is the minimum electrons per pixel for optics to
        # be the bottleneck.  Typical C: ~0.05–0.1 light atoms, ~0.3 heavier,
        # ~1.0 sharp vacuum edge.
        contrast_row = dwell_row + 1
        ttk.Label(src, text='Image contrast C:').grid(
            row=contrast_row, column=0, sticky='w', padx=(20, 4), pady=(2, 4))
        self.var_contrast = tk.StringVar(value='0.1')
        ttk.Entry(src, textvariable=self.var_contrast).grid(
            row=contrast_row, column=1, sticky='ew', padx=4)
        ttk.Label(src, text='(0\u20131)', style='Info.TLabel').grid(
            row=contrast_row, column=2, sticky='w', padx=2)

        # Rose criterion SNR threshold k  (detection requires SNR = C·√N ≥ k)
        snr_row = contrast_row + 1
        ttk.Label(src, text='Rose criterion k:').grid(
            row=snr_row, column=0, sticky='w', padx=(20, 4), pady=(2, 4))
        self.var_rose_k = tk.StringVar(value='5')
        ttk.Entry(src, textvariable=self.var_rose_k).grid(
            row=snr_row, column=1, sticky='ew', padx=4)
        ttk.Label(src, text='SNR = C\u00b7\u221aN \u2265 k',
                  style='Info.TLabel').grid(
            row=snr_row, column=2, sticky='w', padx=2)

        # Fluence per pixel  F = I·T / (e·dx²)  [e/Å²]
        fluence_row = contrast_row + 2
        ttk.Label(src,
                  text='F = I\u00b7T / (e\u00b7dx\u00b2) :',
                  style='Info.TLabel').grid(
            row=fluence_row, column=0, sticky='w', padx=(20, 4), pady=(6, 1))
        self.var_fluence_display = tk.StringVar(value='\u2014')
        ttk.Label(src, textvariable=self.var_fluence_display,
                  style='Header.TLabel').grid(
            row=fluence_row, column=1, columnspan=2, sticky='w', padx=4)

        # Dose per pixel  D = F·Sₑ(E₀)  [J/kg]  — amorphous carbon, NIST ESTAR
        dose_row = contrast_row + 3
        ttk.Label(src,
                  text='D = F\u00b7S\u2091(E\u2080) :',
                  style='Info.TLabel').grid(
            row=dose_row, column=0, sticky='w', padx=(20, 4), pady=(1, 4))
        self.var_dose_display = tk.StringVar(value='\u2014')
        dose_val_frame = ttk.Frame(src)
        dose_val_frame.grid(row=dose_row, column=1, columnspan=2, sticky='w')
        ttk.Label(dose_val_frame, textvariable=self.var_dose_display,
                  style='Header.TLabel').pack(side=tk.LEFT)
        ttk.Label(dose_val_frame, text='  (amorph. C, NIST)',
                  style='Info.TLabel').pack(side=tk.LEFT)

        # Formula reminder: electrons per probe footprint used in dose resolution.
        # N_pixel = I·T/e counts arrivals per dwell; multiplying by (d₅₀/dx)²
        # scales to the probe area so the Rose criterion applies at the right scale.
        nprobe_row = contrast_row + 4
        ttk.Label(src,
                  text='N\u209a\u2b63\u2092\u2d62\u1d49 = (I\u00b7T/e)\u00b7(d\u2085\u2080/dx)\u00b2 :',
                  style='Info.TLabel').grid(
            row=nprobe_row, column=0, sticky='w', padx=(20, 4), pady=(1, 6))
        ttk.Label(src,
                  text='electrons per probe footprint  (used in d\u209a\u209c\u2099\u2099\u2091 / d\u2091\u2091\u2091)',
                  style='Info.TLabel').grid(
            row=nprobe_row, column=1, columnspan=2, sticky='w', padx=4)

        # Traces: recompute current whenever any relevant parameter changes
        for _v in (self.var_br_on, self.var_br_custom, self.var_ss_on,
                   self.var_phys_size, self.var_demag, self.var_energy, self.var_alpha):
            _v.trace_add('write', self._update_current_display)
        self._update_br_from_preset()   # initialise var_br and current display

        # Traces: recompute fluence & dose when current inputs or pixel size change
        for _v in (self.var_br_on, self.var_br_custom, self.var_ss_on,
                   self.var_phys_size, self.var_demag, self.var_energy, self.var_alpha,
                   self.var_dwell_us, self.var_dx):
            _v.trace_add('write', self._update_fluence_dose_display)
        self._update_fluence_dose_display()

        # Trace: live-update d_dose / d_eff in the info bar when k changes
        self.var_rose_k.trace_add('write', self._refresh_dose_display)

        for _v in (self.var_cc, self.var_dE, self.var_energy):
            _v.trace_add('write', self._update_fs_display)
        self.var_fs_on.trace_add('write', self._update_fs_display)
        self._update_fs_display()

        for _v in (self.var_phys_size, self.var_demag):
            _v.trace_add('write', self._update_apparent_size_display)
        self.var_ss_on.trace_add('write', self._update_apparent_size_display)
        self._update_apparent_size_display()

        ttk.Separator(frame, orient=tk.HORIZONTAL).grid(
            row=12, column=0, columnspan=3, sticky='ew', pady=4)

        # --- Scherzer / Optimal Probe subsection ---
        sch = ttk.LabelFrame(frame, text='Scherzer Optimal Probe', padding=6)
        sch.grid(row=13, column=0, columnspan=3, sticky='ew', padx=4, pady=4)

        # Phase error input
        ttk.Label(sch, text='Phase error \u03c6 (rad):', width=26).grid(
            row=0, column=0, sticky='w', padx=4, pady=2)
        self.var_phi_max = tk.StringVar(value=f'{-np.pi / 2:.6f}')
        ttk.Entry(sch, textvariable=self.var_phi_max, width=10).grid(
            row=0, column=1, sticky='w', padx=4, pady=2)
        ttk.Label(sch, text='(default \u2212\u03c0/2)', style='Info.TLabel').grid(
            row=0, column=2, sticky='w', padx=4)

        # d₀ formula label + live value
        ttk.Label(
            sch,
            text='d\u2080 = 0.61(\u03c0/(8|\u03c6|))\u00bc C\u209b\u00bc \u03bb\u00be :',
            style='Info.TLabel',
        ).grid(row=1, column=0, sticky='w', padx=4, pady=1)
        self.var_d0_display = tk.StringVar(value='\u2014')
        ttk.Label(sch, textvariable=self.var_d0_display,
                  style='ScherzerSTEM.TLabel').grid(
            row=1, column=1, columnspan=2, sticky='w', padx=4)

        # α₀ formula label + live value
        ttk.Label(
            sch,
            text='\u03b1\u2080 = (8|\u03c6|\u03bb/(\u03c0C\u209b))\u00bc :',
            style='Info.TLabel',
        ).grid(row=2, column=0, sticky='w', padx=4, pady=1)
        self.var_alpha0_display = tk.StringVar(value='\u2014')
        ttk.Label(sch, textvariable=self.var_alpha0_display,
                  style='ScherzerSTEM.TLabel').grid(
            row=2, column=1, columnspan=2, sticky='w', padx=4)

        # Δf formula label + live value
        ttk.Label(
            sch,
            text='\u0394f = \u2212(2|\u03c6|/\u03c0)\u00bd(C\u209b\u03bb)\u00bd :',
            style='Info.TLabel',
        ).grid(row=3, column=0, sticky='w', padx=4, pady=1)
        self.var_df0_display = tk.StringVar(value='\u2014')
        ttk.Label(sch, textvariable=self.var_df0_display,
                  style='ScherzerSTEM.TLabel').grid(
            row=3, column=1, columnspan=2, sticky='w', padx=4)

        # Traces: phi_max and energy → refresh display + set simulation fields
        self.var_phi_max.trace_add('write', self._update_scherzer_display)
        self.var_energy.trace_add('write', self._update_scherzer_display)

    def _update_wl_display(self, *_):
        try:
            e_kev = float(self.var_energy.get())
            if e_kev <= 0:
                raise ValueError
            wl = calc_wavelength(e_kev)
            self.var_wl_display.set(f'\u03bb = {wl*1000:.4f} pm  ({wl:.6f} nm)')
        except (ValueError, ZeroDivisionError):
            self.var_wl_display.set('—')

    def _update_dk_display(self, *_):
        try:
            dx = float(self.var_dx.get())
            n  = int(self.var_numk.get())
            if dx <= 0 or n <= 0:
                raise ValueError
            dk = 1.0 / (dx * n)
            self.var_dk_display.set(
                f'dk = 1 / (dx \u00d7 N) = {dk:.5f} nm\u207b\u00b9/pixel')
        except (ValueError, ZeroDivisionError):
            self.var_dk_display.set('dk = 1 / (dx \u00d7 N) = —')

    def _update_fs_display(self, *_):
        """
        Recompute and display σ_ΔE and σ_f from Cc, δE FWHM, and beam energy.

        Full chain (Reimer & Kohl, 5th ed., eq. 6.40):

            σ_ΔE  = δE_FWHM / (2·√(2·ln 2))    — FWHM → std-dev of energy spread
            f_t   = (1 + E/E₀) / (1 + E/2·E₀)  — relativistic correction; E₀=511 keV
            σ_f   = Cc · σ_ΔE / E₀ · f_t        — std-dev of defocus distribution (nm)

        δE_FWHM is the zero-loss peak width from EELS (what the user enters).
        E₀ (denominator) is the beam kinetic energy in eV; E₀ (in f_t) is m₀c²=511 keV.
        The live readout shows both σ_ΔE (eV) and σ_f (nm) so the chain is auditable.

        This value is shown read-only; _get_physics recomputes it independently
        using the same formula to avoid coupling display state to calculation.
        """
        try:
            cc    = float(self.var_cc.get())
            dE    = float(self.var_dE.get())           # δE FWHM in eV
            e_kev = float(self.var_energy.get())
            if e_kev <= 0:
                raise ValueError
            # Step 1: FWHM → std-dev of the Gaussian energy-spread distribution.
            # 2·√(2·ln2) = 2.3548; this is the standard FWHM-to-σ factor.
            sigma_dE = dE / (2.0 * np.sqrt(2.0 * np.log(2.0)))

            # Step 2: relativistic f_t factor (Reimer eq. 6.40).
            # Accounts for the relativistic increase in chromatic sensitivity at
            # higher beam energies.  f_t → 1 at non-relativistic energies.
            E0_keV = 511.0                             # m₀c² in keV
            f_t = (1.0 + e_kev / E0_keV) / (1.0 + e_kev / (2.0 * E0_keV))

            # Step 3: focal-spread std-dev σ_f in nm.
            e0_eV = e_kev * 1000.0                    # beam energy in eV (denominator)
            sigma_f = cc * sigma_dE / e0_eV * f_t

            if self.var_fs_on.get():
                self.var_fs_display.set(
                    f'\u03c3\u2090E={sigma_dE:.3f}\u00a0eV  '
                    f'\u03c3_f={sigma_f:.4f}\u00a0nm  '
                    f'(f\u209c={f_t:.3f})')
            else:
                self.var_fs_display.set('disabled')
        except (ValueError, ZeroDivisionError):
            self.var_fs_display.set('\u2014')

    def _update_apparent_size_display(self, *_):
        """
        Recompute and display the apparent source σ_s from the FWHM input and demag.

        The user enters the physical source FWHM (as quoted by the gun manufacturer
        or measured from a ronchigram).  The chain is:

            FWHM_apparent = FWHM_physical / M
            σ_s           = FWHM_apparent / (2·√(2·ln 2))   ≈ FWHM_apparent / 2.355

        σ_s is what enters the Gaussian convolution in get_probe.  Both values
        are shown so the user can sanity-check against instrument specifications.
        """
        try:
            fwhm  = float(self.var_phys_size.get())
            demag = float(self.var_demag.get())
            if demag <= 0:
                raise ValueError
            # FWHM → σ_s conversion; 2√(2 ln 2) ≈ 2.3548
            fwhm_apparent = fwhm / demag
            sigma_s       = fwhm_apparent / (2.0 * np.sqrt(2.0 * np.log(2.0)))
            if self.var_ss_on.get():
                self.var_apparent_size_display.set(
                    f'FWHM={fwhm_apparent:.4f}\u00a0nm  '
                    f'\u03c3_s={sigma_s:.4f}\u00a0nm')
            else:
                self.var_apparent_size_display.set('disabled')
        except (ValueError, ZeroDivisionError):
            self.var_apparent_size_display.set('\u2014')

    def _compute_current_pA(self, br, e_kev, alpha_mrad, phys_size_nm, demag):
        """
        Probe current in pA.

        Formula: I = B_r · V₀ · π² · α² · r_s²
            B_r  — reduced brightness in A/(m²·sr·V); gun-type dependent.
                   Modern CFEG ~2×10⁹, Schottky ~5×10⁸, LaB₆ ~1×10⁷,
                   W hairpin ~5×10⁴.  These are conservative upper bounds of
                   published vendor specs; real-world values can be ~2–3×
                   lower depending on tip age and emission conditions.
            V₀   — accelerating voltage in V  (= e_kev × 1000)
            α    — semi-convergence angle in rad  (= alpha_mrad × 1e-3)
            r_s  — effective source radius in m, from the Langmuir convention:
                   r_s = FWHM_apparent / 2 = phys_size_nm · 1e-9 / (2·M)
                   where phys_size_nm is the physical source FWHM and M is the
                   demagnification.  This treats the Gaussian source as a
                   uniform disk of radius FWHM/2, which is the convention used
                   by Reimer, Williams & Carter, and vendor data sheets when
                   quoting B_r.

        Note on the two source-size conventions in this code:
          • Brightness / current formula (this function): r_s = FWHM/(2M)
          • Coherence convolution (get_probe):            σ_s = FWHM/(2.355·M)
        These are DIFFERENT conventions serving different physical roles
        (aperture-accepted phase-space area vs. spatial-coherence envelope).
        The factor of ~1.18 between r_s and σ_s is intentional and matches
        the respective reference definitions.

        The factor π²·α²·r_s² is the phase-space area (emittance) accepted by
        the aperture.  Multiplied by B_r·V₀ it gives the current through a
        circular aperture of half-angle α from a source of effective radius r_s.

        Shot-noise consequence: the beam is composed of discrete electrons
        arriving at rate R = I/e (e = 1.602×10⁻¹⁹ C).  In a pixel dwell time
        T the count N = I·T/e is Poisson-distributed with std √N.  The
        fractional uncertainty in I (and hence in the apparent σ²) is 1/√N.
        This is used in _update_res_plot to compute ±1σ error bars on d₅₀.

        Returns float in pA, or None if any parameter is invalid.
        """
        if demag <= 0 or phys_size_nm <= 0 or alpha_mrad <= 0 or e_kev <= 0:
            return None
        v0    = e_kev * 1000.0
        alpha = alpha_mrad * 1e-3
        # Langmuir / reduced-brightness convention: the effective source "radius"
        # in the brightness equation is r_s = FWHM / 2 (half the FWHM of the
        # virtual source), which treats the Gaussian source as a uniform disk
        # of that radius.  This is the standard convention used by Reimer,
        # Williams & Carter, and vendor data sheets when quoting B_r.
        #
        # Note: this is a DIFFERENT source-size convention from the one used in
        # get_probe, where the Gaussian std-dev σ_s = FWHM/2.355 enters the
        # coherence convolution kernel.  The two conventions describe different
        # physical roles (aperture-accepted current vs. spatial coherence
        # envelope), so using different "sizes" is correct, not inconsistent.
        r_s = phys_size_nm * 1e-9 / (2.0 * demag)   # effective radius in metres
        return br * v0 * np.pi**2 * alpha**2 * r_s**2 * 1e12

    def _update_br_from_preset(self, *_):
        """Set var_br from the selected radio button and toggle the custom entry."""
        key = self.var_br_preset.get()
        if key == 'custom':
            self.ent_br_custom.configure(state='normal')
        else:
            self.ent_br_custom.configure(state='disabled')
            br_val = next(v for _, k, v in BR_PRESETS if k == key)
            self.var_br_custom.set(f'{br_val:.0e}')
        self._update_current_display()

    def _update_current_display(self, *_):
        """Recompute and display estimated probe current from current GUI state."""
        if not self.var_br_on.get() or not self.var_ss_on.get():
            self.var_current_display.set('\u2014')
            return
        try:
            br        = float(self.var_br_custom.get())
            e_kev     = float(self.var_energy.get())
            alpha     = float(self.var_alpha.get())
            phys_size = float(self.var_phys_size.get())
            demag     = float(self.var_demag.get())
            i_pA = self._compute_current_pA(br, e_kev, alpha, phys_size, demag)
            if i_pA is None:
                self.var_current_display.set('\u2014')
            else:
                self.var_current_display.set(f'{i_pA:.3g} pA')
        except (ValueError, ZeroDivisionError):
            self.var_current_display.set('\u2014')

    def _update_fluence_dose_display(self, *_):
        """
        Compute and display fluence (e/Å²) and radiation dose (J/kg) per pixel.

        Fluence:  F = I·T / (e·dx²)
            I  — probe current (A), from brightness formula
            T  — pixel dwell time (s)
            e  — electron charge (1.602×10⁻¹⁹ C)
            dx — pixel size (Å)

        Dose:  D = F · Sₑ(E₀)
            Sₑ — mass electronic stopping power (MeV·cm²/g), NIST ESTAR, amorphous C
            D  — in J/kg (= Gy);  specimen density ρ cancels in the mass stopping power.
        """
        dash = '\u2014'
        if not self.var_br_on.get() or not self.var_ss_on.get():
            self.var_fluence_display.set(dash)
            self.var_dose_display.set(dash)
            return
        try:
            br        = float(self.var_br_custom.get())
            e_kev     = float(self.var_energy.get())
            alpha     = float(self.var_alpha.get())
            phys_size = float(self.var_phys_size.get())
            demag     = float(self.var_demag.get())
            dwell_us  = float(self.var_dwell_us.get())
            dx_nm     = float(self.var_dx.get())

            i_pA = self._compute_current_pA(br, e_kev, alpha, phys_size, demag)
            if i_pA is None:
                raise ValueError('no current')

            I_A      = i_pA * 1e-12           # pA → A
            T_s      = dwell_us * 1e-6         # µs → s
            e_C      = 1.602e-19               # C
            dx_A     = dx_nm * 10.0            # nm → Å

            fluence  = I_A * T_s / (e_C * dx_A ** 2)   # e/Å²

            S_e      = stopping_power_carbon(e_kev)     # MeV·cm²/g
            dose_Gy  = fluence * S_e * 1.602e6          # J/kg

            self.var_fluence_display.set(f'{fluence:.3g} e/\u212b\u00b2')

            if dose_Gy >= 1e6:
                self.var_dose_display.set(f'{dose_Gy / 1e6:.3g} MGy')
            elif dose_Gy >= 1e3:
                self.var_dose_display.set(f'{dose_Gy / 1e3:.3g} kGy')
            else:
                self.var_dose_display.set(f'{dose_Gy:.3g} Gy')
        except (ValueError, ZeroDivisionError):
            self.var_fluence_display.set(dash)
            self.var_dose_display.set(dash)

    def _update_scherzer_display(self, *_):
        """
        Recompute Scherzer optimal d0, alpha0, Df from current phi, energy, and Cs.
        Also propagates alpha0 to var_alpha and Df to the defocus aberration field.
        """
        dash = '\u2014'
        try:
            phi = float(self.var_phi_max.get())
            wl  = calc_wavelength(float(self.var_energy.get()))
            cs  = self._get_cs_nm()
            if cs is None or cs <= 0 or phi == 0.0:
                raise ValueError('invalid params')

            d0     = calc_optimal_probe_size(cs, wl, phi)
            alpha0 = calc_optimal_alpha_mrad(cs, wl, phi)
            df0    = calc_scherzer_stem_general(cs, wl, phi)

            self.var_d0_display.set(
                f'{d0 * 1000:.2f} pm  ({d0:.5f} nm)')
            self.var_alpha0_display.set(f'{alpha0:.3f} mrad')
            self.var_df0_display.set(
                f'{df0:.4f} nm  ({df0 * 1000:.1f} pm)')

            # Push values into simulation fields
            if getattr(self, 'var_alpha_mode', None) and self.var_alpha_mode.get() == 'auto':
                self.var_alpha.set(f'{alpha0:.3f}')
            if hasattr(self, 'aber_rows') and (2, 0) in self.aber_rows:
                _, ax_var, _ = self.aber_rows[(2, 0)]
                ax_var.set(f'{df0:.4f}')
        except Exception:
            self.var_d0_display.set(dash)
            self.var_alpha0_display.set(dash)
            self.var_df0_display.set(dash)

    # ------------------------------------------------------------------
    # Left panel — Aberrations table
    # ------------------------------------------------------------------

    def _build_aberrations_section(self):
        outer = ttk.LabelFrame(self.left_frame,
                               text='Aberration Coefficients', padding=8)
        outer.pack(fill=tk.X, padx=8, pady=4)

        # ── Top toolbar: import button + input-mode radio buttons ────────────
        toolbar = ttk.Frame(outer)
        toolbar.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(toolbar, text='Import Aberrations from File…',
                   command=self._on_import_aberrations).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(toolbar, text='Load Defaults',
                   command=self._on_load_defaults).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Label(toolbar, text='Input mode:').pack(side=tk.LEFT)
        self._aber_mode_var = tk.StringVar(value='polar')
        ttk.Radiobutton(toolbar, text='Ax / Ay',
                        variable=self._aber_mode_var, value='xy',
                        command=self._on_aber_mode_change).pack(side=tk.LEFT, padx=4)
        ttk.Radiobutton(toolbar, text='\u03c1 / \u03b8',
                        variable=self._aber_mode_var, value='polar',
                        command=self._on_aber_mode_change).pack(side=tk.LEFT, padx=4)

        tbl = ttk.Frame(outer)
        tbl.pack(fill=tk.X)

        col_cfg = [
            ('On',              4,  'center'),
            ('\u2460 (m,n)',    6,  'center'),
            ('\u2461 Krivanek', 6,  'center'),
            ('\u2462 Haider',   8,  'center'),
            ('\u2463 Traditional name', 20, 'w'),
            ('Ax [nm]',        10,  'center'),
            ('Ay [nm]',        10,  'center'),
        ]
        hdr_labels = []
        for col, (text, width, anchor) in enumerate(col_cfg):
            lbl = ttk.Label(tbl, text=text, style='Header.TLabel',
                            width=width, anchor=anchor)
            lbl.grid(row=0, column=col, padx=2, pady=2, sticky='ew')
            hdr_labels.append(lbl)
        # Keep refs to the two mutable header labels
        self._aber_hdr_col5 = hdr_labels[5]
        self._aber_hdr_col6 = hdr_labels[6]

        ttk.Separator(tbl, orient=tk.HORIZONTAL).grid(
            row=1, column=0, columnspan=len(col_cfg), sticky='ew', pady=2)

        self.aber_rows        = {}
        self._aber_polar_vars = {}   # key → (rho_v, theta_v, ent5, ent6, is_sym)
        self._aber_updating   = False

        def _make_xy_to_polar(ax_v, ay_v, rho_v, theta_v, sym):
            def _update(*_):
                if self._aber_updating:
                    return
                self._aber_updating = True
                try:
                    ax = float(ax_v.get())
                    ay = float(ay_v.get())
                    # Symmetric terms have no angle, so preserve the sign of ax as ρ
                    rho_v.set(f'{ax:.6g}' if sym else f'{float(np.hypot(ax, ay)):.6g}')
                    if not sym:
                        theta_v.set(f'{float(np.degrees(np.arctan2(ay, ax))):.2f}')
                except ValueError:
                    pass
                finally:
                    self._aber_updating = False
            return _update

        def _make_polar_to_xy(rho_v, theta_v, ax_v, ay_v, sym):
            def _update(*_):
                if self._aber_updating:
                    return
                self._aber_updating = True
                try:
                    rho = float(rho_v.get())
                    if sym:
                        ax_v.set(f'{rho:.6g}')
                    else:
                        theta_rad = float(theta_v.get()) * (np.pi / 180.0)
                        ax_v.set(f'{float(rho * np.cos(theta_rad)):.6g}')
                        ay_v.set(f'{float(rho * np.sin(theta_rad)):.6g}')
                except ValueError:
                    pass
                finally:
                    self._aber_updating = False
            return _update

        for r_idx, defn in enumerate(ABERRATION_DEFS, start=2):
            (m, n), krivanek, haider, trad, def_ax, def_ay, init_on = defn
            key  = (m, n)
            sym  = (n == 0)

            en_var  = tk.BooleanVar(value=init_on)
            ax_var  = tk.StringVar(value=str(def_ax) if init_on else '0.0')
            ay_var  = tk.StringVar(value=str(def_ay) if init_on else '0.0')

            init_rho   = float(np.hypot(def_ax, def_ay)) if init_on else 0.0
            init_theta = (float(np.degrees(np.arctan2(def_ay, def_ax)))
                          if (init_on and not sym) else 0.0)
            rho_var   = tk.StringVar(value=f'{init_rho:.6g}')
            theta_var = tk.StringVar(value='N/A' if sym else f'{init_theta:.2f}')

            # xy → polar (fires when user edits Ax/Ay fields)
            cb_xy2p = _make_xy_to_polar(ax_var, ay_var, rho_var, theta_var, sym)
            ax_var.trace_add('write', cb_xy2p)
            ay_var.trace_add('write', cb_xy2p)

            # polar → xy (fires when user edits ρ/θ fields)
            cb_p2xy = _make_polar_to_xy(rho_var, theta_var, ax_var, ay_var, sym)
            rho_var.trace_add('write', cb_p2xy)
            if not sym:
                theta_var.trace_add('write', cb_p2xy)

            ttk.Checkbutton(tbl, variable=en_var).grid(
                row=r_idx, column=0, padx=2, pady=1)
            ttk.Label(tbl, text=f'({m},{n})', anchor='center').grid(
                row=r_idx, column=1, padx=2)
            ttk.Label(tbl, text=krivanek, anchor='center').grid(
                row=r_idx, column=2, padx=2)
            ttk.Label(tbl, text=haider, anchor='center').grid(
                row=r_idx, column=3, padx=2)
            ttk.Label(tbl, text=trad, anchor='w').grid(
                row=r_idx, column=4, padx=4, sticky='w')
            ent5 = ttk.Entry(tbl, textvariable=ax_var, width=10)
            ent5.grid(row=r_idx, column=5, padx=2)
            ent6 = ttk.Entry(tbl, textvariable=ay_var, width=10)
            ent6.grid(row=r_idx, column=6, padx=2)

            self.aber_rows[key]        = (en_var, ax_var, ay_var)
            self._aber_polar_vars[key] = (rho_var, theta_var, ent5, ent6, sym)

        # Refresh Scherzer display when Cs changes
        if (4, 0) in self.aber_rows:
            _, cs_ax, _ = self.aber_rows[(4, 0)]
            cs_ax.trace_add('write', self._update_scherzer_display)

        # Apply default mode (polar) so headers and entry bindings are correct on startup
        self._on_aber_mode_change()

        # Naming conventions legend — placed at the bottom so the table is
        # immediately visible when the panel is scrolled to the aberrations section
        ttk.Separator(outer, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=6)
        legend = ttk.LabelFrame(outer, text='Naming Conventions', padding=6)
        legend.pack(fill=tk.X, pady=(0, 4))

        conventions = [
            ('\u2460 Eikonal (m,n)',
             'Integer index pair used in this code.  m = aberration order,\n'
             'n = rotational symmetry.  m+n must be even, n \u2264 m.\n'
             'Example: (4,0) = 4th-order, rotationally symmetric = spherical aberration.'),
            ('\u2461 Krivanek notation',
             'C\u2099\u2098 where n = m\u208b\u2081 = wave-optical order, m = azimuthal symmetry.\n'
             'Two orthogonal components are labeled a/b (or x/y).\n'
             'Example: C\u2083\u2080 = 3rd-order spherical (= Cs),  C\u2081\u2082 = 2-fold astigmatism.'),
            ('\u2462 Typke\u2013Dierksen / Haider notation',
             'Used by CEOS correctors.  Letter prefix encodes symmetry type:\n'
             '  A = astigmatism,  B = coma,  C = spherical,  D = 3-lobe,\n'
             '  S = star,  R = rosette.\n'
             'Subscript = radial order (1-based).  Example: B\u2082 = coma, Cs = C\u2083.'),
            ('\u2463 Traditional / descriptive name',
             'Plain-language name widely used in the electron microscopy\n'
             'community (e.g. defocus, coma, spherical aberration).'),
        ]
        for title, body in conventions:
            row_f = ttk.Frame(legend)
            row_f.pack(fill=tk.X, pady=2)
            ttk.Label(row_f, text=title, style='Header.TLabel').pack(anchor='w')
            ttk.Label(row_f, text=body, style='Info.TLabel',
                      wraplength=420, justify=tk.LEFT).pack(anchor='w', padx=12)

    # ------------------------------------------------------------------
    # Left panel — Action buttons & status
    # ------------------------------------------------------------------

    def _on_alpha_mode_change(self, *_):
        mode = self.var_alpha_mode.get()
        self._alpha_entry.config(state='normal' if mode == 'manual' else 'disabled')
        if mode == 'auto':
            self._update_scherzer_display()

    def _on_load_defaults(self):
        for defn in ABERRATION_DEFS:
            (m, n), _, _, _, def_ax, def_ay, init_on = defn
            key = (m, n)
            if key not in self.aber_rows:
                continue
            en_v, ax_v, ay_v = self.aber_rows[key]
            ax_v.set(str(def_ax))
            ay_v.set(str(def_ay))
            en_v.set(init_on)

    def _on_aber_mode_change(self, *_):
        mode = self._aber_mode_var.get()
        if mode == 'xy':
            self._aber_hdr_col5.config(text='Ax [nm]')
            self._aber_hdr_col6.config(text='Ay [nm]')
            for key, (rho_v, theta_v, ent5, ent6, sym) in self._aber_polar_vars.items():
                en_v, ax_v, ay_v = self.aber_rows[key]
                ent5.config(textvariable=ax_v, state='normal')
                ent6.config(textvariable=ay_v, state='normal')
        else:  # polar
            self._aber_hdr_col5.config(text='\u03c1 [nm]')
            self._aber_hdr_col6.config(text='\u03b8 [\u00b0]')
            for key, (rho_v, theta_v, ent5, ent6, sym) in self._aber_polar_vars.items():
                ent5.config(textvariable=rho_v, state='normal')
                ent6.config(textvariable=theta_v,
                            state='disabled' if sym else 'normal')

    def _on_import_aberrations(self):
        path = filedialog.askopenfilename(
            title='Import aberrations',
            filetypes=[('Text files', '*.txt'), ('All files', '*.*')],
        )
        if not path:
            return
        try:
            data, skipped = parse_st_aberrations(path)
        except Exception as exc:
            messagebox.showerror('Import failed', str(exc))
            return
        # Zero out all rows not present in the file before applying new values
        for mn_key, (en_v, ax_v, ay_v) in self.aber_rows.items():
            if mn_key not in data:
                ax_v.set('0.0')
                ay_v.set('0.0')
                en_v.set(False)
        # Set all terms except defocus first — importing Cs triggers _update_scherzer_display
        # which overwrites (2,0); re-applying defocus last keeps the file's value.
        for mn, (ax, ay) in data.items():
            if mn == (2, 0) or mn not in self.aber_rows:
                continue
            en_var, ax_var, ay_var = self.aber_rows[mn]
            ax_var.set(f'{ax:.6g}')
            ay_var.set(f'{ay:.6g}')
            en_var.set(True)
        if (2, 0) in data and (2, 0) in self.aber_rows:
            en_var, ax_var, ay_var = self.aber_rows[(2, 0)]
            ax, ay = data[(2, 0)]
            ax_var.set(f'{ax:.6g}')
            ay_var.set(f'{ay:.6g}')
            en_var.set(True)
        # Switch to ρ/θ mode so the table shows the same representation as the file
        self._aber_mode_var.set('polar')
        self._on_aber_mode_change()
        msg = f'Loaded {len(data)} aberration term(s).'
        if skipped:
            msg += f'\nUnrecognised abbreviations skipped: {", ".join(skipped)}'
        messagebox.showinfo('Import complete', msg)

    def _build_action_buttons(self):
        frame = ttk.Frame(self.left_frame)
        frame.pack(fill=tk.X, padx=8, pady=8)

        self.btn_probe = ttk.Button(
            frame, text='  Calculate Probe  ',
            style='Accent.TButton', command=self._on_calc_probe)
        self.btn_probe.pack(fill=tk.X, pady=3)

        self.btn_res = ttk.Button(
            frame, text='  Calculate Resolution vs \u03b1  ',
            style='Accent.TButton', command=self._on_calc_resolution)
        self.btn_res.pack(fill=tk.X, pady=3)

        self.btn_defoc = ttk.Button(
            frame, text='  Calculate Resolution vs Defocus  ',
            style='Accent.TButton', command=self._on_calc_defocus)
        self.btn_defoc.pack(fill=tk.X, pady=3)

        self.btn_map = ttk.Button(
            frame, text='  Calculate d\u2085\u2080 Contour Map  ',
            style='Accent.TButton', command=self._on_calc_map)
        self.btn_map.pack(fill=tk.X, pady=3)

        self.btn_stop = ttk.Button(
            frame, text='  ■  Stop Calculation  ',
            style='Stop.TButton', command=self._on_stop)
        self.btn_stop.pack(fill=tk.X, pady=3)
        self.btn_stop.configure(state='disabled')

        ttk.Separator(frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)
        self.btn_export = ttk.Button(
            frame, text='  \U0001f4be  Export All Data to Excel (.xlsx)  ',
            command=self._export_to_spreadsheet)
        self.btn_export.pack(fill=tk.X, pady=3)

        self.progress = ttk.Progressbar(frame, mode='indeterminate', length=200)
        self.progress.pack(fill=tk.X, pady=4)

        self.var_status = tk.StringVar(value='Ready.')
        ttk.Label(frame, textvariable=self.var_status,
                  style='Info.TLabel', wraplength=420,
                  justify=tk.LEFT).pack(fill=tk.X)

    # ------------------------------------------------------------------
    # Right panel — Tab 1: Probe viewer
    # ------------------------------------------------------------------

    def _build_probe_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=' Probe View ')

        # ---- Info bar: d50 + both Scherzer values ----
        info_bar = ttk.Frame(tab)
        info_bar.pack(fill=tk.X, padx=6, pady=(2, 0))

        # Helpers to keep line length down
        def _iblbl(text, style='Header.TLabel'):
            ttk.Label(info_bar, text=text, style=style).pack(side=tk.LEFT)
        def _ibval(var, style='Info.TLabel', gap=8):
            ttk.Label(info_bar, textvariable=var, style=style).pack(
                side=tk.LEFT, padx=(2, gap))

        _iblbl('d\u2085\u2080:')
        self.var_d50 = tk.StringVar(value='—')
        ttk.Label(info_bar, textvariable=self.var_d50,
                  style='Title.TLabel').pack(side=tk.LEFT, padx=(2, 10))

        # d_dose: dose-limited resolution from Rose criterion (SNR ≥ k).
        # d_dose = k·d₅₀/(C·√N)  — shown only when brightness+dwell+contrast valid.
        # d_eff: combined optics+dose resolution in quadrature: √(d₅₀²+d_dose²).
        # Both diverge as I→0 (N→0), correctly reflecting that zero current
        # means no resolvable information regardless of optic quality.
        _iblbl('d\u209a\u209c\u2099\u2099\u2091:')   # d_dose label
        self.var_d_dose_display = tk.StringVar(value='—')
        _ibval(self.var_d_dose_display, gap=10)

        _iblbl('d\u2091\u2091\u2091:')   # d_eff label
        self.var_d_eff_display = tk.StringVar(value='—')
        _ibval(self.var_d_eff_display, gap=10)

        _iblbl('\u0394f\u209b(TEM):')
        self.var_scherzer_tem_display = tk.StringVar(value='—')
        _ibval(self.var_scherzer_tem_display, 'ScherzerTEM.TLabel', gap=10)

        _iblbl('\u0394f\u209b(STEM):')
        self.var_scherzer_stem_display = tk.StringVar(value='—')
        _ibval(self.var_scherzer_stem_display, 'ScherzerSTEM.TLabel', gap=10)

        _iblbl('Strehl:')
        self.var_strehl = tk.StringVar(value='—')
        self.lbl_strehl = ttk.Label(info_bar, textvariable=self.var_strehl,
                                    style='StrehlGood.TLabel')
        self.lbl_strehl.pack(side=tk.LEFT, padx=(2, 10))

        # Phase-error statistics: RMS (rad), RMS (waves), PV (rad), piston (rad)
        _iblbl('RMS\u03c7:')
        self.var_sigma_rad = tk.StringVar(value='—')
        _ibval(self.var_sigma_rad, gap=0)
        ttk.Label(info_bar, text='rad /').pack(side=tk.LEFT, padx=(2, 0))
        self.var_sigma_waves = tk.StringVar(value='—')
        _ibval(self.var_sigma_waves, gap=0)
        ttk.Label(info_bar, text='\u03bb').pack(side=tk.LEFT, padx=(2, 8))

        _iblbl('PV\u03c7:')
        self.var_pv_rad = tk.StringVar(value='—')
        _ibval(self.var_pv_rad, gap=0)
        ttk.Label(info_bar, text='rad').pack(side=tk.LEFT, padx=(2, 8))

        _iblbl('\u27e8\u03c7\u27e9:')
        self.var_piston_display = tk.StringVar(value='—')
        _ibval(self.var_piston_display, gap=0)
        ttk.Label(info_bar, text='rad').pack(side=tk.LEFT, padx=(2, 8))

        # Total RMS = sqrt(⟨χ²⟩) — includes piston; decomposes as sqrt(σ² + ⟨χ⟩²)
        _iblbl('\u03c3\u209c\u2092\u209c\u2090\u2097:')
        self.var_total_phase_display = tk.StringVar(value='—')
        _ibval(self.var_total_phase_display, gap=0)
        ttk.Label(info_bar, text='rad').pack(side=tk.LEFT, padx=(2, 0))

        # ---- Controls row: display mode + colormap ----
        ctrl_row = ttk.Frame(tab)
        ctrl_row.pack(fill=tk.X, padx=6, pady=(1, 2))

        mode_bar = ttk.LabelFrame(ctrl_row, text='Display mode', padding=2)
        mode_bar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 6))

        self.var_probe_mode = tk.StringVar(value='intensity')
        for label, val in [('2D slice — x',  'slice_x'),
                            ('2D slice — y',  'slice_y'),
                            ('Intensity map', 'intensity'),
                            ('Contour map',   'contour'),
                            ('3D surface',    '3d')]:
            ttk.Radiobutton(mode_bar, text=label, variable=self.var_probe_mode,
                            value=val, command=self._on_probe_mode_change).pack(
                side=tk.LEFT, padx=4)

        cmap_bar = ttk.LabelFrame(ctrl_row, text='Colour scheme', padding=2)
        cmap_bar.pack(side=tk.LEFT, fill=tk.Y)

        self.var_colormap = tk.StringVar(value='inferno')
        for label, cmap in COLORMAP_OPTIONS:
            ttk.Radiobutton(cmap_bar, text=label, variable=self.var_colormap,
                            value=cmap, command=self._on_probe_mode_change).pack(
                side=tk.LEFT, padx=4)

        # ---- Figure ----
        self.fig_probe = Figure(figsize=(10, 6.5))
        self.fig_probe.set_tight_layout({'pad': 0.4, 'h_pad': 0.4, 'w_pad': 0.4})
        self.ax_phase  = self.fig_probe.add_subplot(2, 2, 1)
        self.ax_probe  = self.fig_probe.add_subplot(2, 2, 2)

        canvas = FigureCanvasTkAgg(self.fig_probe, master=tab)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        NavigationToolbar2Tk(canvas, tab).pack(fill=tk.X)
        self.canvas_probe = canvas

    # ------------------------------------------------------------------
    # Right panel — Tab 2: Resolution vs alpha
    # ------------------------------------------------------------------

    def _build_resolution_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=' Resolution vs \u03b1 ')

        ctrl = ttk.LabelFrame(tab, text='Convergence angle sweep', padding=6)
        ctrl.pack(fill=tk.X, padx=8, pady=4)

        row = ttk.Frame(ctrl)
        row.pack(fill=tk.X)

        def _lbl(text): ttk.Label(row, text=text).pack(side=tk.LEFT, padx=(8, 2))
        def _ent(var, width=6):
            ttk.Entry(row, textvariable=var, width=width).pack(side=tk.LEFT)

        _lbl('\u03b1 min (mrad):');  self.var_amin   = tk.StringVar(value='5');  _ent(self.var_amin)
        _lbl('\u03b1 max (mrad):');  self.var_amax   = tk.StringVar(value='40'); _ent(self.var_amax)
        _lbl('Steps:');               self.var_asteps = tk.StringVar(value='20'); _ent(self.var_asteps, 5)

        self.var_fast_sweep = tk.BooleanVar(value=False)
        ttk.Checkbutton(row, variable=self.var_fast_sweep,
                        text='Fast sweep (skip focus spread & source size)').pack(
            side=tk.LEFT, padx=10)

        ttk.Separator(row, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Label(row, text='Moving avg:').pack(side=tk.LEFT)
        self.var_smooth_res = tk.StringVar(value='1')
        for _k in ('1', '2', '3', '4'):
            ttk.Radiobutton(row, text=_k, variable=self.var_smooth_res,
                            value=_k).pack(side=tk.LEFT, padx=2)
        self.var_smooth_res.trace_add('write', self._on_res_smooth_change)

        ttk.Separator(row, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Label(row, text='Theory overlay:').pack(side=tk.LEFT)
        self.var_wc_overlay = tk.StringVar(value='off')
        for _val, _txt in (('off', 'Off'),
                            ('total', 'W&C total'),
                            ('components', 'W&C components')):
            ttk.Radiobutton(row, text=_txt, variable=self.var_wc_overlay,
                            value=_val).pack(side=tk.LEFT, padx=2)
        # Toggling the overlay rerenders from cached sweep data — no re-sweep needed.
        self.var_wc_overlay.trace_add('write', self._on_res_smooth_change)

        self.btn_res2 = ttk.Button(row, text='Run Sweep',
                                    style='Accent.TButton',
                                    command=self._on_calc_resolution)
        self.btn_res2.pack(side=tk.LEFT, padx=10)

        self.res_progress = ttk.Progressbar(ctrl, mode='determinate', length=400)
        self.res_progress.pack(fill=tk.X, pady=4)

        self.fig_res = Figure(figsize=(10, 5), tight_layout=True)
        self.ax_res  = self.fig_res.add_subplot(1, 1, 1)
        self._init_res_axes()

        canvas = FigureCanvasTkAgg(self.fig_res, master=tab)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        NavigationToolbar2Tk(canvas, tab).pack(fill=tk.X)
        self.canvas_res = canvas

    def _init_res_axes(self):
        self.ax_res.set_xlabel('Semi-convergence angle \u03b1 (mrad)', fontsize=11)
        self.ax_res.set_ylabel('Probe diameter d\u2085\u2080 (pm)', fontsize=11)
        self.ax_res.set_title('Probe resolution vs. convergence angle')
        self.ax_res.grid(True, alpha=0.3)

    # ------------------------------------------------------------------
    # Right panel — Tab 3: Resolution vs Defocus
    # ------------------------------------------------------------------

    def _build_defocus_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=' Resolution vs Defocus ')

        ctrl = ttk.LabelFrame(tab,
                               text='Defocus sweep  (uses \u03b1 from Beam & Optics panel)',
                               padding=6)
        ctrl.pack(fill=tk.X, padx=8, pady=4)

        ttk.Label(ctrl,
                  text='Convergence angle \u03b1 is taken from the left panel.  '
                       'All other aberrations are held fixed; only \u0394f is swept.  '
                       'Both Scherzer lines are marked on the plot.',
                  style='Info.TLabel', wraplength=800, justify=tk.LEFT).pack(
            anchor='w', pady=(0, 4))

        param_row = ttk.Frame(ctrl)
        param_row.pack(fill=tk.X)

        def _lbl(text): ttk.Label(param_row, text=text).pack(side=tk.LEFT, padx=(8, 2))
        def _ent(var, width=8):
            ttk.Entry(param_row, textvariable=var, width=width).pack(side=tk.LEFT)

        _lbl('\u0394f min (nm):');  self.var_dfmin   = tk.StringVar(value='-50');  _ent(self.var_dfmin)
        _lbl('\u0394f max (nm):');  self.var_dfmax   = tk.StringVar(value='50');   _ent(self.var_dfmax)
        _lbl('Steps:');              self.var_dfsteps = tk.StringVar(value='30');   _ent(self.var_dfsteps, 5)

        self.var_fast_defoc = tk.BooleanVar(value=False)
        ttk.Checkbutton(param_row, variable=self.var_fast_defoc,
                        text='Fast (skip focus spread & source size)').pack(
            side=tk.LEFT, padx=10)

        ttk.Separator(param_row, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Label(param_row, text='Moving avg:').pack(side=tk.LEFT)
        self.var_smooth_defoc = tk.StringVar(value='1')
        for _k in ('1', '2', '3', '4'):
            ttk.Radiobutton(param_row, text=_k, variable=self.var_smooth_defoc,
                            value=_k).pack(side=tk.LEFT, padx=2)
        self.var_smooth_defoc.trace_add('write', self._on_defoc_smooth_change)

        self.btn_defoc2 = ttk.Button(param_row, text='Run Defocus Sweep',
                                      style='Accent.TButton',
                                      command=self._on_calc_defocus)
        self.btn_defoc2.pack(side=tk.LEFT, padx=10)

        self.defoc_progress = ttk.Progressbar(ctrl, mode='determinate', length=400)
        self.defoc_progress.pack(fill=tk.X, pady=4)

        alpha_row = ttk.Frame(ctrl)
        alpha_row.pack(fill=tk.X)
        ttk.Label(alpha_row, text='Using \u03b1 =', style='Header.TLabel').pack(
            side=tk.LEFT, padx=4)
        self.var_defoc_alpha_display = tk.StringVar(value='—')
        ttk.Label(alpha_row, textvariable=self.var_defoc_alpha_display,
                  style='Title.TLabel').pack(side=tk.LEFT)

        self.fig_defoc = Figure(figsize=(10, 5), tight_layout=True)
        self.ax_defoc  = self.fig_defoc.add_subplot(1, 1, 1)
        self._init_defoc_axes()

        canvas = FigureCanvasTkAgg(self.fig_defoc, master=tab)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        NavigationToolbar2Tk(canvas, tab).pack(fill=tk.X)
        self.canvas_defoc = canvas

    def _init_defoc_axes(self):
        self.ax_defoc.set_xlabel('Defocus \u0394f (nm)', fontsize=11)
        self.ax_defoc.set_ylabel('Probe diameter d\u2085\u2080 (pm)', fontsize=11)
        self.ax_defoc.set_title('Probe resolution vs. defocus')
        self.ax_defoc.grid(True, alpha=0.3)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_aberrations(self):
        result = {}
        for (m, n), (en, vx, vy) in self.aber_rows.items():
            if en.get():
                try:
                    result[(m, n)] = (float(vx.get()), float(vy.get()))
                except ValueError:
                    pass
        return result

    def _get_physics(self):
        """
        Read all beam/optics parameters from the GUI and return them as a dict.

        Focal spread (fs) — temporal coherence
            The user enters δE as the FWHM of the zero-loss peak (EELS).
            The full conversion (Reimer & Kohl 5th ed., eq. 6.40) is:
                σ_ΔE = δE_FWHM / (2·√(2·ln 2))         [FWHM → std-dev]
                f_t  = (1 + E/511) / (1 + E/1022)       [relativistic; E in keV]
                σ_f  = Cc · σ_ΔE / E₀ · f_t             [nm; E₀ = e_kev × 1000 eV]
            Set to None when the Temporal checkbox is off.

        Source size (ss) — spatial coherence
            Computed as  σ = Physical_Size / Demagnification  (both in nm).
            This is the apparent source radius as seen at the sample plane.
            Set to None when the Spatial checkbox is off.

        Returns
        -------
        dict with keys: e_kev, dx, num_k, alpha, wl, fs, ss
        """
        e_kev = float(self.var_energy.get())
        dx    = float(self.var_dx.get())
        num_k = int(self.var_numk.get())
        alpha = float(self.var_alpha.get())
        # Focal spread σ_f — full relativistic chain (Reimer eq. 6.40):
        #   σ_ΔE = δE_FWHM / (2·√(2·ln2))   — FWHM (EELS ZLP) → std-dev
        #   f_t  = (1+E/511) / (1+E/1022)    — relativistic correction (E in keV)
        #   σ_f  = Cc · σ_ΔE / E₀ · f_t      — defocus std-dev in nm
        if self.var_fs_on.get():
            cc       = float(self.var_cc.get())
            dE       = float(self.var_dE.get())            # FWHM in eV
            sigma_dE = dE / (2.0 * np.sqrt(2.0 * np.log(2.0)))   # → std-dev
            f_t      = (1.0 + e_kev / 511.0) / (1.0 + e_kev / 1022.0)
            fs       = cc * sigma_dE / (e_kev * 1000.0) * f_t
        else:
            fs   = None
        # Apparent source σ_s — user enters physical FWHM; convert to std-dev.
        # σ_s = FWHM_physical / (2·√(2·ln2) · M)  ≈ FWHM / (2.355 · M)
        if self.var_ss_on.get():
            fwhm  = float(self.var_phys_size.get())
            demag = float(self.var_demag.get())
            ss    = fwhm / (2.0 * np.sqrt(2.0 * np.log(2.0)) * demag)
        else:
            ss = None
        wl    = calc_wavelength(e_kev)
        return dict(e_kev=e_kev, dx=dx, num_k=num_k, alpha=alpha, wl=wl, fs=fs, ss=ss)

    def _get_cs_nm(self):
        """Return Cs (nm) from aberration table, or None if disabled/invalid."""
        en, vx, _ = self.aber_rows.get((4, 0), (None, None, None))
        if en is None or not en.get():
            return None
        try:
            return float(vx.get())
        except ValueError:
            return None

    def _compute_both_scherzer(self):
        """
        Returns (df_tem, df_stem, tem_str, stem_str).
        Values in nm; strings formatted to 3 decimal places.
        """
        try:
            wl = calc_wavelength(float(self.var_energy.get()))
        except (ValueError, ZeroDivisionError):
            return None, None, '—', '—'

        cs = self._get_cs_nm()
        if cs is None:
            return None, None, 'Cs disabled', 'Cs disabled'

        try:
            phi = float(self.var_phi_max.get())
        except Exception:
            phi = -np.pi / 2

        df_tem  = calc_scherzer_tem(cs, wl)
        df_stem = calc_scherzer_stem_general(cs, wl, phi)

        tem_str  = (f'{df_tem:.3f} nm  ({df_tem*1000:.1f} pm)'
                    if df_tem  is not None else 'Cs \u2264 0')
        stem_str = (f'{df_stem:.3f} nm  ({df_stem*1000:.1f} pm)'
                    if df_stem is not None else 'Cs \u2264 0')

        return df_tem, df_stem, tem_str, stem_str

    def _set_default_scherzer_defocus(self):
        """
        After GUI is built, initialise defocus, convergence angle and Scherzer
        display labels using the current phi_max, energy and Cs values.
        """
        self._update_scherzer_display()

    @staticmethod
    def _build_kgrid(num_k, dx):
        """
        Build the 2-D reciprocal-space coordinate grids used by all physics routines.

        The grid follows NumPy FFT ordering (0, 1, …, N/2−1, −N/2, …, −1) so that
        it matches the output of np.fft.ifft2 without any fftshift needed during
        the probe calculation.  fftshift is applied only for display purposes in
        _render_probe_view.

        dk = 1 / (dx × N)  is the reciprocal-space pixel size in nm⁻¹.

        Returns
        -------
        akx : ndarray (N×N) — kx coordinates in nm⁻¹
        aky : ndarray (N×N) — ky coordinates in nm⁻¹  (= akx.T by symmetry)
        dk  : float         — reciprocal-space pixel size in nm⁻¹
        """
        ikx = (np.fft.fftfreq(num_k) * num_k).astype(int)   # integer pixel indices
        dk  = 1.0 / (dx * num_k)                              # nm⁻¹ per pixel
        akx = np.tile(ikx, (num_k, 1)) * dk                   # 2-D kx grid
        aky = akx.T                                            # 2-D ky grid
        return akx, aky, dk

    def _set_buttons(self, probe_state, res_state, defoc_state, demag_state=None,
                     map_state=None, current_state=None):
        if demag_state is None:
            demag_state = defoc_state
        if map_state is None:
            map_state = defoc_state
        if current_state is None:
            current_state = defoc_state
        self.btn_probe.config(state=probe_state)
        self.btn_res.config(state=res_state)
        self.btn_res2.config(state=res_state)
        self.btn_defoc.config(state=defoc_state)
        self.btn_defoc2.config(state=defoc_state)
        self.btn_demag.config(state=demag_state)
        self.btn_cur.config(state=current_state)
        self.btn_map.config(state=map_state)
        # Stop button is enabled when any calculation is running (i.e. calc buttons disabled)
        stop_state = 'normal' if probe_state == 'disabled' else 'disabled'
        self.btn_stop.config(state=stop_state)

    def _on_stop(self):
        """Signal all running calculation threads to stop after the current step."""
        self._stop_event.set()
        self.var_status.set('Stopping\u2026')
        self.btn_stop.configure(state='disabled')

    # ------------------------------------------------------------------
    # Probe view — mode / colormap switching
    # ------------------------------------------------------------------

    def _on_probe_mode_change(self):
        if self._probe_render_data is not None:
            self._render_probe_view(*self._probe_render_data)

    def _render_probe_view(self, phi, probe, d50, alpha, k_alpha, dk, alen, k_max):
        mode = self.var_probe_mode.get()
        cmap = self.var_colormap.get()

        # Rebuild with a 2×2 gridspec: top row = 2D images, bottom row = radial avgs.
        # height_ratios [2.5, 1] gives the bottom row ~28 % of figure height so labels
        # and the secondary degree y-axis are readable without excessive empty space.
        # right=0.93 leaves room for the twinx degree tick labels on ax_pr.
        self.fig_probe.clear()
        gs = self.fig_probe.add_gridspec(2, 2,
                                          height_ratios=[2.5, 1],
                                          hspace=0.38, wspace=0.38,
                                          top=0.94, bottom=0.08,
                                          left=0.07, right=0.93)
        self.ax_phase = self.fig_probe.add_subplot(gs[0, 0])
        if mode == '3d':
            self.ax_probe = self.fig_probe.add_subplot(gs[0, 1], projection='3d')
        else:
            self.ax_probe = self.fig_probe.add_subplot(gs[0, 1])
        ax_pr = self.fig_probe.add_subplot(gs[1, 0])   # phase plate radial avg
        ax_rr = self.fig_probe.add_subplot(gs[1, 1])   # probe radial avg

        # ---- Phase plate (left, always 2-D, bwr) ----
        ax_p = self.ax_phase
        ax_p.imshow(np.fft.fftshift(np.sin(phi)), vmin=-1, vmax=1,
                    cmap='bwr', origin='lower',
                    extent=(-k_max, k_max - dk, -k_max, k_max - dk))
        ax_p.set_xlim(-1.5 * k_alpha, 1.5 * k_alpha)
        ax_p.set_ylim(-1.5 * k_alpha, 1.5 * k_alpha)
        ax_p.set_xlabel('$q_x$  [nm$^{-1}$]')
        ax_p.set_ylabel('$q_y$  [nm$^{-1}$]')
        ax_p.set_title(f'Phase plate  (\u03b1 = {alpha:.1f} mrad)')
        ax_p_top = ax_p.twiny()
        ax_p_top.set_xlim(-1.5 * alpha, 1.5 * alpha)
        ax_p_top.set_xlabel('\u03b1  (mrad)', fontsize=8)
        ax_p_top.tick_params(labelsize=7)

        # ---- Probe view (right) ----
        ax_r     = self.ax_probe
        probe_sh = np.fft.fftshift(probe)
        ny, nx   = probe_sh.shape
        half     = max(5 * d50, 4 * dk)

        x_arr = np.linspace(-alen / 2, alen / 2, nx)
        y_arr = np.linspace(-alen / 2, alen / 2, ny)
        xi    = (np.searchsorted(x_arr, -half), np.searchsorted(x_arr, half))
        yi    = (np.searchsorted(y_arr, -half), np.searchsorted(y_arr, half))
        x_roi = x_arr[xi[0]:xi[1]]
        y_roi = y_arr[yi[0]:yi[1]]
        p_roi = probe_sh[yi[0]:yi[1], xi[0]:xi[1]]

        title_base = f'd\u2085\u2080 = {d50*1000:.2f} pm'

        if mode == 'intensity':
            im = ax_r.imshow(p_roi, vmin=0, cmap=cmap, origin='lower',
                             extent=(x_roi[0], x_roi[-1], y_roi[0], y_roi[-1]))
            self.fig_probe.colorbar(im, ax=ax_r, fraction=0.046, pad=0.04,
                                    label='Intensity (a.u.)')
            ax_r.set_xlabel('$x$  [nm]')
            ax_r.set_ylabel('$y$  [nm]')
            ax_r.set_title(f'Probe intensity  ({title_base})')

        elif mode == 'contour':
            X, Y = np.meshgrid(x_roi, y_roi)
            cs_plot = ax_r.contourf(X, Y, p_roi, levels=20, cmap=cmap)
            ax_r.contour(X, Y, p_roi, levels=20, colors='white',
                         linewidths=0.4, alpha=0.4)
            self.fig_probe.colorbar(cs_plot, ax=ax_r, fraction=0.046, pad=0.04,
                                    label='Intensity (a.u.)')
            ax_r.set_xlabel('$x$  [nm]')
            ax_r.set_ylabel('$y$  [nm]')
            ax_r.set_title(f'Probe contour  ({title_base})')
            ax_r.set_aspect('equal')

        elif mode == 'slice_x':
            cy = ny // 2
            ax_r.plot(x_arr, probe_sh[cy, :], color='steelblue', linewidth=1.5)
            ax_r.set_xlim(-half, half)
            ax_r.set_xlabel('$x$  [nm]')
            ax_r.set_ylabel('Intensity (a.u.)')
            ax_r.set_title(f'Horizontal slice (y = 0)  ({title_base})')
            ax_r.grid(True, alpha=0.3)
            ax_r.axvline(0, color='gray', lw=0.8, ls='--')

        elif mode == 'slice_y':
            cx = nx // 2
            ax_r.plot(probe_sh[:, cx], y_arr, color='darkorange', linewidth=1.5)
            ax_r.set_ylim(-half, half)
            ax_r.set_ylabel('$y$  [nm]')
            ax_r.set_xlabel('Intensity (a.u.)')
            ax_r.set_title(f'Vertical slice (x = 0)  ({title_base})')
            ax_r.grid(True, alpha=0.3)
            ax_r.axhline(0, color='gray', lw=0.8, ls='--')

        elif mode == '3d':
            step = max(1, p_roi.shape[0] // 128)
            p_ds = p_roi[::step, ::step]
            x_ds = x_roi[::step]
            y_ds = y_roi[::step]
            X, Y = np.meshgrid(x_ds, y_ds)
            ax_r.plot_surface(X, Y, p_ds, cmap=cmap,
                              rstride=1, cstride=1,
                              linewidth=0, antialiased=False)
            ax_r.view_init(elev=35, azim=45)
            ax_r.set_xlabel('$x$ [nm]')
            ax_r.set_ylabel('$y$ [nm]')
            ax_r.set_zlabel('Intensity')
            ax_r.set_title(f'3-D surface  ({title_base})')

        # ── Radial average: phase plate (bottom-left) ────────────────────────
        # Plot χ(|k|) averaged azimuthally from the aperture centre.
        # The phase plate is multiplied by the aperture in get_phaseplate, so
        # χ → 0 outside the aperture; we limit the plot to 1.5× the aperture radius.
        phi_sh   = np.fft.fftshift(phi)
        r_k, phi_avg = radial_profile(phi_sh, pixel_scale=dk,
                                       max_r=1.5 * k_alpha)
        ax_pr.plot(r_k, phi_avg, color='#44448b', linewidth=1.5)
        ax_pr.axhline(0, color='gray', lw=0.7, ls='--', alpha=0.6)
        ax_pr.axvline(k_alpha, color='#8b1a1a', lw=0.9, ls='--', alpha=0.7,
                      label='aperture edge')
        # ── π/2 reference line ───────────────────────────────────────────────
        ax_pr.axhline(-np.pi / 2, color='#e07b00', lw=1.0, ls='-.',
                      alpha=0.85, label='\u03c7 = \u2212\u03c0/2')

        # ── Phase threshold table: q and α where |χ| first reaches each level ─
        # wl (nm) recovered from k_alpha = α_rad / wl → wl = (alpha·1e-3)/k_alpha
        _wl_nm = (alpha * 1e-3) / k_alpha          # wavelength in nm
        _tv_list  = [np.pi/8, np.pi/4, np.pi/2, 3*np.pi/4, np.pi]
        _tl_list  = ['\u03c0/8', '\u03c0/4', '\u03c0/2', '3\u03c0/4', '\u03c0']
        _col1 = '|\u03c7|';  _col2 = 'q(nm\u207b\u00b9)';  _col3 = '\u03b1(mrad)'
        _hdr  = f"{_col1:<6s}  {_col2:>8s}  {_col3:>7s}"
        _sep  = '\u2500' * len(_hdr)
        _rows = [_hdr, _sep]
        for _tv, _tl in zip(_tv_list, _tl_list):
            _qc = None
            # Find first descending crossing of -_tv in phi_avg
            for _j in range(1, len(phi_avg)):
                if phi_avg[_j - 1] > -_tv >= phi_avg[_j]:
                    _frac = (-_tv - phi_avg[_j - 1]) / (phi_avg[_j] - phi_avg[_j - 1])
                    _qc   = r_k[_j - 1] + _frac * (r_k[_j] - r_k[_j - 1])
                    break
            if _qc is not None:
                _ac = _qc * _wl_nm * 1000          # q → α  (mrad)
                _rows.append(f'{_tl:<6s}  {_qc:>8.3f}  {_ac:>7.2f}')
            else:
                _rows.append(f'{_tl:<6s}  {"—":>8s}  {"—":>7s}')
        ax_pr.text(0.02, 0.03, '\n'.join(_rows),
                   transform=ax_pr.transAxes, fontsize=5.5,
                   verticalalignment='bottom', fontfamily='monospace',
                   bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                             alpha=0.82, edgecolor='#aaaaaa', linewidth=0.7))

        ax_pr.set_xlim(0, 1.5 * k_alpha)
        ax_pr.set_xlabel('q  [nm\u207b\u00b9]', fontsize=8)
        ax_pr.set_ylabel('\u03c7  (rad)', fontsize=8)
        ax_pr.set_title('Phase plate — radial avg', fontsize=8)
        ax_pr.tick_params(labelsize=7)
        ax_pr.legend(fontsize=7, loc='upper right')
        ax_pr.grid(True, alpha=0.3)
        ax_pr_top = ax_pr.twiny()
        ax_pr_top.set_xlim(0, 1.5 * alpha)
        ax_pr_top.set_xlabel('\u03b1  (mrad)', fontsize=8)
        ax_pr_top.tick_params(labelsize=7)
        ax_pr_deg = ax_pr.secondary_yaxis(
            'right', functions=(np.degrees, np.radians))
        ax_pr_deg.set_ylabel('\u03c7  (\u00b0)', fontsize=8)
        ax_pr_deg.tick_params(labelsize=7)

        # ── Radial average: probe intensity (bottom-right) ────────────────────
        # Plot I(r) averaged azimuthally.  Limit to 6× d50 so the peak is visible
        # without too much empty space.  Mark r = d50/2 (the half-enclosed-power
        # radius) with a vertical line.
        probe_sh = np.fft.fftshift(probe)
        num_k    = probe_sh.shape[0]
        dx_      = alen / num_k            # real-space pixel size (nm)
        max_r_nm = max(6.0 * d50, 4.0 * dx_)
        r_r, probe_avg = radial_profile(probe_sh, pixel_scale=dx_, max_r=max_r_nm)
        ax_rr.plot(r_r, probe_avg, color='steelblue', linewidth=1.5)
        ax_rr.axvline(d50 / 2, color='#8b1a1a', lw=0.9, ls='--', alpha=0.7,
                      label=f'd\u2085\u2080/2 = {d50/2*1000:.1f} pm')
        ax_rr.set_xlim(0, max_r_nm)
        ax_rr.set_ylim(bottom=0)
        ax_rr.set_xlabel('r  (nm)', fontsize=8)
        ax_rr.set_ylabel('Intensity (a.u.)', fontsize=8)
        ax_rr.set_title('Probe — radial avg', fontsize=8)
        ax_rr.tick_params(labelsize=7)
        ax_rr.legend(fontsize=7, loc='upper right')
        ax_rr.grid(True, alpha=0.3)

        self.canvas_probe.draw()

    # ------------------------------------------------------------------
    # Calculate Probe
    # ------------------------------------------------------------------

    def _on_calc_probe(self):
        if self._calc_running or self._res_running or self._defoc_running or self._demag_running or self._current_running:
            return
        self._stop_event.clear()
        self._calc_running = True
        self._set_buttons('disabled', 'disabled', 'disabled')
        self.var_status.set('Calculating probe…')
        self.progress.start(12)
        threading.Thread(target=self._calc_probe_thread, daemon=True).start()

    def _calc_probe_thread(self):
        try:
            phys  = self._get_physics()
            abers = self._get_aberrations()
            num_k = phys['num_k']
            dx    = phys['dx']
            wl    = phys['wl']
            alpha = phys['alpha']
            akx, aky, dk = self._build_kgrid(num_k, dx)
            k_alpha = 1e-3 * alpha / wl
            k_max   = dk * (num_k >> 1)
            alen    = dx * num_k

            warns = []
            if k_alpha > 0.66 * k_max:
                warns.append(f'Aperture edge ({k_alpha:.2f}/nm) near Nyquist ({k_max:.2f}/nm) — reduce dx.')
            if k_alpha / dk < 5:
                warns.append('Very few beams inside aperture — increase grid size or dx.')

            aperture = get_aperture(k_alpha, dk, akx, aky)
            phi      = get_phaseplate(abers, aperture, akx, aky, wl)

            fs = phys['fs']
            ss = phys['ss']
            if fs is not None:
                # Parallelise the 21-step Gaussian quadrature over defocus.
                # Each worker reconstructs its own k-grid/aperture to avoid
                # pickling large arrays across process boundaries.
                dfs      = fs / 5.0
                fs_args  = []
                for ifs in range(-10, 11):
                    df_offset = dfs * ifs
                    weight    = np.exp(-df_offset**2 / fs**2)
                    # Fold this step's defocus into the (2,0) coefficient
                    abers_step = dict(abers)
                    ax0, ay0   = abers_step.get((2, 0), (0.0, 0.0))
                    abers_step[(2, 0)] = (ax0 + df_offset, ay0)
                    fs_args.append((abers_step, k_alpha, dk, num_k, dx, wl, weight))
                with ProcessPoolExecutor() as pool:
                    step_probes = list(pool.map(_worker_focus_step, fs_args))
                intens_r = sum(step_probes)
                # Apply source-size envelope after summing all focus steps
                if ss is not None:
                    sk       = np.exp(-np.pi**2 * ss**2 * (akx**2 + aky**2))
                    intens_r = np.fft.ifft2(np.fft.fft2(intens_r) * sk).real
                total = np.sum(intens_r)
                probe = intens_r / total if total > 0 else intens_r
            else:
                # No focus spread — single coherent probe, nothing to parallelise
                probe = get_probe(abers, aperture, akx, aky, wl, source_size=ss)

            d50                                                        = get_d50(probe, dx, self._get_edge_mult())
            s_exact, s_marechal, sigma_rad, sigma_waves, pv_rad, mean_phi_rad, sigma_total_rad = \
                calc_strehl(phi, aperture)

            self.root.after(0, self._update_probe_plot,
                             phi, probe, d50, alpha, k_alpha, dk, alen, k_max,
                             warns, s_exact, s_marechal,
                             sigma_rad, sigma_waves, pv_rad, mean_phi_rad, sigma_total_rad)
        except Exception as exc:
            self.root.after(0, self._show_error, str(exc))
        finally:
            self.root.after(0, self._calc_probe_done)

    def _compute_dose_limited_pm(self, I_A, d50_m):
        """
        Dose-limited resolution and combined effective resolution, both in pm.

        Physical motivation:
          Even with perfect optics (d₅₀ → 0), an image cannot resolve features
          smaller than the statistical noise allows.  At zero beam current no
          electrons arrive and nothing is resolvable — d_eff must diverge as I→0.
          The shot-noise model for Δ(d₅₀) does NOT capture this; it only
          perturbs the source-size term.  The dose limit is a separate,
          more fundamental constraint from detector statistics.

        Rose criterion:
          A feature of contrast C (fraction of electrons scattered by the
          feature, 0–1) is reliably detectable when SNR ≥ k, where k is the
          user-entered Rose SNR threshold (default 5, classic Rose criterion).

          The relevant electron count is NOT per pixel but per probe footprint:
            N_pixel = I · T / e              (electrons per dwell, per pixel)
            N_probe = N_pixel · (d₅₀ / dx)² (electrons within one probe area)

          Using N_pixel instead of N_probe overestimates d_dose by √(d₅₀/dx)
          — typically 2–5× — because fewer electrons appear to illuminate the
          feature than actually do.

          Rose criterion applied at the probe scale:
            SNR = C · √N_probe ≥ k

          Crossover (optics-limited ↔ statistics-limited):
            N_probe_cross = (k / C)²   →   d_dose = d₅₀ at crossover

        Dose-limited resolution:
          d_dose = k · d₅₀ / (C · √N_probe)

          Interpretation:
            - N_probe >> N_cross: d_dose << d₅₀ → optics-limited (good regime)
            - N_probe = N_cross:  d_dose = d₅₀  → crossover
            - N_probe << N_cross: d_dose >> d₅₀ → statistics-limited
            - N_probe → 0:        d_dose → ∞    → nothing resolvable

        Combined effective resolution (quadrature sum):
          d_eff = √(d₅₀² + d_dose²)

        Parameters
        ----------
        I_A   : beam current in amperes
        d50_m : d₅₀ probe diameter in metres (pass d50_nm × 1e-9)

        Returns
        -------
        (d_dose_pm, d_eff_pm) as floats, or None if preconditions not met.
        """
        SNR_THRESHOLD = self._get_rose_k()
        try:
            dwell_us = float(self.var_dwell_us.get())
            contrast = float(self.var_contrast.get())
            dx_nm    = float(self.var_dx.get())
            if dwell_us <= 0 or contrast <= 0 or contrast > 1 or dx_nm <= 0:
                return None
            if I_A is None or I_A <= 0 or d50_m <= 0:
                return None
            e_C     = 1.602e-19               # electron charge (C)
            T_s     = dwell_us * 1e-6         # dwell time in seconds
            dx_m    = dx_nm * 1e-9            # pixel size in metres
            N_pixel = I_A * T_s / e_C         # electrons per pixel per dwell
            # Scale to electrons per probe footprint: N_probe = N_pixel·(π/4)·(d₅₀/dx)²
            N_probe = N_pixel * (np.pi / 4.0) * (d50_m / dx_m) ** 2
            if N_probe <= 0:
                return None
            # d_dose: minimum detectable feature size given the probe-area dose.
            # Derived from C·√N_probe ≥ k  →  d_dose = k·d₅₀/(C·√N_probe)
            d_dose_m = SNR_THRESHOLD * d50_m / (contrast * np.sqrt(N_probe))
            d_eff_m  = np.sqrt(d50_m**2 + d_dose_m**2)
            return d_dose_m * 1e12, d_eff_m * 1e12   # metres → pm
        except (ValueError, ZeroDivisionError):
            return None

    def _compute_d50_error_pm(self, d50_m):   # d50_m must be in metres (nm × 1e-9)
        """
        ±1σ shot-noise uncertainty on d₅₀ in pm, or None if preconditions unmet.

        Model:
          The beam current is I = B_r · V₀ · π² · α² · σ²_apparent
            B_r        — reduced brightness (A/m²/sr/V)
            V₀         — accelerating voltage (V)
            α          — semi-convergence angle (rad)
            σ_apparent — apparent source radius = phys_size_nm·1e-9 / (2·M)

          Electrons per pixel: N = I · T / e
            T — pixel dwell time (s),  e = 1.602×10⁻¹⁹ C

          Poisson shot noise gives a fractional uncertainty in the measured
          current (and hence in the apparent source area σ²) of 1/√N:
            Δ(σ²) / σ² = 1/√N

          Propagating through d₅₀² = d_source² + d_diff² + d_aber²:
            2·d₅₀·Δ(d₅₀) = Δ(d_source²) = σ²_apparent / √N
          so:
            Δ(d₅₀) = σ²_apparent / (2 · d₅₀ · √N)   [metres]

          Limiting behaviour:
            - High current (N → ∞): Δ(d₅₀) → 0  (shot noise negligible)
            - Aberration-limited (d_source ≪ d₅₀): σ²/d₅₀ is small, bars collapse
            - Source-limited (d_source ≈ d₅₀): bars are widest
            - Zero current (N → 0): N and σ_apparent both → 0, so Δ(d₅₀) → 0 too;
              the fractional uncertainty 1/√N → ∞ but the absolute error
              vanishes — there is no finite noise floor at zero current.

        Returns float (pm) or None.
        """
        if not self.var_br_on.get() or not self.var_ss_on.get():
            return None
        try:
            br        = float(self.var_br_custom.get())
            e_kev     = float(self.var_energy.get())
            alpha_mrad = float(self.var_alpha.get())
            phys_size = float(self.var_phys_size.get())
            demag     = float(self.var_demag.get())
            dwell_us  = float(self.var_dwell_us.get())
            if br <= 0 or e_kev <= 0 or alpha_mrad <= 0 or phys_size <= 0 \
                    or demag <= 0 or dwell_us <= 0 or d50_m <= 0:
                return None
            v0        = e_kev * 1000.0
            alpha_r   = alpha_mrad * 1e-3
            sigma_m   = phys_size * 1e-9 / (2.0 * demag)   # apparent source radius (m)
            T_s       = dwell_us * 1e-6                     # dwell time (s)
            e_C       = 1.602e-19                           # electron charge (C)
            I_A       = br * v0 * np.pi**2 * alpha_r**2 * sigma_m**2   # current (A)
            N         = I_A * T_s / e_C                    # electrons per pixel
            if N <= 0:
                return None
            delta_m   = sigma_m**2 / (2.0 * d50_m * np.sqrt(N))   # Δ(d₅₀) in metres
            return delta_m * 1e12   # convert to pm
        except (ValueError, ZeroDivisionError):
            return None

    def _update_probe_plot(self, phi, probe, d50, alpha, k_alpha, dk, alen, k_max,
                           warns, s_exact, s_marechal,
                           sigma_rad, sigma_waves, pv_rad, mean_phi_rad, sigma_total_rad):
        self._probe_render_data = (phi, probe, d50, alpha, k_alpha, dk, alen, k_max)

        # Update Scherzer displays
        df_tem, df_stem, tem_str, stem_str = self._compute_both_scherzer()
        self.var_scherzer_tem_display.set(tem_str)
        self.var_scherzer_stem_display.set(stem_str)

        # Shot-noise error on d₅₀ using the source-panel parameters.
        # Full derivation: I = B_r·V₀·π²·α²·σ²_app  →  N = I·T/e electrons/pixel
        # Δ(d₅₀) = σ²_apparent / (2·d₅₀·√N)   [propagated from Δ(σ²)/σ² = 1/√N]
        # d50 from get_d50() is in nm; convert to metres for the error formula.
        d50_m  = d50 * 1e-9   # nm → m; shared by error and dose helpers below
        err_pm = self._compute_d50_error_pm(d50_m)
        if err_pm is not None:
            self.var_d50.set(f'{d50 * 1000:.3f} \u00b1 {err_pm:.3f} pm')
        else:
            self.var_d50.set(f'{d50 * 1000:.3f} pm')

        # Dose-limited and effective resolution at the single probe operating point.
        # Requires beam current from _compute_current_pA (brightness + source size)
        # plus dwell time and contrast from the source panel.
        # d_dose and d_eff both diverge as I→0, encoding that zero current
        # means nothing is resolvable regardless of optic quality.
        dose_result = None
        if self.var_br_on.get() and self.var_ss_on.get():
            try:
                br        = float(self.var_br_custom.get())
                e_kev_val = float(self.var_energy.get())
                alpha_val = float(self.var_alpha.get())
                phys_size = float(self.var_phys_size.get())
                demag_val = float(self.var_demag.get())
                i_pA = self._compute_current_pA(br, e_kev_val, alpha_val,
                                                phys_size, demag_val)
                if i_pA is not None:
                    I_A = i_pA * 1e-12   # pA → A
                    self._last_I_A  = I_A    # store for live k updates
                    self._last_d50_m = d50_m
                    dose_result = self._compute_dose_limited_pm(I_A, d50_m)
            except (ValueError, ZeroDivisionError):
                pass
        if dose_result is not None:
            d_dose_pm, d_eff_pm = dose_result
            self.var_d_dose_display.set(f'{d_dose_pm:.1f} pm')
            self.var_d_eff_display.set(f'{d_eff_pm:.1f} pm')
        else:
            self.var_d_dose_display.set('\u2014')
            self.var_d_eff_display.set('\u2014')

        # Update Strehl display and colour-code the label
        self.var_strehl.set(f'{s_exact:.4f}  (Maréchal: {s_marechal:.4f})')
        if s_exact >= 0.8:
            self.lbl_strehl.configure(style='StrehlGood.TLabel')
        elif s_exact >= 0.5:
            self.lbl_strehl.configure(style='StrehlWarn.TLabel')
        else:
            self.lbl_strehl.configure(style='StrehlBad.TLabel')

        # Phase-error statistics
        self.var_sigma_rad.set(f'{sigma_rad:.4f}')
        self.var_sigma_waves.set(f'{sigma_waves:.4f}')
        self.var_pv_rad.set(f'{pv_rad:.4f}')
        self.var_piston_display.set(f'{mean_phi_rad:.4f}')
        self.var_total_phase_display.set(f'{sigma_total_rad:.4f}')

        self._render_probe_view(phi, probe, d50, alpha, k_alpha, dk, alen, k_max)

        status = 'Probe calculated.'
        if warns:
            status += '  \u26a0 ' + '  '.join(warns)
        self.var_status.set(status)
        self.notebook.select(0)

    def _calc_probe_done(self):
        self._calc_running = False
        self.progress.stop()
        self._set_buttons('normal', 'normal', 'normal')

    # ------------------------------------------------------------------
    # Calculate Resolution vs Alpha
    # ------------------------------------------------------------------

    def _on_calc_resolution(self):
        if self._calc_running or self._res_running or self._defoc_running or self._demag_running or self._current_running:
            return
        self._stop_event.clear()
        self._res_running = True
        self._set_buttons('disabled', 'disabled', 'disabled')
        self.var_status.set('Calculating resolution curve…')
        self.progress.start(12)
        self.res_progress['value'] = 0
        threading.Thread(target=self._calc_res_thread, daemon=True).start()

    def _calc_res_thread(self):
        """
        Background thread for the convergence-angle sweep (Tab 2).

        Runs in a daemon thread so it does not block the Tkinter main loop.
        Each alpha step is independent, so all steps are submitted to a
        ProcessPoolExecutor at once.  Child processes run on separate CPU
        cores, each with their own GIL, bypassing the Python threading GIL
        limitation that would otherwise serialise the NumPy work.

        Progress bar updates are sent back to the main thread via
        root.after(0, ...) which is the only thread-safe way to touch Tkinter
        widgets from a non-main thread.  as_completed() yields futures as
        they finish (out of order), so the bar ticks in real time; results
        are stored by index so the final list is always in alpha order.

        The k-grid is built once on the main process and passed to every
        worker to avoid redundant allocation inside each child process.
        Aberrations and physics parameters are snapshotted at sweep-start
        so mid-sweep GUI changes do not affect the running calculation.
        """
        try:
            phys        = self._get_physics()
            abers       = self._get_aberrations()
            alpha_min   = float(self.var_amin.get())
            alpha_max   = float(self.var_amax.get())
            alpha_steps = max(2, int(self.var_asteps.get()))
            fast        = self.var_fast_sweep.get()

            # Read pixel dwell time and image contrast for shot-noise error bars
            # and dose-limited resolution curves.
            # N(α) = I(α) · T / e  electrons per pixel.
            # If brightness/source-size are disabled or either parameter is
            # invalid, dwell_us/contrast are set to None and the dose curves
            # and error bars are suppressed silently.
            try:
                dwell_us = float(self.var_dwell_us.get())
                if dwell_us <= 0:
                    dwell_us = None
            except ValueError:
                dwell_us = None
            try:
                contrast = float(self.var_contrast.get())
                if contrast <= 0 or contrast > 1:
                    contrast = None
            except ValueError:
                contrast = None

            alphas = np.linspace(alpha_min, alpha_max, alpha_steps)
            num_k  = phys['num_k']
            dx     = phys['dx']
            wl     = phys['wl']
            # Fast mode skips the envelope functions for quick qualitative sweeps
            fs     = None if fast else phys['fs']
            ss     = None if fast else phys['ss']
            akx, aky, dk = self._build_kgrid(num_k, dx)

            # Pack each step's arguments into a tuple for the worker function
            edge_mult = self._get_edge_mult()
            args = [(a, abers, akx, aky, dk, wl, dx, fs, ss, edge_mult) for a in alphas]
            d50s    = [None] * len(alphas)   # pre-allocate so index assignment works
            strehls = [None] * len(alphas)   # parallel list for exact Strehl values
            completed = 0
            stopped = False
            with ProcessPoolExecutor() as pool:
                futures = {pool.submit(_worker_alpha, arg): i
                           for i, arg in enumerate(args)}
                for future in as_completed(futures):
                    if self._stop_event.is_set():
                        for f in futures:
                            f.cancel()
                        stopped = True
                        break
                    i = futures[future]
                    d50s[i], strehls[i] = future.result()
                    completed += 1
                    pct = completed / len(alphas) * 100
                    self.root.after(0, lambda p=pct: self.res_progress.configure(value=p))

            if not stopped:
                self.root.after(0, self._update_res_plot,
                                alphas, d50s, strehls, phys, fast, dwell_us, contrast)
            else:
                self.root.after(0, lambda: self.var_status.set('α sweep stopped.'))
        except Exception as exc:
            self.root.after(0, self._show_error, str(exc))
        finally:
            self.root.after(0, self._calc_res_done)

    def _calc_wc_geometric(self, alphas_mrad):
        """
        Williams & Carter geometric (ray-optics) probe size components.

        Model (W&C 2nd ed., Ch. 22, eq. 22.2):
            d_geo(α) = √( d_diff² + d_s² + d_c² + d_g² )

        Components:
            d_diff = 1.22 · λ / α          — diffraction-limited disc (nm → pm)
            d_s    = 0.5 · Cs · α³         — spherical aberration disc
            d_c    = Cc · (σ_ΔE/E) · f_t · α  — chromatic aberration disc;
                     uses the same FWHM→σ and relativistic f_t chain as _get_physics
            d_g    = FWHM_physical / M     — geometric image of the source (constant vs α)

        All components are in pm.  Disabled or unavailable components are set to
        zero so d_geo is always computed from whatever IS available.

        Parameters
        ----------
        alphas_mrad : array-like — convergence semi-angles in mrad

        Returns
        -------
        dict with keys 'total', 'diff', 's', 'c', 'g' — all 1-D arrays in pm.
        Returns None if the beam energy cannot be read.
        """
        alphas_r = np.array(alphas_mrad, dtype=float) * 1e-3   # mrad → rad
        # Guard against α = 0 (would give /0 in d_diff); replace with NaN.
        alphas_r = np.where(alphas_r > 0, alphas_r, np.nan)

        try:
            e_kev = float(self.var_energy.get())
            wl    = calc_wavelength(e_kev)   # nm
        except (ValueError, tk.TclError):
            return None

        # -- d_diff: diffraction disc  (Abbe / W&C convention: 1.22·λ/α) ------
        d_diff_pm = 1.22 * wl / alphas_r * 1000.0   # nm → pm

        # -- d_s: spherical aberration disc  (0.5·Cs·α³) ----------------------
        # Reads the (4,0) Cs entry from the aberration table; zero if disabled.
        cs_nm = self._get_cs_nm()   # None when Cs is absent or disabled
        if cs_nm is not None and cs_nm != 0.0:
            d_s_pm = 0.5 * cs_nm * alphas_r ** 3 * 1000.0   # nm → pm
        else:
            d_s_pm = np.zeros_like(alphas_r)

        # -- d_c: chromatic aberration disc  (Cc·σ_ΔE/E·f_t·α) ---------------
        # Uses the same FWHM→σ_ΔE and relativistic f_t formula as _get_physics
        # (Reimer eq. 6.40) so the two panels are internally consistent.
        d_c_pm = np.zeros_like(alphas_r)
        if self.var_fs_on.get():
            try:
                cc       = float(self.var_cc.get())
                dE_fwhm  = float(self.var_dE.get())
                sigma_dE = dE_fwhm / (2.0 * np.sqrt(2.0 * np.log(2.0)))   # FWHM → σ
                E0_keV   = 511.0                                            # m₀c² in keV
                f_t      = (1.0 + e_kev / E0_keV) / (1.0 + e_kev / (2.0 * E0_keV))
                e0_eV    = e_kev * 1000.0
                # d_c = Cc[nm]·(σ_ΔE[eV]/E[eV])·f_t·α[rad]  →  nm → pm
                d_c_pm   = cc * (sigma_dE / e0_eV) * f_t * alphas_r * 1000.0
            except (ValueError, tk.TclError):
                pass

        # -- d_g: geometric image of the source  (FWHM_apparent = FWHM/M) -----
        # W&C use the demagnified physical source size as the geometric term.
        # Convention: FWHM_apparent (the full FWHM at the specimen), not σ_s.
        d_g_pm = np.zeros_like(alphas_r)
        if self.var_ss_on.get():
            try:
                fwhm_nm = float(self.var_phys_size.get())
                demag   = float(self.var_demag.get())
                if demag > 0:
                    # d_g is constant with α — it is the demagnified source FWHM.
                    d_g_pm = np.full_like(alphas_r, fwhm_nm / demag * 1000.0)
            except (ValueError, tk.TclError):
                pass

        d_total_pm = np.sqrt(d_diff_pm**2 + d_s_pm**2 + d_c_pm**2 + d_g_pm**2)

        return {'total': d_total_pm,
                'diff':  d_diff_pm,
                's':     d_s_pm,
                'c':     d_c_pm,
                'g':     d_g_pm}

    def _update_res_plot(self, alphas, d50s, strehls, phys, fast, dwell_us=None, contrast=None):
        self._res_last_data = (alphas, d50s, strehls, phys, fast, dwell_us, contrast)
        # Apply 1D moving average, cropping edges by (k-1) total
        try:
            _k = max(1, int(self.var_smooth_res.get()))
        except (ValueError, tk.TclError):
            _k = 1
        alphas  = np.array(alphas,  dtype=float)
        d50s    = np.array(d50s,    dtype=float)
        strehls = np.array(strehls, dtype=float)
        if _k > 1:
            _kernel  = np.ones(_k) / _k
            d50s     = np.convolve(d50s,    _kernel, mode='valid')
            strehls  = np.convolve(strehls, _kernel, mode='valid')
            _cl = (_k - 1) // 2;  _cr = _k - 1 - _cl
            alphas = alphas[_cl : len(alphas) - _cr] if _cr > 0 else alphas[_cl:]

        # Clear the primary axis and remove any previously twinned axes (both
        # twinx — right y-axis — and twiny — top x-axis) to avoid stacking.
        self.ax_res.clear()
        for twin in self.ax_res.get_shared_x_axes().get_siblings(self.ax_res):
            if twin is not self.ax_res:
                twin.remove()
        for twin in self.ax_res.get_shared_y_axes().get_siblings(self.ax_res):
            if twin is not self.ax_res:
                twin.remove()

        # ── Left axis: d50 (probe resolution) ───────────────────────────────
        ax_d  = self.ax_res                # left  = d50
        ax_s  = self.ax_res.twinx()       # right = Strehl

        d50_color    = 'steelblue'
        strehl_color = '#7b2d8b'   # purple — distinct from steelblue

        d50_pm    = [d * 1000 for d in d50s]
        min_i     = int(np.argmin(d50s))
        min_alpha = alphas[min_i]
        min_d50   = d50_pm[min_i]
        d_range   = max(d50_pm) - min(d50_pm) if max(d50_pm) != min(d50_pm) else 1.0
        a_range   = alphas[-1] - alphas[0]

        # ── Shot-noise error bars on d₅₀ ────────────────────────────────────
        # Derivation:
        #   Beam current at each α: I(α) = B_r · V₀ · π² · α² · σ²_apparent
        #   Electrons per pixel:    N(α) = I(α) · T / e
        #     where T = dwell time (s), e = 1.602×10⁻¹⁹ C.
        #
        #   Shot noise is Poisson, so the fractional variance of the measured
        #   current equals 1/N:  Var(I_meas)/I² = 1/N.
        #
        #   The brightness formula gives σ² = I / (B_r · V₀ · π² · α²),
        #   so the fractional variance of the apparent source area also = 1/N:
        #     Δ(σ²)/σ² = ΔI/I = 1/√N  (one-sigma fractional uncertainty).
        #
        #   The source contributes d_source² = (σ/M)² to the probe quadrature:
        #     d₅₀² = d_source² + d_diff² + d_aber²
        #   Differentiating:
        #     2·d₅₀·Δ(d₅₀) = Δ(d_source²) = d_source²·(1/√N)
        #   So the ±1σ half-width of the error bar on d₅₀ is:
        #     Δ(d₅₀) = (d_source² / (2·d₅₀)) · (1/√N)
        #            = (d_source / d₅₀)² · d₅₀ / (2√N)
        #
        #   Physical interpretation:
        #   - At large α (aberration-limited): d_source/d₅₀ ≪ 1, bars collapse.
        #   - At small α (source-limited):    d_source/d₅₀ ≈ 1, bars are widest.
        #   - At zero current (N→0): Δ(d₅₀)→∞ in relative terms; both the mean
        #     and the fluctuation of σ_apparent vanish with I, so the apparent
        #     source size is zero at zero current — not a finite floor.
        #     The linear propagation breaks down for N<1; bars are only
        #     meaningful when N ≳ 1.
        yerr = None
        dwell_note = ''
        br_on = self.var_br_on.get()
        ss_on = self.var_ss_on.get()
        if dwell_us is not None and br_on and ss_on:
            try:
                br        = float(self.var_br_custom.get())
                phys_size = float(self.var_phys_size.get())
                demag     = float(self.var_demag.get())
                e_kev     = phys['e_kev']
                T_s       = dwell_us * 1e-6          # dwell time in seconds
                e_C       = 1.602e-19                # electron charge (C)
                d50_m     = np.array(d50s)           # d₅₀ in metres
                sigma_m   = (phys_size * 1e-9) / (2.0 * demag)   # apparent source radius (m)
                d_src_m   = sigma_m                  # d_source radius contribution (m)
                # d₅₀ from the quadrature: d₅₀² = d_source² + ...
                # d_source here is the source radius term (σ_apparent), so we
                # use d_src_m directly without extra factor.
                # Current at each α, in amperes:
                v0        = e_kev * 1000.0
                alphas_r  = np.array(alphas) * 1e-3  # mrad → rad
                I_A       = br * v0 * np.pi**2 * alphas_r**2 * sigma_m**2
                # Electrons per pixel at each α:
                N         = I_A * T_s / e_C
                # Guard against N≤0 (disabled brightness or tiny current):
                N_safe    = np.where(N > 0, N, np.inf)
                # Source area contribution to probe quadrature (m²):
                d_src_sq  = d_src_m**2
                # ±1σ half-width on d₅₀ (metres), propagated from Δ(σ²)/σ²=1/√N:
                delta_d50 = (d_src_sq / (2.0 * d50_m)) / np.sqrt(N_safe)
                yerr      = delta_d50 * 1000.0       # convert to pm for plot
                dwell_note = f'  ±1σ shot noise (T={dwell_us:.0f} µs)'
            except Exception:
                yerr = None   # suppress bars silently if any parameter is missing

        if yerr is not None:
            ax_d.errorbar(alphas, d50_pm, yerr=yerr,
                          fmt='o-', color=d50_color, linewidth=2, markersize=5,
                          ecolor=d50_color, elinewidth=1.2, capsize=3, alpha=0.85,
                          label='d\u2085\u2080' + dwell_note)
        else:
            ax_d.plot(alphas, d50_pm, 'o-',
                      color=d50_color, linewidth=2, markersize=5,
                      label='d\u2085\u2080')
        ax_d.axvline(min_alpha, color='red', linewidth=1.4, linestyle='--', alpha=0.8, zorder=3)
        ax_d.plot(min_alpha, min_d50, 'r*', markersize=14, zorder=5)
        ax_d.annotate(
            f'min d\u2085\u2080: {min_d50:.1f} pm\n\u03b1 = {min_alpha:.1f} mrad',
            xy=(min_alpha, min_d50),
            xytext=(min_alpha + 0.06 * a_range, min_d50 + 0.12 * d_range + 0.5),
            arrowprops=dict(arrowstyle='->', color='red', lw=1.4),
            fontsize=9, color='red',
        )
        ax_d.set_ylabel('Probe diameter d\u2085\u2080 (pm)', fontsize=11, color=d50_color)
        ax_d.tick_params(axis='y', labelcolor=d50_color)

        # ── Right axis: Strehl ratio ─────────────────────────────────────────
        ax_s.plot(alphas, strehls, 'D-',
                  color=strehl_color, linewidth=2, markersize=4,
                  label='Strehl ratio')
        # Dashed horizontal line marking the diffraction-limited threshold S = 0.8
        ax_s.axhline(0.8, color=strehl_color, linewidth=1.0, linestyle='--', alpha=0.6)
        ax_s.set_ylabel('Strehl ratio', fontsize=11, color=strehl_color)
        ax_s.tick_params(axis='y', labelcolor=strehl_color)
        ax_s.set_ylim(-0.05, 1.05)   # fixed range so the threshold line is always visible

        # Vertical line at the α with maximum Strehl; annotate distance from min-d50 α
        max_s_i     = int(np.argmax(strehls))
        max_s_alpha = alphas[max_s_i]
        max_s_val   = strehls[max_s_i]
        delta_alpha = max_s_alpha - min_alpha
        ax_d.axvline(max_s_alpha, color=strehl_color, linewidth=1.4,
                     linestyle='--', alpha=0.8, zorder=3)
        sign = '+' if delta_alpha >= 0 else '\u2212'
        ax_d.text(
            max_s_alpha + 0.01 * a_range, min_d50 + 0.70 * d_range,
            f'max S = {max_s_val:.3f}\n\u03b1 = {max_s_alpha:.1f} mrad\n'
            f'({sign}{abs(delta_alpha):.1f} mrad from min d\u2085\u2080)',
            color=strehl_color, fontsize=8, va='bottom',
        )

        # ── Dose-limited resolution curves ──────────────────────────────────
        # Computed only when brightness, source size, dwell, and contrast are
        # all valid.  See _compute_dose_limited_pm for full derivation.
        #
        # d_dose(α) = 5 · d₅₀(α) / (C · √N(α))
        #   where N(α) = I(α) · T / e,  I(α) = B_r·V₀·π²·α²·σ²_apparent
        #
        # d_eff(α) = √(d₅₀(α)² + d_dose(α)²)
        #   — quadrature sum; the effective resolution a user would measure
        #     in an image, accounting for both optics and dose.
        #
        # Crossover: where d_dose = d₅₀, i.e. N = (5/C)².
        #   Left of crossover  → dose-limited  (statistics dominate)
        #   Right of crossover → optics-limited (probe quality dominates)
        dose_color = '#c0392b'   # red-orange, distinct from steelblue/purple
        eff_color  = '#e67e22'   # amber for d_eff
        dose_curves_drawn = False
        if dwell_us is not None and contrast is not None:
            br_on = self.var_br_on.get()
            ss_on = self.var_ss_on.get()
            if br_on and ss_on:
                try:
                    br        = float(self.var_br_custom.get())
                    phys_size = float(self.var_phys_size.get())
                    demag     = float(self.var_demag.get())
                    e_kev_val = phys['e_kev']
                    v0        = e_kev_val * 1000.0
                    sigma_m   = phys_size * 1e-9 / (2.0 * demag)
                    alphas_r  = np.array(alphas) * 1e-3
                    I_arr     = br * v0 * np.pi**2 * alphas_r**2 * sigma_m**2
                    T_s       = dwell_us * 1e-6
                    e_C       = 1.602e-19
                    N_arr     = I_arr * T_s / e_C        # electrons per pixel per dwell
                    SNR_TH    = self._get_rose_k()
                    d50_arr   = np.array(d50s) * 1e-9   # nm → m
                    dx_m      = phys['dx'] * 1e-9        # pixel size in m
                    # Scale to probe footprint: N_probe = N_pixel · (π/4) · (d₅₀/dx)²
                    # π/4 converts square-pixel count to disk count (d50 is a diameter).
                    # d_dose diverges as N→0, correctly encoding that
                    # zero current means infinite effective resolution limit.
                    N_probe_arr = N_arr * (np.pi / 4.0) * (d50_arr / dx_m) ** 2
                    N_safe    = np.where(N_probe_arr > 0, N_probe_arr, np.nan)
                    d_dose_arr = SNR_TH * d50_arr / (contrast * np.sqrt(N_safe))
                    d_eff_arr  = np.sqrt(d50_arr**2 + d_dose_arr**2)
                    d_dose_pm  = d_dose_arr * 1e12
                    d_eff_pm   = d_eff_arr  * 1e12

                    ax_d.plot(alphas, d_dose_pm, '--',
                              color=dose_color, linewidth=1.8,
                              label=f'd\u209a\u209c\u2099\u2099\u2091 (C={contrast:.2f}, SNR\u2265{SNR_TH:.4g})')
                    ax_d.plot(alphas, d_eff_pm, '-',
                              color=eff_color, linewidth=2,
                              label='d\u2091\u2091\u2091 = \u221a(d\u2085\u2080\u00b2+d\u209a\u209c\u2099\u2099\u2091\u00b2)')

                    # Mark the crossover α where d_dose = d₅₀, i.e. N = (5/C)².
                    # Left of this point the image is statistics-limited;
                    # right of it the probe quality is the binding constraint.
                    N_cross = (SNR_TH / contrast) ** 2
                    cross_mask = N_probe_arr >= N_cross
                    if cross_mask.any() and not cross_mask.all():
                        # Find the first α index where N transitions above N_cross
                        cross_i = int(np.argmax(cross_mask))
                        cross_alpha = alphas[cross_i]
                        ax_d.axvline(cross_alpha, color=dose_color,
                                     linewidth=1.2, linestyle=':', alpha=0.8)
                        ax_d.text(
                            cross_alpha + 0.01 * a_range,
                            min_d50 + 0.45 * d_range,
                            f'dose\u2192optics\n\u03b1={cross_alpha:.1f} mrad\nN={N_cross:.0f} e\u207b',
                            color=dose_color, fontsize=7, va='bottom')
                    dose_curves_drawn = True
                except (ValueError, ZeroDivisionError, TypeError):
                    pass   # missing or invalid parameter → skip dose curves

        # ── Williams & Carter geometric theory overlay ───────────────────────
        # Drawn only when the "Theory overlay" radio is not 'off'.
        # Computed instantly from current GUI state — no FFT sweep required.
        # W&C model: d_geo = √(d_diff² + d_s² + d_c² + d_g²)  (ray-optics).
        # Caveat: geometric theory overestimates probe size at small aberrations
        # (high Strehl); the wave-optical d₅₀ is the more rigorous result.
        wc_mode = self.var_wc_overlay.get()
        if wc_mode != 'off':
            wc = self._calc_wc_geometric(alphas)
            if wc is not None:
                wc_color   = '#e07b00'   # orange — distinct from steelblue/purple/red
                comp_alpha = 0.55        # transparency for individual component lines

                ax_d.plot(alphas, wc['total'], '-',
                          color=wc_color, linewidth=2.2, linestyle=(0, (5, 2)),
                          label='W&C geometric total', zorder=4)

                if wc_mode == 'components':
                    comp_styles = [
                        ('diff', 'd\u209a\u209c\u2099\u2099 = 1.22\u03bb/\u03b1',
                         '#2196F3', (0, (3, 2))),
                        ('s',    'd_s = \u00bdC\u209b\u03b1\u00b3',
                         '#9C27B0', (0, (1, 2))),
                        ('c',    'd_c = C\u1d04\u00b7(\u03c3\u2090E/E)\u00b7f\u209c\u00b7\u03b1',
                         '#4CAF50', (0, (4, 2, 1, 2))),
                        ('g',    'd_g = FWHM/M',
                         '#FF5722', (0, (2, 2))),
                    ]
                    for key, lbl, col, ls in comp_styles:
                        arr = wc[key]
                        # Skip components that are identically zero (disabled)
                        if np.any(arr > 0):
                            ax_d.plot(alphas, arr, linestyle=ls,
                                      color=col, linewidth=1.4, alpha=comp_alpha,
                                      label=lbl, zorder=3)

                # Flag the geometric approximation's limitation on the plot
                ax_d.text(0.01, 0.01,
                          'W&C = geometric (ray-optics) approx.\n'
                          'Overestimates at high Strehl; wave-optic d\u2085\u2080 is rigorous',
                          transform=ax_d.transAxes,
                          fontsize=6.5, color=wc_color, alpha=0.8,
                          va='bottom', ha='left')

        # ── Shared x-axis, title, grid, combined legend ──────────────────────
        e_kev  = phys['e_kev']
        wl     = phys['wl']
        suffix = '  [fast: no focus spread / source size]' if fast else ''
        ax_d.set_title(
            f'Probe resolution vs. convergence angle\n'
            f'E = {e_kev:.0f} keV,  \u03bb = {wl*1000:.4f} pm{suffix}',
            fontsize=10)
        ax_d.set_xlabel('Semi-convergence angle \u03b1 (mrad)', fontsize=11)
        ax_d.set_xlim(alphas[0] * 0.97, alphas[-1] * 1.03)
        ax_d.grid(True, alpha=0.3)

        # Combine legend handles from both axes
        handles = ax_d.get_lines() + ax_s.get_lines()
        labels  = [h.get_label() for h in handles]
        ax_d.legend(handles, labels, fontsize=8, loc='upper right')

        # ── Top x-axis: beam current at each α ───────────────────────────────
        # Only drawn when Gun Brightness and Spatial coherence are both enabled,
        # because current requires a defined apparent source size (σ = size/M).
        try:
            br_on = self.var_br_on.get()
            ss_on = self.var_ss_on.get()
            if br_on and ss_on:
                br        = float(self.var_br_custom.get())
                phys_size = float(self.var_phys_size.get())
                demag     = float(self.var_demag.get())
                e_kev     = phys['e_kev']
                ax_top = ax_d.twiny()
                ax_top.set_xlim(ax_d.get_xlim())
                # Place current ticks at the same α positions as the bottom axis
                bot_ticks = [t for t in ax_d.get_xticks()
                             if alphas[0] <= t <= alphas[-1]]
                curr_labels = []
                for a_t in bot_ticks:
                    i_pA = self._compute_current_pA(br, e_kev, a_t, phys_size, demag)
                    curr_labels.append(f'{i_pA:.2g}' if i_pA is not None else '—')
                ax_top.set_xticks(bot_ticks)
                ax_top.set_xticklabels(curr_labels, fontsize=7)
                ax_top.set_xlabel('Probe current (pA)', fontsize=9, color='#555555')
                ax_top.tick_params(axis='x', labelcolor='#555555', labelsize=7)
        except Exception:
            pass   # brightness not yet configured — skip top axis silently

        self.canvas_res.draw()

        self.res_progress['value'] = 100
        self.var_status.set(
            f'Resolution curve done.  Optimum: d\u2085\u2080 = {min_d50:.1f} pm '
            f'at \u03b1 = {min_alpha:.1f} mrad')
        self.notebook.select(1)

    def _calc_res_done(self):
        self._res_running = False
        self.progress.stop()
        self._set_buttons('normal', 'normal', 'normal')

    # ------------------------------------------------------------------
    # Calculate Resolution vs Defocus  (Tab 3)
    # ------------------------------------------------------------------

    def _on_calc_defocus(self):
        if self._calc_running or self._res_running or self._defoc_running or self._demag_running or self._current_running:
            return
        self._stop_event.clear()
        self._defoc_running = True
        self._set_buttons('disabled', 'disabled', 'disabled')
        try:
            alpha = float(self.var_alpha.get())
        except ValueError:
            alpha = 0.0
        self.var_defoc_alpha_display.set(f'{alpha:.2f} mrad')
        self.var_status.set('Calculating resolution vs defocus…')
        self.progress.start(12)
        self.defoc_progress['value'] = 0
        threading.Thread(target=self._calc_defocus_thread, daemon=True).start()

    def _calc_defocus_thread(self):
        """
        Background thread for the defocus sweep (Tab 3).

        The convergence angle α is fixed (taken from the left panel), so the
        aperture is identical for every step and is built once before the pool
        is created, then passed to every worker to avoid redundant computation.

        The defocus aberration (2,0) is excluded from abers_base; each worker
        injects its own Δf value.  All other aberrations are held fixed.

        The two Scherzer reference defoci (TEM and STEM) are computed here on
        the main process — they depend only on Cs and λ, not on the sweep — and
        are passed through to _update_defoc_plot for annotation on the final
        plot.  The generalised STEM formula uses the current φ from the
        Scherzer panel so that the reference line respects the user's chosen
        phase-error tolerance.

        See _calc_res_thread for a full explanation of the parallelisation
        pattern (ProcessPoolExecutor + as_completed + root.after).
        """
        try:
            phys       = self._get_physics()
            # Build base aberration dict excluding defocus — each worker adds its own Δf
            abers_base = {}
            for (m, n), (en, vx, vy) in self.aber_rows.items():
                if (m, n) == (2, 0):
                    continue   # defocus is the sweep variable; skip here
                if en.get():
                    try:
                        abers_base[(m, n)] = (float(vx.get()), float(vy.get()))
                    except ValueError:
                        pass

            df_min   = float(self.var_dfmin.get())
            df_max   = float(self.var_dfmax.get())
            df_steps = max(2, int(self.var_dfsteps.get()))
            fast     = self.var_fast_defoc.get()

            defoci   = np.linspace(df_min, df_max, df_steps)
            num_k    = phys['num_k']
            dx       = phys['dx']
            wl       = phys['wl']
            alpha    = phys['alpha']
            fs       = None if fast else phys['fs']
            ss       = None if fast else phys['ss']
            akx, aky, dk = self._build_kgrid(num_k, dx)
            k_alpha  = 1e-3 * alpha / wl
            # Aperture is fixed for all defocus steps — build it once
            aperture = get_aperture(k_alpha, dk, akx, aky)

            cs  = self._get_cs_nm()
            try:
                phi_sweep = float(self.var_phi_max.get())
            except Exception:
                phi_sweep = -np.pi / 2
            # Compute Scherzer reference lines for plot annotation (not used in calculation)
            df_scherzer_tem  = calc_scherzer_tem(cs, wl) if cs is not None else None
            df_scherzer_stem = (calc_scherzer_stem_general(cs, wl, phi_sweep)
                                if cs is not None else None)

            edge_mult = self._get_edge_mult()
            args = [(df, abers_base, aperture, akx, aky, wl, dx, fs, ss, edge_mult)
                    for df in defoci]
            d50s    = [None] * len(defoci)
            strehls = [None] * len(defoci)   # parallel list for exact Strehl values
            completed = 0
            stopped = False
            with ProcessPoolExecutor() as pool:
                futures = {pool.submit(_worker_defocus, arg): i
                           for i, arg in enumerate(args)}
                for future in as_completed(futures):
                    if self._stop_event.is_set():
                        for f in futures:
                            f.cancel()
                        stopped = True
                        break
                    i = futures[future]
                    d50s[i], strehls[i] = future.result()
                    completed += 1
                    pct = completed / len(defoci) * 100
                    self.root.after(0, lambda p=pct: self.defoc_progress.configure(value=p))

            if not stopped:
                self.root.after(0, self._update_defoc_plot,
                                 defoci, d50s, strehls, phys, fast,
                                 df_scherzer_tem, df_scherzer_stem)
            else:
                self.root.after(0, lambda: self.var_status.set('Defocus sweep stopped.'))
        except Exception as exc:
            self.root.after(0, self._show_error, str(exc))
        finally:
            self.root.after(0, self._calc_defocus_done)

    def _update_defoc_plot(self, defoci, d50s, strehls, phys, fast,
                            df_scherzer_tem, df_scherzer_stem):
        self._defoc_last_data = (defoci, d50s, strehls, phys, fast,
                                  df_scherzer_tem, df_scherzer_stem)
        try:
            _k = max(1, int(self.var_smooth_defoc.get()))
        except (ValueError, tk.TclError):
            _k = 1
        defoci  = np.array(defoci,  dtype=float)
        d50s    = np.array(d50s,    dtype=float)
        strehls = np.array(strehls, dtype=float)
        if _k > 1:
            _kernel  = np.ones(_k) / _k
            d50s     = np.convolve(d50s,    _kernel, mode='valid')
            strehls  = np.convolve(strehls, _kernel, mode='valid')
            _cl = (_k - 1) // 2;  _cr = _k - 1 - _cl
            defoci = defoci[_cl : len(defoci) - _cr] if _cr > 0 else defoci[_cl:]

        # Clear the primary axis and remove any previously twinned axis so that
        # re-running the sweep doesn't stack additional right-hand axes.
        self.ax_defoc.clear()
        for twin in self.ax_defoc.get_shared_x_axes().get_siblings(self.ax_defoc):
            if twin is not self.ax_defoc:
                twin.remove()

        # ── Left axis: d50 (probe resolution) ───────────────────────────────
        ax_d = self.ax_defoc             # left  = d50
        ax_s = self.ax_defoc.twinx()    # right = Strehl

        d50_color    = 'darkorange'
        strehl_color = '#7b2d8b'   # purple — distinct from darkorange

        d50_pm  = [d * 1000 for d in d50s]
        f_range = defoci[-1] - defoci[0] if defoci[-1] != defoci[0] else 1.0
        d_range = max(d50_pm) - min(d50_pm) if max(d50_pm) != min(d50_pm) else 1.0
        min_i   = int(np.argmin(d50s))
        min_df  = defoci[min_i]
        min_d50 = d50_pm[min_i]

        ax_d.plot(defoci, d50_pm, 's-',
                  color=d50_color, linewidth=2, markersize=5,
                  label='d\u2085\u2080')
        ax_d.axvline(min_df, color='red', linewidth=1.4, linestyle='--', alpha=0.8, zorder=3)
        ax_d.plot(min_df, min_d50, 'r*', markersize=14, zorder=5)
        ax_d.annotate(
            f'min d\u2085\u2080: {min_d50:.1f} pm\n\u0394f = {min_df:.2f} nm',
            xy=(min_df, min_d50),
            xytext=(min_df + 0.06 * f_range, min_d50 + 0.12 * d_range + 0.5),
            arrowprops=dict(arrowstyle='->', color='red', lw=1.4),
            fontsize=9, color='red',
        )
        ax_d.set_ylabel('Probe diameter d\u2085\u2080 (pm)', fontsize=11, color=d50_color)
        ax_d.tick_params(axis='y', labelcolor=d50_color)

        # ── Right axis: Strehl ratio ─────────────────────────────────────────
        ax_s.plot(defoci, strehls, 'D-',
                  color=strehl_color, linewidth=2, markersize=4,
                  label='Strehl ratio')
        # Dashed horizontal line marking the diffraction-limited threshold S = 0.8
        ax_s.axhline(0.8, color=strehl_color, linewidth=1.0, linestyle='--', alpha=0.6)
        ax_s.set_ylabel('Strehl ratio', fontsize=11, color=strehl_color)
        ax_s.tick_params(axis='y', labelcolor=strehl_color)
        ax_s.set_ylim(-0.05, 1.05)   # fixed range so the threshold line is always visible

        # Vertical line at the defocus with maximum Strehl; annotate distance from min-d50
        max_s_i   = int(np.argmax(strehls))
        max_s_df  = defoci[max_s_i]
        max_s_val = strehls[max_s_i]
        delta_df  = max_s_df - min_df
        ax_d.axvline(max_s_df, color=strehl_color, linewidth=1.4,
                     linestyle='--', alpha=0.8, zorder=3)
        sign = '+' if delta_df >= 0 else '\u2212'
        ax_d.text(
            max_s_df + 0.01 * f_range, min_d50 + 0.70 * d_range,
            f'max S = {max_s_val:.3f}\n\u0394f = {max_s_df:.2f} nm\n'
            f'({sign}{abs(delta_df):.2f} nm from min d\u2085\u2080)',
            color=strehl_color, fontsize=8, va='bottom',
        )

        # ── Scherzer reference vertical lines (drawn on d50 axis so they
        #    span the full plot height consistently) ──────────────────────────
        if df_scherzer_tem is not None:
            ax_d.axvline(df_scherzer_tem, color='#8b1a1a',
                         linewidth=1.6, linestyle='--',
                         label=f'\u0394f\u209b(TEM) = {df_scherzer_tem:.3f} nm')
            ax_d.text(
                df_scherzer_tem + 0.01 * f_range, min_d50 + 0.55 * d_range,
                f'\u0394f\u209b(TEM)\n{df_scherzer_tem:.3f} nm',
                color='#8b1a1a', fontsize=8, va='bottom')

        if df_scherzer_stem is not None:
            ax_d.axvline(df_scherzer_stem, color='#1a6e1a',
                         linewidth=1.6, linestyle='-.',
                         label=f'\u0394f\u209b(STEM) = {df_scherzer_stem:.3f} nm')
            ax_d.text(
                df_scherzer_stem + 0.01 * f_range, min_d50 + 0.25 * d_range,
                f'\u0394f\u209b(STEM)\n{df_scherzer_stem:.3f} nm',
                color='#1a6e1a', fontsize=8, va='bottom')

        # ── Shared x-axis, title, grid, combined legend ──────────────────────
        e_kev  = phys['e_kev']
        wl     = phys['wl']
        alpha  = phys['alpha']
        suffix = '  [fast: no focus spread / source size]' if fast else ''
        ax_d.set_title(
            f'Probe resolution vs. defocus\n'
            f'E = {e_kev:.0f} keV,  \u03bb = {wl*1000:.4f} pm,  '
            f'\u03b1 = {alpha:.1f} mrad{suffix}',
            fontsize=10)
        ax_d.set_xlabel('Defocus \u0394f  (nm)', fontsize=11)
        ax_d.grid(True, alpha=0.3)

        # Combine legend handles from both axes
        handles = ax_d.get_lines() + ax_s.get_lines()
        labels  = [h.get_label() for h in handles]
        ax_d.legend(handles, labels, fontsize=8, loc='upper right')

        self.canvas_defoc.draw()

        self.defoc_progress['value'] = 100
        stem_note = (f'  \u0394f\u209b(STEM) = {df_scherzer_stem:.3f} nm'
                     if df_scherzer_stem is not None else '')
        self.var_status.set(
            f'Defocus sweep done.  Optimum: d\u2085\u2080 = {min_d50:.1f} pm '
            f'at \u0394f = {min_df:.2f} nm  (\u03b1 = {alpha:.1f} mrad){stem_note}')
        self.notebook.select(2)

    def _calc_defocus_done(self):
        self._defoc_running = False
        self.progress.stop()
        self._set_buttons('normal', 'normal', 'normal')

    # ------------------------------------------------------------------
    # Right panel — Tab 4: Probe size vs Demagnification
    # ------------------------------------------------------------------

    def _build_demag_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=' d\u2085\u2080 vs Demagnification ')

        ctrl = ttk.LabelFrame(
            tab,
            text='Demagnification sweep  (uses \u03b1, aberrations, Cc/\u03b4E from left panel)',
            padding=6)
        ctrl.pack(fill=tk.X, padx=8, pady=4)

        ttk.Label(ctrl,
                  text='Sweeps the demagnification factor M while holding all other '
                       'parameters fixed.  Apparent source size \u03c3 = Physical Size / M.  '
                       'Strehl ratio is independent of M (coherent wavefront only) and '
                       'appears as a horizontal reference line on the right y-axis.',
                  style='Info.TLabel', wraplength=900, justify=tk.LEFT).pack(
            anchor='w', pady=(0, 4))

        param_row = ttk.Frame(ctrl)
        param_row.pack(fill=tk.X)

        def _lbl(text): ttk.Label(param_row, text=text).pack(side=tk.LEFT, padx=(8, 2))
        def _ent(var, width=6):
            ttk.Entry(param_row, textvariable=var, width=width).pack(side=tk.LEFT)

        _lbl('M min:');    self.var_demag_min  = tk.StringVar(value='10');  _ent(self.var_demag_min)
        _lbl('M max:');    self.var_demag_max  = tk.StringVar(value='100'); _ent(self.var_demag_max)
        _lbl('Step mult:'); self.var_demag_mult = tk.StringVar(value='2.0'); _ent(self.var_demag_mult, 5)

        self.var_fast_demag = tk.BooleanVar(value=False)
        ttk.Checkbutton(param_row, variable=self.var_fast_demag,
                        text='Fast (skip focus spread)').pack(side=tk.LEFT, padx=10)

        ttk.Separator(param_row, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Label(param_row, text='Scale:').pack(side=tk.LEFT)
        self.var_demag_scale = tk.StringVar(value='log')
        for _lbl2, _val in (('Linear', 'linear'), ('Log\u2081\u2080', 'log')):
            ttk.Radiobutton(param_row, text=_lbl2, variable=self.var_demag_scale,
                            value=_val).pack(side=tk.LEFT, padx=2)
        self.var_demag_scale.trace_add('write', self._on_demag_smooth_change)

        ttk.Separator(param_row, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Label(param_row, text='Moving avg:').pack(side=tk.LEFT)
        self.var_smooth_demag = tk.StringVar(value='1')
        for _k in ('1', '2', '3', '4'):
            ttk.Radiobutton(param_row, text=_k, variable=self.var_smooth_demag,
                            value=_k).pack(side=tk.LEFT, padx=2)
        self.var_smooth_demag.trace_add('write', self._on_demag_smooth_change)

        self.btn_demag = ttk.Button(param_row, text='Run Sweep',
                                     style='Accent.TButton',
                                     command=self._on_calc_demag)
        self.btn_demag.pack(side=tk.LEFT, padx=10)

        self.demag_progress = ttk.Progressbar(ctrl, mode='determinate', length=400)
        self.demag_progress.pack(fill=tk.X, pady=4)

        self.fig_demag = Figure(figsize=(10, 5), constrained_layout=True)
        self.ax_demag  = self.fig_demag.add_subplot(1, 1, 1)
        self.ax_demag.set_xlabel('Demagnification M', fontsize=11)
        self.ax_demag.set_ylabel('Probe diameter d\u2085\u2080 (pm)', fontsize=11)
        self.ax_demag.set_title('Probe size vs. demagnification')
        self.ax_demag.grid(True, alpha=0.3)

        canvas = FigureCanvasTkAgg(self.fig_demag, master=tab)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        NavigationToolbar2Tk(canvas, tab).pack(fill=tk.X)
        self.canvas_demag = canvas

    def _on_calc_demag(self):
        if self._calc_running or self._res_running or self._defoc_running or self._demag_running or self._current_running:
            return
        self._stop_event.clear()
        self._demag_running = True
        self._set_buttons('disabled', 'disabled', 'disabled', 'disabled')
        self.var_status.set('Calculating probe size vs demagnification…')
        self.progress.start(12)
        self.demag_progress['value'] = 0
        threading.Thread(target=self._calc_demag_thread, daemon=True).start()

    def _calc_demag_thread(self):
        """
        Background thread for the demagnification sweep (Tab 4).

        The convergence angle α and all aberrations are fixed, so the aperture
        and the coherent phase plate are both identical for every step.  Only the
        apparent source size σ = physical_size / M varies.

        Because Strehl is derived from the phase plate alone, it is constant
        across the sweep.  Workers still compute it per-step for consistency;
        the plot draws it as a horizontal line whose y-value equals the
        wavefront quality at the chosen α and aberration set.
        """
        try:
            phys       = self._get_physics()
            abers      = self._get_aberrations()
            demag_min  = float(self.var_demag_min.get())
            demag_max  = float(self.var_demag_max.get())
            fast       = self.var_fast_demag.get()

            # Geometric steps: demag_min, demag_min·mult, demag_min·mult², …, ≤ demag_max
            try:
                _mult = max(float(self.var_demag_mult.get()), 1.001)
            except ValueError:
                _mult = 2.0
            _m = demag_min
            _mlist = []
            while _m <= demag_max * 1.0001:
                _mlist.append(_m)
                _m *= _mult
            demags = np.array(_mlist if len(_mlist) >= 2 else [demag_min, demag_max],
                              dtype=float)
            num_k    = phys['num_k']
            dx       = phys['dx']
            wl       = phys['wl']
            alpha    = phys['alpha']
            fs       = None if fast else phys['fs']
            # Physical source size from the left-panel entry (not the apparent size)
            try:
                phys_size = float(self.var_phys_size.get())
            except ValueError:
                phys_size = 1.0

            akx, aky, dk = self._build_kgrid(num_k, dx)
            k_alpha  = 1e-3 * alpha / wl
            aperture = get_aperture(k_alpha, dk, akx, aky)

            edge_mult = self._get_edge_mult()
            args = [(m, phys_size, abers, aperture, akx, aky, wl, dx, fs, edge_mult)
                    for m in demags]
            n_steps = len(demags)
            d50s    = [None] * n_steps
            strehls = [None] * n_steps
            completed = 0
            stopped = False
            with ProcessPoolExecutor() as pool:
                futures = {pool.submit(_worker_demag, arg): i
                           for i, arg in enumerate(args)}
                for future in as_completed(futures):
                    if self._stop_event.is_set():
                        for f in futures:
                            f.cancel()
                        stopped = True
                        break
                    i = futures[future]
                    d50s[i], strehls[i] = future.result()
                    completed += 1
                    pct = completed / n_steps * 100
                    self.root.after(0, lambda p=pct: self.demag_progress.configure(value=p))

            if not stopped:
                self.root.after(0, self._update_demag_plot,
                                 demags, d50s, strehls, phys, fast, phys_size)
            else:
                self.root.after(0, lambda: self.var_status.set('Demag sweep stopped.'))
        except Exception as exc:
            self.root.after(0, self._show_error, str(exc))
        finally:
            self.root.after(0, self._calc_demag_done)

    def _update_demag_plot(self, demags, d50s, strehls, phys, fast, phys_size):
        """
        Plot demagnification M (left y-axis) and probe current I (right y-axis)
        against probe size d₅₀ on the x-axis.  Scale radio applies to both y-axes.
        """
        self._demag_last_data = (demags, d50s, strehls, phys, fast, phys_size)
        try:
            _k = max(1, int(self.var_smooth_demag.get()))
        except (ValueError, tk.TclError):
            _k = 1
        demags  = np.array(demags,  dtype=float)
        d50s    = np.array(d50s,    dtype=float)
        strehls = np.array(strehls, dtype=float)
        if _k > 1:
            _kernel = np.ones(_k) / _k
            d50s    = np.convolve(d50s,    _kernel, mode='valid')
            strehls = np.convolve(strehls, _kernel, mode='valid')
            _cl = (_k - 1) // 2;  _cr = _k - 1 - _cl
            demags = demags[_cl : len(demags) - _cr] if _cr > 0 else demags[_cl:]

        self.ax_demag.clear()
        for twin in self.ax_demag.get_shared_x_axes().get_siblings(self.ax_demag):
            if twin is not self.ax_demag:
                twin.remove()
        for twin in self.ax_demag.get_shared_y_axes().get_siblings(self.ax_demag):
            if twin is not self.ax_demag:
                twin.remove()
        for _sax in list(getattr(self.ax_demag, '_secondary_axes', [])):
            try:
                _sax.remove()
            except Exception:
                pass

        ax_d    = self.ax_demag
        d50_pm  = d50s * 1000.0
        min_i   = int(np.argmin(d50_pm))
        min_m   = demags[min_i]
        min_d50 = d50_pm[min_i]
        d50_range = max(d50_pm) - min(d50_pm) if max(d50_pm) != min(d50_pm) else 1.0
        m_range   = max(demags)  - min(demags)  if max(demags) != min(demags)  else 1.0

        m_color = '#1a6e7a'   # teal for M axis
        i_color = '#b85c00'   # orange for current axis

        # ── Left y-axis: M vs d50 ────────────────────────────────────────────
        ax_d.plot(d50_pm, demags, 'o-', color=m_color, linewidth=2,
                  markersize=5, label='M')
        ax_d.plot(min_d50, min_m, 'r*', markersize=14, zorder=5)
        ax_d.annotate(
            f'min d\u2085\u2080: {min_d50:.1f} pm\nM = {min_m:.0f}',
            xy=(min_d50, min_m),
            xytext=(min_d50 + 0.12 * d50_range, min_m - 0.15 * m_range),
            arrowprops=dict(arrowstyle='->', color='red', lw=1.4),
            fontsize=9, color='red')
        ax_d.set_xlabel('Probe diameter d\u2085\u2080  (pm)', fontsize=11)
        ax_d.set_ylabel('Demagnification  M', fontsize=11, color=m_color)
        ax_d.tick_params(axis='y', labelcolor=m_color)

        # y-axis scale (linear or log) — applied to both left and right axes
        try:
            y_scale = self.var_demag_scale.get()
        except Exception:
            y_scale = 'linear'
        ax_d.set_yscale(y_scale)
        # Largest M at bottom, smallest M at top — current axis mirrors automatically
        ax_d.invert_yaxis()

        # ── Right y-axis: current via secondary_yaxis (exact M↔I correspondence) ──
        # For every M tick on the left axis the right axis shows I = K/M² at the
        # identical vertical position.  No separate curve is plotted; the axis
        # labels themselves carry the current information.
        try:
            if self.var_br_on.get():
                br    = float(self.var_br_custom.get())
                _v0   = phys['e_kev'] * 1000.0          # V
                _a_r  = phys['alpha'] * 1e-3             # rad
                _sf   = phys_size * 1e-9 / 2.0          # apparent-source radius factor (m)
                _K_pA = br * _v0 * np.pi**2 * _a_r**2 * _sf**2 * 1e12  # pA·M²

                def _m_to_i(m_arr):
                    """Forward: left-axis M → right-axis I (pA)."""
                    m_arr = np.asarray(m_arr, dtype=float)
                    with np.errstate(divide='ignore', invalid='ignore'):
                        return np.where(m_arr > 0, _K_pA / m_arr**2, np.nan)

                def _i_to_m(i_arr):
                    """Inverse: right-axis I (pA) → left-axis M."""
                    i_arr = np.asarray(i_arr, dtype=float)
                    with np.errstate(divide='ignore', invalid='ignore'):
                        return np.where(i_arr > 0, np.sqrt(_K_pA / i_arr), np.nan)

                ax_r = ax_d.secondary_yaxis('right', functions=(_m_to_i, _i_to_m))
                ax_r.set_ylabel('Probe current  (pA)', fontsize=11, color=i_color)
                ax_r.tick_params(axis='y', labelcolor=i_color)
        except Exception:
            pass   # brightness not configured — skip silently

        # ── Title, grid, legend ───────────────────────────────────────────────
        e_kev  = phys['e_kev']
        wl     = phys['wl']
        alpha  = phys['alpha']
        suffix = '  [fast: no focus spread]' if fast else ''
        try:
            _mult_lbl = f'  step mult \u00d7{float(self.var_demag_mult.get()):g}'
        except Exception:
            _mult_lbl = ''
        ax_d.set_title(
            f'Demagnification vs. probe size\n'
            f'E = {e_kev:.0f} keV,  \u03bb = {wl*1000:.4f} pm,  '
            f'\u03b1 = {alpha:.1f} mrad,  '
            f'phys. size = {phys_size:.1f} nm{_mult_lbl}{suffix}',
            fontsize=10)
        ax_d.grid(True, alpha=0.3)

        handles = [h for h in ax_d.get_lines() if not h.get_label().startswith('_')]
        ax_d.legend(handles, [h.get_label() for h in handles],
                    fontsize=8, loc='upper right')

        self.canvas_demag.draw()
        self.demag_progress['value'] = 100
        self.var_status.set(
            f'Demag sweep done.  Min d\u2085\u2080 = {min_d50:.1f} pm at M = {min_m:.0f}')
        self.notebook.select(3)

    def _calc_demag_done(self):
        self._demag_running = False
        self.progress.stop()
        self._set_buttons('normal', 'normal', 'normal', 'normal')

    # ------------------------------------------------------------------
    # d50 vs Current (Tab 4) — sweeps demagnification M, plots d50 vs I
    # ------------------------------------------------------------------

    def _build_current_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=' d\u2085\u2080 vs Current ')

        ctrl = ttk.LabelFrame(
            tab,
            text='Current sweep  (varies demagnification M; uses \u03b1, aberrations, '
                 'Cc/\u03b4E, brightness, source size from left panel)',
            padding=6)
        ctrl.pack(fill=tk.X, padx=8, pady=4)

        ttk.Label(ctrl,
                  text='Sweeps M from M start to M stop.  Probe current I \u221d (1/M)\u00b2 '
                       'is plotted on the bottom x-axis; M on the top x-axis.  '
                       'Beam current requires Gun Brightness and Source Size to be enabled.',
                  style='Info.TLabel', wraplength=900, justify=tk.LEFT).pack(
            anchor='w', pady=(0, 4))

        param_row = ttk.Frame(ctrl)
        param_row.pack(fill=tk.X)

        def _lbl(t): ttk.Label(param_row, text=t).pack(side=tk.LEFT, padx=(8, 2))
        def _ent(v, w=6): ttk.Entry(param_row, textvariable=v, width=w).pack(side=tk.LEFT)

        _lbl('M start:');  self.var_cur_mstart = tk.StringVar(value='10');  _ent(self.var_cur_mstart)
        _lbl('M stop:');   self.var_cur_mstop  = tk.StringVar(value='100'); _ent(self.var_cur_mstop)
        _lbl('Step mult:'); self.var_cur_mult  = tk.StringVar(value='2.0'); _ent(self.var_cur_mult, 5)

        self.var_fast_cur = tk.BooleanVar(value=False)
        ttk.Checkbutton(param_row, variable=self.var_fast_cur,
                        text='Fast (skip focus spread)').pack(side=tk.LEFT, padx=10)

        ttk.Separator(param_row, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Label(param_row, text='Scale:').pack(side=tk.LEFT)
        self.var_cur_scale = tk.StringVar(value='log')
        for _lbl2, _val in (('Linear', 'linear'), ('Log\u2081\u2080', 'log')):
            ttk.Radiobutton(param_row, text=_lbl2, variable=self.var_cur_scale,
                            value=_val).pack(side=tk.LEFT, padx=2)
        self.var_cur_scale.trace_add('write', self._on_current_smooth_change)

        ttk.Separator(param_row, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Label(param_row, text='Moving avg:').pack(side=tk.LEFT)
        self.var_smooth_cur = tk.StringVar(value='1')
        for _k in ('1', '2', '3', '4'):
            ttk.Radiobutton(param_row, text=_k, variable=self.var_smooth_cur,
                            value=_k).pack(side=tk.LEFT, padx=2)
        self.var_smooth_cur.trace_add('write', self._on_current_smooth_change)

        self.btn_cur = ttk.Button(param_row, text='Run Sweep',
                                  style='Accent.TButton',
                                  command=self._on_calc_current)
        self.btn_cur.pack(side=tk.LEFT, padx=10)

        self.cur_progress = ttk.Progressbar(ctrl, mode='determinate', length=400)
        self.cur_progress.pack(fill=tk.X, pady=4)

        self.fig_cur = Figure(figsize=(10, 5), tight_layout=True)
        self.ax_cur  = self.fig_cur.add_subplot(1, 1, 1)
        self.ax_cur.set_xlabel('Probe current  (pA)', fontsize=11)
        self.ax_cur.set_ylabel('Probe diameter d\u2085\u2080  (pm)', fontsize=11)
        self.ax_cur.set_title('Probe size vs. current')
        self.ax_cur.grid(True, alpha=0.3)

        canvas = FigureCanvasTkAgg(self.fig_cur, master=tab)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        NavigationToolbar2Tk(canvas, tab).pack(fill=tk.X)
        self.canvas_cur = canvas

    def _on_calc_current(self):
        if self._calc_running or self._res_running or self._defoc_running \
                or self._demag_running or self._current_running:
            return
        self._stop_event.clear()
        self._current_running = True
        self._set_buttons('disabled', 'disabled', 'disabled', 'disabled',
                          current_state='disabled')
        self.var_status.set('Calculating probe size vs current\u2026')
        self.progress.start(12)
        self.cur_progress['value'] = 0
        threading.Thread(target=self._calc_current_thread, daemon=True).start()

    def _calc_current_thread(self):
        """
        Background thread for the current sweep (Tab 4).

        Sweeps demagnification M from M_start to M_stop.  At each step the
        apparent source size σ = phys_size/M changes, which alters both the
        probe d₅₀ (source broadening) and the beam current I ∝ σ².  Results
        are passed to _update_current_plot for the dual-axis display.
        """
        try:
            phys    = self._get_physics()
            abers   = self._get_aberrations()
            m_start = float(self.var_cur_mstart.get())
            m_stop  = float(self.var_cur_mstop.get())
            fast    = self.var_fast_cur.get()

            # Geometric steps: M_start, M_start·mult, M_start·mult², …, ≤ M_stop
            try:
                _mult = max(float(self.var_cur_mult.get()), 1.001)
            except ValueError:
                _mult = 2.0
            _m = m_start
            _mlist = []
            while _m <= m_stop * 1.0001:
                _mlist.append(_m)
                _m *= _mult
            demags = np.array(_mlist if len(_mlist) >= 2 else [m_start, m_stop],
                              dtype=float)
            num_k   = phys['num_k']
            dx      = phys['dx']
            wl      = phys['wl']
            alpha   = phys['alpha']
            fs      = None if fast else phys['fs']
            try:
                phys_size = float(self.var_phys_size.get())
            except ValueError:
                phys_size = 1.0

            akx, aky, dk = self._build_kgrid(num_k, dx)
            k_alpha  = 1e-3 * alpha / wl
            aperture = get_aperture(k_alpha, dk, akx, aky)
            edge_mult = self._get_edge_mult()

            # Reuse _worker_demag: takes (M, phys_size, abers, aperture, akx, aky,
            # wl, dx, fs, edge_mult) and returns (d50_nm, strehl)
            args = [(m, phys_size, abers, aperture, akx, aky, wl, dx, fs, edge_mult)
                    for m in demags]
            n_steps = len(demags)
            d50s    = [None] * n_steps
            strehls = [None] * n_steps
            completed = 0
            stopped   = False

            with ProcessPoolExecutor() as pool:
                futures = {pool.submit(_worker_demag, arg): i
                           for i, arg in enumerate(args)}
                for future in as_completed(futures):
                    if self._stop_event.is_set():
                        for f in futures:
                            f.cancel()
                        stopped = True
                        break
                    i = futures[future]
                    d50s[i], strehls[i] = future.result()
                    completed += 1
                    pct = completed / n_steps * 100
                    self.root.after(0, lambda p=pct: self.cur_progress.configure(value=p))

            if not stopped:
                self.root.after(0, self._update_current_plot,
                                demags, d50s, strehls, phys, fast, phys_size)
            else:
                self.root.after(0, lambda: self.var_status.set('Current sweep stopped.'))
        except Exception as exc:
            self.root.after(0, self._show_error, str(exc))
        finally:
            self.root.after(0, self._calc_current_done)

    def _update_current_plot(self, demags, d50s, strehls, phys, fast, phys_size):
        """
        Plot d₅₀ vs probe current (bottom x-axis) and demagnification M (top x-axis).

        I(M) = B_r·V₀·π²·α²·(phys_size/2M)² is used as the primary x-axis when
        brightness is enabled.  Arrays are sorted by ascending I so the x-axis
        runs conventionally from low to high current (right to left in M).

        The top M axis uses a manual twiny with ticks placed at I-positions
        corresponding to evenly-spaced round M values — avoids the ugly irrational
        labels produced by secondary_xaxis on this nonlinear, direction-inverting
        transform.
        """
        self._current_last_data = (demags, d50s, strehls, phys, fast, phys_size)

        try:
            _k = max(1, int(self.var_smooth_cur.get()))
        except (ValueError, tk.TclError):
            _k = 1
        demags  = np.array(demags,  dtype=float)
        d50s    = np.array(d50s,    dtype=float)
        strehls = np.array(strehls, dtype=float)
        if _k > 1:
            _kernel = np.ones(_k) / _k
            d50s    = np.convolve(d50s,    _kernel, mode='valid')
            strehls = np.convolve(strehls, _kernel, mode='valid')
            _cl = (_k - 1) // 2;  _cr = _k - 1 - _cl
            demags = demags[_cl : len(demags) - _cr] if _cr > 0 else demags[_cl:]

        # Compute I(M) in pA if brightness is configured
        i_pa_arr = None
        _br = None
        _K  = None
        try:
            if self.var_br_on.get():
                _br = float(self.var_br_custom.get())
                _vals = [self._compute_current_pA(
                             _br, phys['e_kev'], phys['alpha'], phys_size, m)
                         for m in demags]
                if all(v is not None and v > 0 for v in _vals):
                    i_pa_arr = np.array(_vals, dtype=float)
                    _v0      = phys['e_kev'] * 1000.0
                    _a_r     = phys['alpha'] * 1e-3
                    _sig_fac = phys_size * 1e-9 / 2.0
                    _K = _br * _v0 * np.pi**2 * _a_r**2 * _sig_fac**2 * 1e12
        except Exception:
            pass

        # Sort all arrays by ascending I (low current at left, conventional axis)
        if i_pa_arr is not None:
            _order   = np.argsort(i_pa_arr)
            i_pa_arr = i_pa_arr[_order]
            d50s     = d50s[_order]
            demags   = demags[_order]
            x_arr    = i_pa_arr
            x_label  = 'Probe current  (pA)'
        else:
            x_arr   = demags
            x_label = 'Demagnification  M  (brightness not set \u2014 I axis unavailable)'

        # Clear axes and any existing twin axes
        self.ax_cur.clear()
        for twin in self.ax_cur.get_shared_x_axes().get_siblings(self.ax_cur):
            if twin is not self.ax_cur:
                twin.remove()
        for twin in self.ax_cur.get_shared_y_axes().get_siblings(self.ax_cur):
            if twin is not self.ax_cur:
                twin.remove()

        ax      = self.ax_cur
        d50_pm  = d50s * 1000.0
        x_range = abs(x_arr[-1] - x_arr[0]) if x_arr[-1] != x_arr[0] else 1.0
        d_range = max(d50_pm) - min(d50_pm) if max(d50_pm) != min(d50_pm) else 1.0
        min_i   = int(np.argmin(d50_pm))
        min_x   = x_arr[min_i]
        min_d50 = d50_pm[min_i]
        min_m   = demags[min_i]

        d50_color  = '#1a6e7a'
        dose_color = '#c0392b'
        eff_color  = '#e67e22'

        ax.plot(x_arr, d50_pm, 'o-', color=d50_color, linewidth=2,
                markersize=5, label='d\u2085\u2080')
        ax.plot(min_x, min_d50, 'r*', markersize=14, zorder=5)
        ax.annotate(
            f'min d\u2085\u2080: {min_d50:.1f} pm\nM = {min_m:.0f}',
            xy=(min_x, min_d50),
            xytext=(min_x + 0.06 * x_range, min_d50 + 0.12 * d_range + 0.5),
            arrowprops=dict(arrowstyle='->', color='red', lw=1.4),
            fontsize=9, color='red')

        # Dose and effective resolution curves
        y_max = float(np.max(d50_pm))   # track max plotted y for ylim
        if i_pa_arr is not None:
            try:
                dwell_us  = float(self.var_dwell_us.get())
                contrast  = float(self.var_contrast.get())
                SNR_TH    = self._get_rose_k()
                dx_m      = phys['dx'] * 1e-9
                e_C       = 1.602e-19
                d50_m     = d50s * 1e-9
                N_pix     = i_pa_arr * 1e-12 * dwell_us * 1e-6 / e_C
                N_probe   = N_pix * (np.pi / 4.0) * (d50_m / dx_m) ** 2
                N_safe    = np.where(N_probe > 0, N_probe, np.nan)
                d_dose    = SNR_TH * d50_m / (contrast * np.sqrt(N_safe))
                d_eff     = np.sqrt(d50_m**2 + d_dose**2) * 1e12
                d_dose_pm = d_dose * 1e12
                ax.plot(x_arr, d_dose_pm, '--', color=dose_color, linewidth=1.8,
                        label=f'd\u209a\u209c\u2099\u2099\u2091 (C={contrast:.2f}, SNR\u2265{SNR_TH:.4g})')
                ax.plot(x_arr, d_eff, '-', color=eff_color, linewidth=2,
                        label='d\u2091\u2091\u2091 = \u221a(d\u2085\u2080\u00b2+d\u209a\u209c\u2099\u2099\u2091\u00b2)')
                finite = np.concatenate([d_dose_pm[np.isfinite(d_dose_pm)],
                                         d_eff[np.isfinite(d_eff)]])
                if len(finite):
                    y_max = max(y_max, float(np.max(finite)))
                # ── Annotate minimum of d_eff curve ──────────────────────────
                _eff_finite = np.where(np.isfinite(d_eff), d_eff, np.inf)
                if np.any(np.isfinite(_eff_finite)):
                    _ei       = int(np.argmin(_eff_finite))
                    _eff_min  = d_eff[_ei]
                    _eff_x    = x_arr[_ei]
                    _eff_d50  = d50_pm[_ei]
                    _eff_delt = _eff_min - _eff_d50   # always ≥ 0
                    ax.plot(_eff_x, _eff_min, '*', color=eff_color,
                            markersize=14, zorder=5)
                    ax.annotate(
                        f'min d\u2091\u2091\u2091: {_eff_min:.1f} pm\n'
                        f'd\u2085\u2080 = {_eff_d50:.1f} pm\n'
                        f'\u0394 = +{_eff_delt:.1f} pm',
                        xy=(_eff_x, _eff_min),
                        xytext=(30, 25), textcoords='offset points',
                        arrowprops=dict(arrowstyle='->', color=eff_color,
                                        lw=1.4),
                        fontsize=9, color=eff_color)
            except (ValueError, ZeroDivisionError, TypeError):
                pass

        # Apply x-axis scale (linear or log); must be set before twiny so xlim is correct
        try:
            x_scale = self.var_cur_scale.get()
        except Exception:
            x_scale = 'linear'
        ax.set_xscale(x_scale)

        # Top x-axis: one tick per actual M data point, placed at its I position.
        # Using the actual demags avoids irrational tick values from linspace rounding.
        if i_pa_arr is not None and _K is not None:
            try:
                i_tick_pos = _K / demags**2   # I value for each data-point M
                m_labels   = [f'{int(round(m))}' for m in demags]
                ax_top = ax.twiny()
                ax_top.set_xscale(x_scale)
                ax_top.set_xlim(ax.get_xlim())
                ax_top.set_xticks(i_tick_pos)
                ax_top.set_xticklabels(m_labels, fontsize=8, rotation=90, va='bottom')
                try:
                    _mult_lbl = float(self.var_cur_mult.get())
                    _mult_str = f'×{_mult_lbl:g}'
                except Exception:
                    _mult_str = ''
                ax_top.set_xlabel(f'Demagnification  M  (step mult {_mult_str})',
                                  fontsize=10)
                ax_top.tick_params(axis='x', length=4)
            except Exception:
                pass

        # y-axis upper limit: a little above the max plotted value
        ax.set_ylim(bottom=0, top=y_max * 1.10)

        # Axis labels, title, grid, legend
        e_kev  = phys['e_kev']
        wl     = phys['wl']
        alpha  = phys['alpha']
        suffix = '  [fast: no focus spread]' if fast else ''
        ax.set_title(
            f'Probe size vs. current (M sweep)\n'
            f'E = {e_kev:.0f} keV,  \u03bb = {wl*1000:.4f} pm,  '
            f'\u03b1 = {alpha:.1f} mrad,  phys. size = {phys_size:.1f} nm{suffix}',
            fontsize=10)
        ax.set_xlabel(x_label, fontsize=11)
        ax.set_ylabel('Probe diameter d\u2085\u2080  (pm)', fontsize=11)
        ax.grid(True, alpha=0.3)

        handles = [h for h in ax.get_lines()
                   if not h.get_label().startswith('_')]
        ax.legend(handles, [h.get_label() for h in handles],
                  fontsize=8, loc='upper left')

        self.canvas_cur.draw()
        self.cur_progress['value'] = 100
        self.var_status.set(
            f'Current sweep done.  Min d\u2085\u2080 = {min_d50:.1f} pm at M = {min_m:.0f}'
            + (f',  I = {min_x:.3g} pA' if i_pa_arr is not None else ''))
        self.notebook.select(4)

    def _calc_current_done(self):
        self._current_running = False
        self.progress.stop()
        self._set_buttons('normal', 'normal', 'normal', 'normal')

    def _on_current_smooth_change(self, *_):
        if self._current_last_data is not None:
            self._update_current_plot(*self._current_last_data)

    # ------------------------------------------------------------------
    # Error
    # ------------------------------------------------------------------

    def _get_edge_mult(self):
        try:
            return float(self.var_d50_blur.get())
        except Exception:
            return 1.0

    def _get_rose_k(self):
        try:
            k = float(self.var_rose_k.get())
            return k if k > 0 else 5.0
        except Exception:
            return 5.0

    def _refresh_dose_display(self, *_):
        if self._last_d50_m is None or self._last_I_A is None:
            return
        dose_result = self._compute_dose_limited_pm(self._last_I_A, self._last_d50_m)
        if dose_result is not None:
            d_dose_pm, d_eff_pm = dose_result
            self.var_d_dose_display.set(f'{d_dose_pm:.1f} pm')
            self.var_d_eff_display.set(f'{d_eff_pm:.1f} pm')
        else:
            self.var_d_dose_display.set('\u2014')
            self.var_d_eff_display.set('\u2014')

    # ------------------------------------------------------------------
    # Tab 5 — d50 contour map (defocus × aperture)
    # ------------------------------------------------------------------

    def _build_map_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=' d\u2085\u2080 Map ')

        # ── Control row 1: ranges, steps, colormap ───────────────────────────
        ctrl = ttk.Frame(tab)
        ctrl.pack(fill=tk.X, padx=8, pady=4)

        ttk.Label(ctrl, text='Δf min (nm):').pack(side=tk.LEFT)
        self.var_map_dfmin = tk.StringVar(value=self.var_dfmin.get())
        ttk.Entry(ctrl, textvariable=self.var_map_dfmin, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Label(ctrl, text='max:').pack(side=tk.LEFT, padx=(6, 0))
        self.var_map_dfmax = tk.StringVar(value=self.var_dfmax.get())
        ttk.Entry(ctrl, textvariable=self.var_map_dfmax, width=7).pack(side=tk.LEFT, padx=2)

        ttk.Label(ctrl, text='  α min (mrad):').pack(side=tk.LEFT, padx=(10, 0))
        self.var_map_amin = tk.StringVar(value=self.var_amin.get())
        ttk.Entry(ctrl, textvariable=self.var_map_amin, width=6).pack(side=tk.LEFT, padx=2)
        ttk.Label(ctrl, text='max:').pack(side=tk.LEFT, padx=(6, 0))
        self.var_map_amax = tk.StringVar(value=self.var_amax.get())
        ttk.Entry(ctrl, textvariable=self.var_map_amax, width=6).pack(side=tk.LEFT, padx=2)

        ttk.Label(ctrl, text='  Steps N×N:').pack(side=tk.LEFT, padx=(10, 0))
        self.var_map_steps = tk.StringVar(value='20')
        ttk.Combobox(ctrl, textvariable=self.var_map_steps, width=5,
                     values=['10', '20', '30', '40', '50', '60', '70', '80', '90', '100',
                             '120', '140', '160', '180', '200']).pack(
                         side=tk.LEFT, padx=2)

        ttk.Label(ctrl, text='  Levels:').pack(side=tk.LEFT, padx=(10, 0))
        self.var_map_levels = tk.StringVar(value='40')
        ttk.Combobox(ctrl, textvariable=self.var_map_levels, width=5,
                     values=['10', '20', '30', '40', '50', '60', '80', '100']).pack(
                         side=tk.LEFT, padx=2)
        self.var_map_levels.trace_add('write', self._on_map_cmap_change)

        ttk.Label(ctrl, text='  Color:').pack(side=tk.LEFT, padx=(10, 0))
        self.var_map_cmap = tk.StringVar(value='viridis')
        for cm in ('viridis', 'plasma', 'inferno', 'coolwarm', 'jet'):
            ttk.Radiobutton(ctrl, text=cm, variable=self.var_map_cmap,
                            value=cm).pack(side=tk.LEFT, padx=2)
        self.var_map_cmap.trace_add('write', self._on_map_cmap_change)

        ttk.Separator(ctrl, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        # Display selector: switch between d50 map and Strehl ratio map
        ttk.Label(ctrl, text='Display:').pack(side=tk.LEFT)
        self.var_map_display = tk.StringVar(value='d50')
        ttk.Radiobutton(ctrl, text='d\u2085\u2080', variable=self.var_map_display,
                        value='d50').pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(ctrl, text='Strehl', variable=self.var_map_display,
                        value='strehl').pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(ctrl, text='d\u2091\u2091\u2091', variable=self.var_map_display,
                        value='deff').pack(side=tk.LEFT, padx=2)
        self.var_map_display.trace_add('write', self._on_map_cmap_change)

        # ── Control row 2: plot mode, smoothing, fast, run ──────────────────
        row2 = ttk.Frame(tab)
        row2.pack(fill=tk.X, padx=8, pady=(0, 4))

        ttk.Label(row2, text='Plot:').pack(side=tk.LEFT)
        self.var_map_mode = tk.StringVar(value='contour')
        ttk.Radiobutton(row2, text='Contour', variable=self.var_map_mode,
                        value='contour').pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(row2, text='3D surface', variable=self.var_map_mode,
                        value='3d').pack(side=tk.LEFT, padx=2)
        self.var_map_mode.trace_add('write', self._on_map_cmap_change)

        ttk.Separator(row2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        ttk.Label(row2, text='2D avg:').pack(side=tk.LEFT)
        self.var_map_smooth = tk.StringVar(value='1')
        for _k in ('1', '2', '3', '4'):
            ttk.Radiobutton(row2, text=_k, variable=self.var_map_smooth,
                            value=_k).pack(side=tk.LEFT, padx=2)
        self.var_map_smooth.trace_add('write', self._on_map_cmap_change)

        ttk.Separator(row2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        # Scherzer reference lines on/off
        ttk.Label(row2, text='Scherzer:').pack(side=tk.LEFT)
        self.var_map_scherzer = tk.StringVar(value='on')
        ttk.Radiobutton(row2, text='On',  variable=self.var_map_scherzer,
                        value='on').pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(row2, text='Off', variable=self.var_map_scherzer,
                        value='off').pack(side=tk.LEFT, padx=2)
        self.var_map_scherzer.trace_add('write', self._on_map_cmap_change)

        ttk.Separator(row2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        # Star marker: respective=current display only, both=all, none=hidden,
        # difference=respective marker + delta-from-Scherzer annotation
        ttk.Label(row2, text='Markers:').pack(side=tk.LEFT)
        self.var_map_markers = tk.StringVar(value='respective')
        for _lbl, _val in (('Respective', 'respective'), ('All', 'both'),
                           ('None', 'none'), ('Difference', 'difference')):
            ttk.Radiobutton(row2, text=_lbl, variable=self.var_map_markers,
                            value=_val).pack(side=tk.LEFT, padx=2)
        self.var_map_markers.trace_add('write', self._on_map_cmap_change)

        ttk.Separator(row2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        self.var_fast_map = tk.BooleanVar(value=False)
        ttk.Checkbutton(row2, variable=self.var_fast_map,
                        text='Fast (skip focus spread & source size)').pack(side=tk.LEFT)
        self.btn_map2 = ttk.Button(row2, text='Run Map',
                                   style='Accent.TButton',
                                   command=self._on_calc_map)
        self.btn_map2.pack(side=tk.LEFT, padx=10)

        self.map_progress = ttk.Progressbar(tab, mode='determinate', length=400)
        self.map_progress.pack(fill=tk.X, padx=8, pady=2)

        # Elapsed / ETA display updated live during calculation
        self.var_map_time = tk.StringVar(value='')
        ttk.Label(tab, textvariable=self.var_map_time,
                  style='Info.TLabel').pack(anchor='w', padx=10)

        # ── Figure ───────────────────────────────────────────────────────────
        fig_map = Figure(figsize=(8, 5), dpi=100)
        self.ax_map = fig_map.add_subplot(111)

        canvas = FigureCanvasTkAgg(fig_map, master=tab)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8, pady=4)
        NavigationToolbar2Tk(canvas, tab)
        self.canvas_map = canvas
        self.fig_map    = fig_map

        canvas.mpl_connect('button_press_event', self._on_map_click)

    def _on_calc_map(self):
        if (self._calc_running or self._res_running or
                self._defoc_running or self._demag_running or
                self._current_running or self._map_running):
            return
        self._stop_event.clear()
        self._map_running = True
        self._set_buttons('disabled', 'disabled', 'disabled', 'disabled', 'disabled')
        self.btn_map2.config(state='disabled')
        self.var_status.set('Calculating d\u2085\u2080 contour map\u2026')
        self.var_map_time.set('Elapsed: 0:00  |  ETA: —')
        self.progress.start(12)
        self.map_progress['value'] = 0
        threading.Thread(target=self._calc_map_thread, daemon=True).start()

    def _calc_map_thread(self):
        try:
            phys  = self._get_physics()
            fast  = self.var_fast_map.get()
            num_k = phys['num_k']
            dx    = phys['dx']
            wl    = phys['wl']
            fs    = None if fast else phys['fs']
            ss    = None if fast else phys['ss']
            edge_mult = self._get_edge_mult()

            # Collect current aberrations (excluding defocus — swept by the map)
            abers_base = {}
            for (m, n), (en, vx, vy) in self.aber_rows.items():
                if (m, n) == (2, 0):
                    continue   # defocus is the x-axis; inject per cell
                if en.get():
                    try:
                        abers_base[(m, n)] = (float(vx.get()), float(vy.get()))
                    except ValueError:
                        pass

            df_min  = float(self.var_map_dfmin.get())
            df_max  = float(self.var_map_dfmax.get())
            a_min   = float(self.var_map_amin.get())
            a_max   = float(self.var_map_amax.get())
            n_steps = max(2, int(self.var_map_steps.get()))

            defoci  = np.linspace(df_min, df_max, n_steps)
            alphas  = np.linspace(a_min,  a_max,  n_steps)
            total   = n_steps * n_steps

            # Build flat list of (df, alpha) cells
            cells = [(df, al) for al in alphas for df in defoci]
            args  = [(df, al, abers_base, num_k, dx, wl, fs, ss, edge_mult)
                     for df, al in cells]

            d50_flat    = [None] * total
            strehl_flat = [None] * total
            completed   = 0
            stopped     = False
            t_start     = time.monotonic()   # wall-clock reference for elapsed/ETA

            def _fmt(seconds):
                """Format seconds as m:ss string."""
                s = max(0, int(seconds))
                return f'{s // 60}:{s % 60:02d}'

            with ProcessPoolExecutor() as pool:
                futures = {pool.submit(_worker_map, arg): i
                           for i, arg in enumerate(args)}
                for future in as_completed(futures):
                    if self._stop_event.is_set():
                        for f in futures:
                            f.cancel()
                        stopped = True
                        break
                    i = futures[future]
                    # Worker now returns (d50, strehl) for each cell
                    d50_flat[i], strehl_flat[i] = future.result()
                    completed += 1
                    pct     = completed / total * 100
                    elapsed = time.monotonic() - t_start
                    # Only show ETA after ≥2 cells so the rate estimate is stable
                    if completed >= 2:
                        rate      = completed / elapsed          # cells per second
                        remaining = (total - completed) / rate  # seconds left
                        time_str  = (f'Elapsed: {_fmt(elapsed)}  |  '
                                     f'ETA: {_fmt(remaining)}')
                    else:
                        time_str = f'Elapsed: {_fmt(elapsed)}  |  ETA: —'
                    self.root.after(0, lambda p=pct: self.map_progress.configure(value=p))
                    self.root.after(0, lambda s=time_str: self.var_map_time.set(s))

            # Record total elapsed for the final status label
            total_elapsed = time.monotonic() - t_start

            if not stopped:
                # Reshape both grids: rows = alpha index, cols = defocus index
                d50_grid    = np.array(d50_flat,    dtype=float).reshape(n_steps, n_steps)
                strehl_grid = np.array(strehl_flat, dtype=float).reshape(n_steps, n_steps)
                self.root.after(0, self._update_map_plot,
                                defoci, alphas, d50_grid, strehl_grid, phys, fast)
            else:
                self.root.after(0, lambda: self.var_status.set('Map calculation stopped.'))
                self.root.after(0, lambda s=f'Elapsed: {_fmt(total_elapsed)}  |  stopped':
                                self.var_map_time.set(s))
        except Exception as exc:
            self.root.after(0, self._show_error, str(exc))
        finally:
            self.root.after(0, self._calc_map_done)

    def _on_map_cmap_change(self, *_):
        if self._map_last_data is not None:
            self._update_map_plot(*self._map_last_data)

    def _update_map_plot(self, defoci, alphas, d50_grid, strehl_grid, phys, fast):
        # Cache the raw grids so colormap / display / smooth changes re-render instantly
        self._map_last_data = (defoci, alphas, d50_grid, strehl_grid, phys, fast)

        # Fully reset the figure each redraw to prevent colorbar axes accumulation
        self.fig_map.clear()
        mode    = self.var_map_mode.get()
        display = self.var_map_display.get()   # 'd50' or 'strehl'
        if mode == '3d':
            self.ax_map = self.fig_map.add_subplot(111, projection='3d')
        else:
            self.ax_map = self.fig_map.add_subplot(111)

        ax   = self.ax_map
        cmap = self.var_map_cmap.get()
        try:
            levels = max(5, int(self.var_map_levels.get()))
        except (ValueError, tk.TclError):
            levels = 40
        try:
            smooth_k = max(1, int(self.var_map_smooth.get()))
        except (ValueError, tk.TclError):
            smooth_k = 1

        # Compute d_eff grid: combines probe size and dose resolution per cell.
        # I(α) = B_r·V₀·π²·α²·σ²  varies with alpha (axis 0 of the grid).
        # N_probe = N_pixel · (π/4) · (d₅₀/dx)²  then d_eff = √(d₅₀² + d_dose²).
        # Computed using the full (un-cropped) alphas so it matches d50_grid shape.
        deff_pm_grid = None
        try:
            if self.var_br_on.get() and self.var_ss_on.get():
                _br       = float(self.var_br_custom.get())
                _ps       = float(self.var_phys_size.get())
                _demag    = float(self.var_demag.get())
                _dwell_us = float(self.var_dwell_us.get())
                _contrast = float(self.var_contrast.get())
                _snr      = self._get_rose_k()
                _dx_m     = phys['dx'] * 1e-9
                _v0       = phys['e_kev'] * 1000.0
                _sigma_m  = _ps * 1e-9 / (2.0 * _demag)
                _T_s      = _dwell_us * 1e-6
                _e_C      = 1.602e-19
                # Current and electrons-per-pixel for each alpha row
                _alphas_r = np.array(alphas) * 1e-3          # (n_alpha,)
                _I        = _br * _v0 * np.pi**2 * _alphas_r**2 * _sigma_m**2
                _N_pix    = _I * _T_s / _e_C                 # (n_alpha,)
                _d50_m    = d50_grid * 1e-9                  # (n_alpha, n_df) nm→m
                # Scale to probe footprint: N_probe = N_pixel · (π/4) · (d₅₀/dx)²
                _N_probe  = _N_pix[:, np.newaxis] * (np.pi / 4.0) * (_d50_m / _dx_m)**2
                _N_safe   = np.where(_N_probe > 0, _N_probe, np.nan)
                _d_dose   = _snr * _d50_m / (_contrast * np.sqrt(_N_safe))
                deff_pm_grid = np.sqrt(_d50_m**2 + _d_dose**2) * 1e12  # pm
        except (ValueError, ZeroDivisionError, TypeError):
            pass   # missing or invalid parameter → leave deff_pm_grid as None
        # Cache d_eff grid for export (None when beam params not available)
        self._deff_last_grid = deff_pm_grid

        # Select the data grid to display; d50/deff in pm, Strehl dimensionless
        if display == 'strehl':
            plot_grid = strehl_grid.copy()
        elif display == 'deff' and deff_pm_grid is not None:
            plot_grid = deff_pm_grid.copy()
        else:
            plot_grid = d50_grid * 1000.0   # nm → pm
            if display == 'deff':
                display = 'd50'   # beam params unavailable — fall back silently

        # Apply separable 2D box filter (mode='valid') — crops coordinate arrays
        # to match, avoiding zero-pad edge artifacts present with mode='same'.
        if smooth_k > 1:
            kernel    = np.ones(smooth_k) / smooth_k
            plot_grid = np.apply_along_axis(
                lambda row: np.convolve(row, kernel, mode='valid'), 1, plot_grid)
            plot_grid = np.apply_along_axis(
                lambda col: np.convolve(col, kernel, mode='valid'), 0, plot_grid)
            _cl = (smooth_k - 1) // 2;  _cr = smooth_k - 1 - _cl
            defoci = defoci[_cl : len(defoci) - _cr] if _cr > 0 else defoci[_cl:]
            alphas = alphas[_cl : len(alphas) - _cr] if _cr > 0 else alphas[_cl:]

        # Store (possibly cropped) coordinate arrays for the click handler
        self._map_defoci = defoci
        self._map_alphas = alphas

        # Locate both extrema in the (smoothed, possibly cropped) grids so that
        # 'both' marker mode can draw whichever is relevant regardless of display.
        # d50 grid must be converted to pm before finding its minimum.
        d50_pm_smooth = d50_grid * 1000.0
        if smooth_k > 1:
            d50_pm_smooth = np.apply_along_axis(
                lambda r: np.convolve(r, np.ones(smooth_k)/smooth_k, mode='valid'),
                1, d50_pm_smooth)
            d50_pm_smooth = np.apply_along_axis(
                lambda c: np.convolve(c, np.ones(smooth_k)/smooth_k, mode='valid'),
                0, d50_pm_smooth)
        strehl_smooth = strehl_grid.copy()
        if smooth_k > 1:
            strehl_smooth = np.apply_along_axis(
                lambda r: np.convolve(r, np.ones(smooth_k)/smooth_k, mode='valid'),
                1, strehl_smooth)
            strehl_smooth = np.apply_along_axis(
                lambda c: np.convolve(c, np.ones(smooth_k)/smooth_k, mode='valid'),
                0, strehl_smooth)
        # Smooth deff grid in the same way (None if beam params unavailable)
        deff_smooth = deff_pm_grid.copy() if deff_pm_grid is not None else None
        if deff_smooth is not None and smooth_k > 1:
            deff_smooth = np.apply_along_axis(
                lambda r: np.convolve(r, np.ones(smooth_k)/smooth_k, mode='valid'),
                1, deff_smooth)
            deff_smooth = np.apply_along_axis(
                lambda c: np.convolve(c, np.ones(smooth_k)/smooth_k, mode='valid'),
                0, deff_smooth)

        # Min-d50 location
        r_d50, c_d50 = np.unravel_index(np.argmin(d50_pm_smooth), d50_pm_smooth.shape)
        min_df, min_al = defoci[c_d50], alphas[r_d50]
        min_d50_val    = d50_pm_smooth[r_d50, c_d50]

        # Max-Strehl location
        r_s, c_s = np.unravel_index(np.argmax(strehl_smooth), strehl_smooth.shape)
        max_s_df, max_s_al = defoci[c_s], alphas[r_s]
        max_s_val          = strehl_smooth[r_s, c_s]

        # Min-d_eff location (only when beam params are available)
        r_deff = c_deff = None
        min_deff_df = min_deff_al = min_deff_val = None
        if deff_smooth is not None:
            r_deff, c_deff   = np.unravel_index(np.argmin(deff_smooth), deff_smooth.shape)
            min_deff_df      = defoci[c_deff]
            min_deff_al      = alphas[r_deff]
            min_deff_val     = deff_smooth[r_deff, c_deff]

        # Primary extremum for status bar and axis annotation (matches current display)
        if display == 'strehl':
            ext_val_label = f'max Strehl = {max_s_val:.3f}'
            ext_df, ext_al = max_s_df, max_s_al
        elif display == 'deff' and min_deff_val is not None:
            ext_val_label = f'min d\u2091\u2091\u2091 = {min_deff_val:.1f} pm'
            ext_df, ext_al = min_deff_df, min_deff_al
        else:
            ext_val_label = (f'min d\u2085\u2080 = {min_d50_val:.1f} pm'
                             f'  (S = {strehl_smooth[r_d50, c_d50]:.3f})')
            ext_df, ext_al = min_df, min_al

        # Scherzer reference values — computed regardless of the on/off toggle so
        # the positions are always available; the toggle controls whether they are drawn.
        show_scherzer = self.var_map_scherzer.get() == 'on'
        markers_mode  = self.var_map_markers.get()   # 'respective','both','none','difference'
        cs = self._get_cs_nm()
        wl = phys['wl']
        df_scherz = None
        al_scherz = None
        if cs is not None:
            try:
                phi_sweep = float(self.var_phi_max.get())
            except Exception:
                phi_sweep = -np.pi / 2
            df_scherz = calc_scherzer_stem_general(cs, wl, phi_sweep)
            al_scherz = calc_optimal_alpha_mrad(cs, wl, phi_sweep)

        if mode == '3d':
            DF, AL = np.meshgrid(defoci, alphas)
            ax.plot_surface(DF, AL, plot_grid, cmap=cmap, linewidth=0,
                            antialiased=True, alpha=0.9)
            ax.set_xlabel('\u0394f  (nm)', fontsize=9, labelpad=6)
            ax.set_ylabel('\u03b1  (mrad)', fontsize=9, labelpad=6)
            if display == 'strehl':
                zlabel = 'Strehl'
            elif display == 'deff':
                zlabel = 'd\u2091\u2091\u2091  (pm)'
            else:
                zlabel = 'd\u2085\u2080  (pm)'
            ax.set_zlabel(zlabel, fontsize=9, labelpad=6)
            # Determine which markers to draw (3D: difference mode not supported,
            # treated as 'respective')
            if markers_mode == 'none':
                draw_d50_star = draw_strehl_star = draw_deff_star = False
            elif markers_mode == 'both':
                draw_d50_star    = True
                draw_strehl_star = True
                draw_deff_star   = (deff_smooth is not None)
            else:   # 'respective' or 'difference' (difference falls back in 3D)
                draw_d50_star    = (display == 'd50')
                draw_strehl_star = (display == 'strehl')
                draw_deff_star   = (display == 'deff' and deff_smooth is not None)
            if draw_d50_star:
                ax.scatter([min_df], [min_al], [plot_grid[r_d50, c_d50]],
                           s=120, color='white', edgecolors='black',
                           linewidths=1.5, zorder=5,
                           label=(f'min d\u2085\u2080 = {min_d50_val:.1f} pm'
                                  f'  (S = {strehl_smooth[r_d50, c_d50]:.3f})'))
            if draw_strehl_star:
                ax.scatter([max_s_df], [max_s_al], [plot_grid[r_s, c_s]],
                           s=120, color='yellow', edgecolors='black',
                           linewidths=1.5, zorder=5,
                           label=f'max Strehl = {max_s_val:.3f}')
            if draw_deff_star:
                ax.scatter([min_deff_df], [min_deff_al], [plot_grid[r_deff, c_deff]],
                           s=120, color='cyan', edgecolors='black',
                           linewidths=1.5, zorder=5,
                           label=f'min d\u2091\u2091\u2091 = {min_deff_val:.1f} pm')
            ax.legend(fontsize=8)
        else:
            # Fixed colour scale for Strehl [0, 1]; auto-scale for d50
            vmin, vmax = (0.0, 1.0) if display == 'strehl' else (None, None)
            cf = ax.contourf(defoci, alphas, plot_grid, levels=levels,
                             cmap=cmap, vmin=vmin, vmax=vmax)
            ax.contour(defoci, alphas, plot_grid, levels=levels,
                       colors='k', linewidths=0.3, alpha=0.4)
            if display == 'strehl':
                cb_label = 'Strehl ratio'
            elif display == 'deff':
                cb_label = 'd\u2091\u2091\u2091 (pm)'
            else:
                cb_label = 'd\u2085\u2080 (pm)'
            self.fig_map.colorbar(cf, ax=ax, label=cb_label,
                                  fraction=0.046, pad=0.04)

            # Diffraction-limited S = 0.8 contour on Strehl map
            if display == 'strehl':
                try:
                    cs08 = ax.contour(defoci, alphas, plot_grid, levels=[0.8],
                                      colors='white', linewidths=1.8, linestyles='--')
                    ax.clabel(cs08, fmt={0.8: 'S = 0.8'}, fontsize=8, inline=True)
                except Exception:
                    pass   # level outside data range — skip silently

            # ── Marker flags ─────────────────────────────────────────────────
            # 'respective': star for the current display only
            # 'both':       all available stars (d50=white, strehl=yellow, deff=cyan)
            # 'none':       no stars
            # 'difference': respective star + line + annotation vs Scherzer point
            if markers_mode == 'none':
                draw_d50_star = draw_strehl_star = draw_deff_star = False
                draw_diff     = False
            elif markers_mode == 'both':
                draw_d50_star    = True
                draw_strehl_star = True
                draw_deff_star   = (deff_smooth is not None)
                draw_diff        = False
            elif markers_mode == 'difference':
                draw_d50_star    = (display == 'd50')
                draw_strehl_star = (display == 'strehl')
                draw_deff_star   = (display == 'deff' and deff_smooth is not None)
                draw_diff        = True
            else:   # 'respective'
                draw_d50_star    = (display == 'd50')
                draw_strehl_star = (display == 'strehl')
                draw_deff_star   = (display == 'deff' and deff_smooth is not None)
                draw_diff        = False

            if draw_d50_star:
                ax.plot(min_df, min_al, marker='*', markersize=16, zorder=6,
                        markerfacecolor='white', markeredgecolor='black',
                        markeredgewidth=1.2, linestyle='none',
                        label=(f'min d\u2085\u2080 = {min_d50_val:.1f} pm'
                               f'  (S = {strehl_smooth[r_d50, c_d50]:.3f},'
                               f' \u0394f={min_df:.1f} nm, \u03b1={min_al:.1f} mrad)'))
            if draw_strehl_star:
                ax.plot(max_s_df, max_s_al, marker='*', markersize=16, zorder=6,
                        markerfacecolor='yellow', markeredgecolor='black',
                        markeredgewidth=1.2, linestyle='none',
                        label=(f'max Strehl = {max_s_val:.3f}'
                               f'  (\u0394f={max_s_df:.1f} nm, \u03b1={max_s_al:.1f} mrad)'))
            if draw_deff_star:
                ax.plot(min_deff_df, min_deff_al, marker='*', markersize=16, zorder=6,
                        markerfacecolor='cyan', markeredgecolor='black',
                        markeredgewidth=1.2, linestyle='none',
                        label=(f'min d\u2091\u2091\u2091 = {min_deff_val:.1f} pm'
                               f'  (\u0394f={min_deff_df:.1f} nm, \u03b1={min_deff_al:.1f} mrad)'))

            # 'difference' mode: draw connecting line from Scherzer point to
            # the display-appropriate optimum, annotated with Δdf and Δα.
            if draw_diff and df_scherz is not None and al_scherz is not None:
                ax.plot(df_scherz, al_scherz, marker='D', markersize=10, zorder=6,
                        markerfacecolor='none', markeredgecolor='white',
                        markeredgewidth=1.5, linestyle='none',
                        label=f'Scherzer (\u0394f={df_scherz:.1f} nm, \u03b1={al_scherz:.1f} mrad)')
                ax.plot([df_scherz, ext_df], [al_scherz, ext_al],
                        'w--', linewidth=1.2, alpha=0.7, zorder=5)
                df_range = defoci[-1] - defoci[0]
                al_range = alphas[-1] - alphas[0]
                ax.annotate(
                    f'\u0394\u0394f = {ext_df - df_scherz:+.1f} nm\n'
                    f'\u0394\u03b1  = {ext_al - al_scherz:+.2f} mrad',
                    xy=(ext_df, ext_al),
                    xytext=(ext_df + 0.04 * df_range, ext_al + 0.04 * al_range),
                    fontsize=8, color='white', zorder=7,
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='black', alpha=0.65))

            # Scherzer STEM defocus — vertical dashed green line (if enabled)
            if show_scherzer:
                if df_scherz is not None and defoci[0] <= df_scherz <= defoci[-1]:
                    ax.axvline(df_scherz, color='#1a6e1a', linestyle='--', linewidth=1.4,
                               label=f'\u0394f Scherzer ({df_scherz:.1f} nm)')
                # Scherzer optimal aperture — horizontal dashed orange line (if enabled)
                if al_scherz is not None and alphas[0] <= al_scherz <= alphas[-1]:
                    ax.axhline(al_scherz, color='#b85c00', linestyle='--', linewidth=1.4,
                               label=f'\u03b1 Scherzer ({al_scherz:.1f} mrad)')

            ax.set_xlabel('\u0394f  (nm)', fontsize=10)
            ax.set_ylabel('\u03b1  (mrad)', fontsize=10)
            ax.legend(fontsize=8, loc='upper right')

        if display == 'strehl':
            map_label = 'Strehl map'
        elif display == 'deff':
            map_label = 'd\u2091\u2091\u2091 map (click to set fields)'
        else:
            map_label = 'd\u2085\u2080 map (click to set fields)'
        title = map_label + ('  [fast]' if fast else '')
        ax.set_title(title, fontsize=10)

        self.fig_map.tight_layout()
        self.canvas_map.draw()
        self.map_progress['value'] = 100
        self.var_status.set(
            f'Map done.  {ext_val_label}  at \u0394f = {ext_df:.1f} nm, \u03b1 = {ext_al:.1f} mrad')
        # Final elapsed shown by _calc_map_thread; update label to 'done'
        current_time = self.var_map_time.get()
        if current_time:
            self.var_map_time.set(current_time.split('|')[0].strip() + '  |  done')
        self.notebook.select(5)

    def _calc_map_done(self):
        self._map_running = False
        self.progress.stop()
        self.btn_map2.config(state='normal')
        self._set_buttons('normal', 'normal', 'normal', 'normal', 'normal')

    def _on_map_click(self, event):
        if self.var_map_mode.get() == '3d':
            return   # 3D axes don't provide reliable xdata/ydata for picking
        if event.inaxes is not self.ax_map:
            return
        if not hasattr(self, '_map_defoci') or self._map_defoci is None:
            return
        # Snap to nearest grid point
        df_click = event.xdata
        al_click = event.ydata
        if df_click is None or al_click is None:
            return
        df_idx = int(np.argmin(np.abs(self._map_defoci - df_click)))
        al_idx = int(np.argmin(np.abs(self._map_alphas - al_click)))
        df_snap = self._map_defoci[df_idx]
        al_snap = self._map_alphas[al_idx]
        # Update left-panel defocus and convergence angle fields
        _, vx, _ = self.aber_rows[(2, 0)]
        vx.set(f'{df_snap:.4f}')
        self.var_alpha.set(f'{al_snap:.2f}')
        self.var_status.set(
            f'Map click \u2192 \u0394f = {df_snap:.2f} nm, \u03b1 = {al_snap:.2f} mrad set in left panel')

    def _on_res_smooth_change(self, *_):
        if self._res_last_data is not None:
            self._update_res_plot(*self._res_last_data)

    def _on_defoc_smooth_change(self, *_):
        if self._defoc_last_data is not None:
            self._update_defoc_plot(*self._defoc_last_data)

    def _on_demag_smooth_change(self, *_):
        if self._demag_last_data is not None:
            self._update_demag_plot(*self._demag_last_data)

    # ------------------------------------------------------------------
    # Export all data to a multi-sheet Excel workbook
    # ------------------------------------------------------------------

    def _export_to_spreadsheet(self):
        """
        Write every calculated dataset to a single .xlsx workbook.

        Sheet layout
        ------------
        Parameters      — all current GUI inputs (energy, aberrations, coherence, …)
        Res_vs_Alpha    — summary stats, then α (mrad), d₅₀ (pm), Strehl [, I (pA)]
        Res_vs_Defocus  — summary stats, then Δf (nm), d₅₀ (pm), Strehl
        Demag_Sweep     — summary stats, then M, d₅₀ (pm), Strehl [, I (pA)]
        Current_Sweep   — summary stats, then M, I (pA), d₅₀ (pm), Strehl
        Map_d50_pm      — summary stats, then 2-D table: rows=α, cols=Δf, values=d₅₀
        Map_Strehl      — summary stats, then same layout with Strehl values
        Map_deff_pm     — summary stats, then same layout with d_eff values (if available)

        Sheets for tabs with no data are silently omitted.
        openpyxl is imported lazily so the rest of the GUI works without it.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                'Missing library',
                'openpyxl is required for Excel export.\n'
                'Install it with:  pip install openpyxl')
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel workbook', '*.xlsx'), ('All files', '*.*')],
            title='Export all data to spreadsheet')
        if not path:
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # discard default empty sheet

        # ── Styling helpers ───────────────────────────────────────────────
        _hdr_fill  = PatternFill(start_color='1A6E7A', end_color='1A6E7A',
                                 fill_type='solid')
        _sum_fill  = PatternFill(start_color='D9EEF0', end_color='D9EEF0',
                                 fill_type='solid')   # pale teal for summary rows
        _hdr_font  = Font(bold=True, color='FFFFFF', size=10)
        _ttl_font  = Font(bold=True, size=10)
        _sum_font  = Font(bold=True, size=9, color='1A4E56')

        def _write_summary(ws, summary_kv):
            """
            Write summary key-value-unit rows at the top of ws with pale teal
            background.  Returns the next available row number after a blank gap.
            summary_kv is a list of (label, value, unit) tuples.
            """
            for r_i, (k, v, u) in enumerate(summary_kv, 1):
                c1 = ws.cell(row=r_i, column=1, value=k)
                c2 = ws.cell(row=r_i, column=2, value=v)
                c3 = ws.cell(row=r_i, column=3, value=u)
                for c in (c1, c2, c3):
                    c.fill = _sum_fill
                c1.font = _sum_font
            return len(summary_kv) + 2   # +1 blank row, then header on next

        def _write_table(ws, headers, rows, start_row=1):
            """Styled header row at start_row followed by data rows."""
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=c, value=h)
                cell.font = _hdr_font
                cell.fill = _hdr_fill
                cell.alignment = Alignment(horizontal='center')
            for r_i, row in enumerate(rows, start_row + 1):
                for c_i, val in enumerate(row, 1):
                    ws.cell(row=r_i, column=c_i, value=val)
            for col_cells in ws.columns:
                max_w = max(
                    (len(str(cell.value or '')) for cell in col_cells),
                    default=8)
                ws.column_dimensions[
                    col_cells[0].column_letter].width = max(max_w + 2, 12)

        def _write_kv(ws, kv_list):
            """Key / Value / Unit rows for the Parameters sheet."""
            for c, h in enumerate(['Parameter', 'Value', 'Unit'], 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = _ttl_font
            for r_i, (k, v, u) in enumerate(kv_list, 2):
                ws.cell(row=r_i, column=1, value=k)
                ws.cell(row=r_i, column=2, value=v)
                ws.cell(row=r_i, column=3, value=u)
            ws.column_dimensions['A'].width = 38
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 16

        # ── Reverse-lookup: (m,n) → Krivanek label ───────────────────────
        _MN_TO_LABEL = {v: k for k, v in KRIVANEK_ABBR_TO_MN.items()}

        # ── Sheet 1: Parameters ───────────────────────────────────────────
        ws_par = wb.create_sheet(title='Parameters')
        kv = []

        def _p(name, var, unit=''):
            try:    val = float(var.get())
            except Exception: val = var.get()
            kv.append((name, val, unit))

        def _pb(name, var):
            kv.append((name, 'Yes' if var.get() else 'No', ''))

        kv.append(('--- Beam / Optics ---', '', ''))
        _p('Beam energy',              self.var_energy,    'keV')
        _p('Convergence half-angle',   self.var_alpha,     'mrad')
        _p('Pixel size dx',            self.var_dx,        'nm')
        _p('Grid size N_k',            self.var_numk,      'pixels')
        kv.append(('', '', ''))
        kv.append(('--- Temporal Coherence ---', '', ''))
        _pb('Temporal coherence on',   self.var_fs_on)
        _p('Chromatic coeff. Cc',      self.var_cc,        'nm')
        _p('Energy spread dE',         self.var_dE,        'eV')
        kv.append(('', '', ''))
        kv.append(('--- Spatial Coherence / Source ---', '', ''))
        _pb('Source size on',          self.var_ss_on)
        _p('Physical source size',     self.var_phys_size, 'nm')
        _p('Demagnification M',        self.var_demag,     '')
        kv.append(('', '', ''))
        kv.append(('--- Gun Brightness ---', '', ''))
        _pb('Brightness on',           self.var_br_on)
        _p('Reduced brightness Br',    self.var_br_custom, 'A/(m2 sr V)')
        kv.append(('', '', ''))
        kv.append(('--- Dose / SNR ---', '', ''))
        _p('Dwell time',               self.var_dwell_us,  'us')
        _p('Contrast',                 self.var_contrast,  '')
        _p('Rose criterion k',         self.var_rose_k,    '')
        kv.append(('', '', ''))
        kv.append(('--- Aberrations (enabled only) ---', '', ''))
        for key, (vx, vy) in self._get_aberrations().items():
            label = _MN_TO_LABEL.get(key, f'({key[0]},{key[1]})')
            kv.append((f'{label}  x-component', round(float(vx), 6), 'nm'))
            if float(vy) != 0.0:
                kv.append((f'{label}  y-component', round(float(vy), 6), 'nm'))

        _write_kv(ws_par, kv)

        # ── Sheet 2: Resolution vs α ──────────────────────────────────────
        if self._res_last_data is not None:
            alphas, d50s, strehls, phys, fast, dwell_us, contrast = \
                self._res_last_data
            _d50_pm = np.array(d50s, dtype=float) * 1000.0
            _al     = np.array(alphas, dtype=float)
            _st     = np.array(strehls, dtype=float)
            _mi     = int(np.argmin(_d50_pm))
            _si     = int(np.argmax(_st))
            summ = [
                ('Min d50 (pm)',              round(float(_d50_pm[_mi]), 3), 'pm'),
                ('\u03b1 at min d50 (mrad)',  round(float(_al[_mi]),    4), 'mrad'),
                ('Max Strehl',                round(float(_st[_si]),    5), ''),
                ('\u03b1 at max Strehl (mrad)', round(float(_al[_si]), 4), 'mrad'),
            ]
            hdrs = ['\u03b1 (mrad)', 'd50 (pm)', 'Strehl']
            rows = [[round(float(a), 4), round(float(d) * 1000.0, 3),
                     round(float(s), 5)]
                    for a, d, s in zip(alphas, d50s, strehls)]
            if self.var_br_on.get():
                try:
                    _br = float(self.var_br_custom.get())
                    _ps = float(self.var_phys_size.get())
                    _dm = float(self.var_demag.get())
                    hdrs.append('I (pA)')
                    for i, a in enumerate(alphas):
                        ip = self._compute_current_pA(
                            _br, phys['e_kev'], float(a), _ps, _dm)
                        rows[i].append(round(float(ip), 4) if ip else '')
                except Exception:
                    pass
            ws_ra = wb.create_sheet(title='Res_vs_Alpha')
            sr = _write_summary(ws_ra, summ)
            _write_table(ws_ra, hdrs, rows, start_row=sr)

        # ── Sheet 3: Resolution vs Defocus ────────────────────────────────
        if self._defoc_last_data is not None:
            defoci, d50s, strehls, phys, fast, df_tem, df_stem = \
                self._defoc_last_data
            _d50_pm = np.array(d50s, dtype=float) * 1000.0
            _df     = np.array(defoci, dtype=float)
            _st     = np.array(strehls, dtype=float)
            _mi     = int(np.argmin(_d50_pm))
            _si     = int(np.argmax(_st))
            summ = [
                ('Min d50 (pm)',             round(float(_d50_pm[_mi]), 3), 'pm'),
                ('\u0394f at min d50 (nm)',  round(float(_df[_mi]),    3), 'nm'),
                ('Max Strehl',               round(float(_st[_si]),    5), ''),
                ('\u0394f at max Strehl (nm)', round(float(_df[_si]), 3), 'nm'),
            ]
            if df_tem is not None:
                summ.append(('Scherzer TEM (nm)', round(df_tem, 3), 'nm'))
            if df_stem is not None:
                summ.append(('Scherzer STEM (nm)', round(df_stem, 3), 'nm'))
            hdrs = ['\u0394f (nm)', 'd50 (pm)', 'Strehl']
            rows = [[round(float(df), 3), round(float(d) * 1000.0, 3),
                     round(float(s), 5)]
                    for df, d, s in zip(defoci, d50s, strehls)]
            ws_df = wb.create_sheet(title='Res_vs_Defocus')
            sr = _write_summary(ws_df, summ)
            _write_table(ws_df, hdrs, rows, start_row=sr)

        # ── Sheet 4: Demagnification sweep ────────────────────────────────
        if self._demag_last_data is not None:
            demags, d50s, strehls, phys, fast, phys_size = self._demag_last_data
            _d50_pm = np.array(d50s, dtype=float) * 1000.0
            _mg     = np.array(demags, dtype=float)
            _mi     = int(np.argmin(_d50_pm))
            summ = [
                ('Min d50 (pm)',         round(float(_d50_pm[_mi]), 3), 'pm'),
                ('M at min d50',         round(float(_mg[_mi]),     3), ''),
            ]
            hdrs = ['M', 'd50 (pm)', 'Strehl']
            rows = [[round(float(m), 3), round(float(d) * 1000.0, 3),
                     round(float(s), 5)]
                    for m, d, s in zip(demags, d50s, strehls)]
            if self.var_br_on.get():
                try:
                    _br = float(self.var_br_custom.get())
                    hdrs.append('I (pA)')
                    i_vals = []
                    for i, m in enumerate(demags):
                        ip = self._compute_current_pA(
                            _br, phys['e_kev'], phys['alpha'],
                            float(phys_size), float(m))
                        ip_r = round(float(ip), 4) if ip else ''
                        rows[i].append(ip_r)
                        i_vals.append(ip_r)
                    if i_vals and i_vals[_mi] != '':
                        summ.append(('I at min d50 (pA)',
                                     i_vals[_mi], 'pA'))
                except Exception:
                    pass
            ws_dm = wb.create_sheet(title='Demag_Sweep')
            sr = _write_summary(ws_dm, summ)
            _write_table(ws_dm, hdrs, rows, start_row=sr)

        # ── Sheet 5: Current sweep ────────────────────────────────────────
        if self._current_last_data is not None:
            demags, d50s, strehls, phys, fast, phys_size = \
                self._current_last_data
            _d50_pm = np.array(d50s, dtype=float) * 1000.0
            _mg     = np.array(demags, dtype=float)
            _mi     = int(np.argmin(_d50_pm))
            hdrs = ['M', 'I (pA)', 'd50 (pm)', 'Strehl']
            rows = []
            i_vals = []
            for m, d, s in zip(demags, d50s, strehls):
                ip = ''
                if self.var_br_on.get():
                    try:
                        _br = float(self.var_br_custom.get())
                        ip_v = self._compute_current_pA(
                            _br, phys['e_kev'], phys['alpha'],
                            float(phys_size), float(m))
                        if ip_v:
                            ip = round(float(ip_v), 4)
                    except Exception:
                        pass
                i_vals.append(ip)
                rows.append([round(float(m), 3), ip,
                             round(float(d) * 1000.0, 3),
                             round(float(s), 5)])
            # sort by ascending I for summary (mirror what the plot does)
            summ = [
                ('Min d50 (pm)',    round(float(_d50_pm[_mi]), 3), 'pm'),
                ('M at min d50',    round(float(_mg[_mi]),     3), ''),
            ]
            if i_vals[_mi] != '':
                summ.append(('I at min d50 (pA)', i_vals[_mi], 'pA'))
            # d_eff minimum — recompute from data
            try:
                _dw   = float(self.var_dwell_us.get())
                _con  = float(self.var_contrast.get())
                _snr  = self._get_rose_k()
                _dx_m = phys['dx'] * 1e-9
                _eC   = 1.602e-19
                _ipa  = np.array([v for v in i_vals if v != ''], dtype=float)
                if len(_ipa) == len(d50s):
                    _d50_m  = np.array(d50s, dtype=float) * 1e-9
                    _Np     = _ipa * 1e-12 * _dw * 1e-6 / _eC
                    _Nprob  = _Np * (_d50_m / _dx_m) ** 2
                    _Nsafe  = np.where(_Nprob > 0, _Nprob, np.nan)
                    _ddose  = _snr * _d50_m / (_con * np.sqrt(_Nsafe))
                    _deff   = np.sqrt(_d50_m**2 + _ddose**2) * 1e12
                    _finite = np.where(np.isfinite(_deff), _deff, np.inf)
                    _ei     = int(np.argmin(_finite))
                    if np.isfinite(_deff[_ei]):
                        summ += [
                            ('Min d_eff (pm)',            round(float(_deff[_ei]), 3), 'pm'),
                            ('d50 at min d_eff (pm)',     round(float(_d50_pm[_ei]), 3), 'pm'),
                            ('Delta d_eff - d50 (pm)',    round(float(_deff[_ei] - _d50_pm[_ei]), 3), 'pm'),
                            ('I at min d_eff (pA)',       i_vals[_ei] if i_vals[_ei] != '' else '—', 'pA'),
                            ('M at min d_eff',            round(float(_mg[_ei]), 3), ''),
                        ]
            except Exception:
                pass
            ws_cs = wb.create_sheet(title='Current_Sweep')
            sr = _write_summary(ws_cs, summ)
            _write_table(ws_cs, hdrs, rows, start_row=sr)

        # ── Sheets 6, 7, 8: Map grids ────────────────────────────────────
        if self._map_last_data is not None:
            from openpyxl.utils import get_column_letter
            defoci, alphas, d50_grid, strehl_grid, phys, fast = \
                self._map_last_data
            d50_pm = np.array(d50_grid,    dtype=float) * 1000.0
            sg     = np.array(strehl_grid, dtype=float)
            df_arr = list(defoci)
            al_arr = list(alphas)

            # Compute global summary values for all three grids
            _mi_flat  = int(np.argmin(d50_pm))
            _mi_r, _mi_c = np.unravel_index(_mi_flat, d50_pm.shape)
            _si_flat  = int(np.argmax(sg))
            _si_r, _si_c = np.unravel_index(_si_flat, sg.shape)
            map_summ_d50 = [
                ('Min d50 (pm)',             round(float(np.min(d50_pm)), 3), 'pm'),
                ('\u03b1 at min d50 (mrad)', round(float(al_arr[_mi_r]), 4),  'mrad'),
                ('\u0394f at min d50 (nm)',  round(float(df_arr[_mi_c]), 3),  'nm'),
                ('Max Strehl',               round(float(np.max(sg)),    5),  ''),
                ('\u03b1 at max Strehl (mrad)', round(float(al_arr[_si_r]), 4), 'mrad'),
                ('\u0394f at max Strehl (nm)', round(float(df_arr[_si_c]), 3), 'nm'),
            ]

            grids_to_write = [
                ('Map_d50_pm', d50_pm, 3, map_summ_d50),
                ('Map_Strehl', sg,     5, map_summ_d50),
            ]

            # d_eff map if it was computed
            if self._deff_last_grid is not None:
                dg = np.array(self._deff_last_grid, dtype=float)
                _ei_flat = int(np.argmin(dg))
                _ei_r, _ei_c = np.unravel_index(_ei_flat, dg.shape)
                map_summ_eff = list(map_summ_d50) + [
                    ('Min d_eff (pm)',             round(float(np.min(dg)), 3),  'pm'),
                    ('\u03b1 at min d_eff (mrad)', round(float(al_arr[_ei_r]), 4), 'mrad'),
                    ('\u0394f at min d_eff (nm)',  round(float(df_arr[_ei_c]), 3), 'nm'),
                ]
                grids_to_write += [('Map_deff_pm', dg, 3, map_summ_eff)]

            for ws_title, grid, fmt, summ in grids_to_write:
                ws_m = wb.create_sheet(title=ws_title)
                # Write summary block at top
                data_start = _write_summary(ws_m, summ)
                # Corner label on data header row
                ws_m.cell(row=data_start, column=1,
                          value='\u03b1 \\ \u0394f').font = _ttl_font
                # Column headers = defocus values
                for c_i, df in enumerate(df_arr, 2):
                    cell = ws_m.cell(row=data_start, column=c_i,
                                     value=round(float(df), 3))
                    cell.font = _hdr_font
                    cell.fill = _hdr_fill
                    cell.alignment = Alignment(horizontal='center')
                # Row headers = alpha values; data cells = grid values
                for r_i, (a, row_vals) in enumerate(
                        zip(al_arr, grid), data_start + 1):
                    cell = ws_m.cell(row=r_i, column=1,
                                     value=round(float(a), 4))
                    cell.font = _ttl_font
                    for c_i, val in enumerate(row_vals, 2):
                        ws_m.cell(row=r_i, column=c_i,
                                  value=round(float(val), fmt))
                # Column widths
                ws_m.column_dimensions['A'].width = 14
                for c_i in range(2, len(df_arr) + 2):
                    ws_m.column_dimensions[
                        get_column_letter(c_i)].width = 10

        # ── Save ──────────────────────────────────────────────────────────
        try:
            wb.save(path)
            self.var_status.set(
                f'Exported {len(wb.sheetnames)} sheets \u2192 '
                f'{os.path.basename(path)}')
            messagebox.showinfo(
                'Export complete',
                f'Exported {len(wb.sheetnames)} sheet(s) to:\n{path}')
        except Exception as exc:
            messagebox.showerror('Export failed', str(exc))

    def _show_error(self, msg):
        self.var_status.set(f'Error: {msg}')
        messagebox.showerror('Calculation Error', msg)


# ============================================================
# Entry point
# ============================================================

def main():
    root = tk.Tk()
    app = STEMProbeApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
